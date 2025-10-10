#!/usr/bin/env python3
"""
Script para gerar arquivos Excel estruturados com os dados tratados e organizados
"""

import json
import os
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class GeradorExcelEstruturado:
    def __init__(self):
        self.dados_tratados = {}
        self.dados_organizados = {}
        self.relatorio_tratamento = {}
        self.relatorio_final = {}

    def carregar_dados_tratados(self):
        """Carrega todos os dados tratados e relatórios"""
        print("📂 Carregando dados tratados...")

        # Carregar dados tratados
        arquivo_tratados = "dados_tratados_organizados_20250919_225631.json"
        if os.path.exists(arquivo_tratados):
            with open(arquivo_tratados, "r", encoding="utf-8") as f:
                self.dados_tratados = json.load(f)
            print(f"✅ Dados tratados carregados: {len(self.dados_tratados)} abas")

        # Carregar dados organizados
        arquivo_organizados = "dados_organizados_por_categoria_20250919_225631.json"
        if os.path.exists(arquivo_organizados):
            with open(arquivo_organizados, "r", encoding="utf-8") as f:
                self.dados_organizados = json.load(f)
            print(
                f"✅ Dados organizados carregados: {len(self.dados_organizados)} categorias"
            )

        # Carregar relatório de tratamento
        arquivo_relatorio_tratamento = "relatorio_tratamento_dados_20250919_225631.json"
        if os.path.exists(arquivo_relatorio_tratamento):
            with open(arquivo_relatorio_tratamento, "r", encoding="utf-8") as f:
                self.relatorio_tratamento = json.load(f)
            print(f"✅ Relatório de tratamento carregado")

        # Carregar relatório final
        arquivo_relatorio_final = "relatorio_final_consolidado_20250919_225755.json"
        if os.path.exists(arquivo_relatorio_final):
            with open(arquivo_relatorio_final, "r", encoding="utf-8") as f:
                self.relatorio_final = json.load(f)
            print(f"✅ Relatório final carregado")

        return len(self.dados_tratados) > 0

    def criar_estilo_cabecalho(self):
        """Cria estilo para cabeçalhos"""
        return {
            "font": Font(bold=True, color="FFFFFF"),
            "fill": PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            ),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            ),
        }

    def criar_estilo_dados(self):
        """Cria estilo para dados"""
        return {
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            ),
            "alignment": Alignment(vertical="center"),
        }

    def aplicar_estilo_planilha(self, ws, num_colunas):
        """Aplica estilo a uma planilha"""
        # Estilo do cabeçalho
        header_style = self.criar_estilo_cabecalho()
        data_style = self.criar_estilo_dados()

        # Aplicar estilo ao cabeçalho
        for col in range(1, num_colunas + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_style["font"]
            cell.fill = header_style["fill"]
            cell.alignment = header_style["alignment"]
            cell.border = header_style["border"]

        # Aplicar estilo aos dados
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = data_style["border"]
                cell.alignment = data_style["alignment"]

        # Ajustar largura das colunas
        for col in range(1, num_colunas + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

    def gerar_excel_dados_tratados(self):
        """Gera arquivo Excel com todos os dados tratados"""
        print(f"\n📊 Gerando Excel: Dados Tratados Completos")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        arquivo_excel = f"dados_tratados_completos_{timestamp}.xlsx"

        with pd.ExcelWriter(arquivo_excel, engine="openpyxl") as writer:
            for chave_aba, dados in self.dados_tratados.items():
                if not dados:
                    continue

                print(f"  📋 Processando aba: {chave_aba} ({len(dados)} registros)")

                # Converter para DataFrame
                df = pd.DataFrame(dados)

                # Limitar nome da aba (Excel tem limite de 31 caracteres)
                nome_aba = chave_aba[:31] if len(chave_aba) > 31 else chave_aba

                # Escrever no Excel
                df.to_excel(writer, sheet_name=nome_aba, index=False)

                # Aplicar estilo
                ws = writer.sheets[nome_aba]
                self.aplicar_estilo_planilha(ws, len(df.columns))

        print(f"✅ Excel gerado: {arquivo_excel}")
        return arquivo_excel

    def gerar_excel_dados_organizados(self):
        """Gera arquivo Excel com dados organizados por categoria"""
        print(f"\n📊 Gerando Excel: Dados Organizados por Categoria")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        arquivo_excel = f"dados_organizados_por_categoria_{timestamp}.xlsx"

        with pd.ExcelWriter(arquivo_excel, engine="openpyxl") as writer:
            for categoria, abas in self.dados_organizados.items():
                if not abas:
                    continue

                print(f"  📋 Processando categoria: {categoria}")

                # Combinar todos os dados da categoria
                dados_categoria = []
                for chave_aba, dados in abas.items():
                    for registro in dados:
                        registro_com_categoria = registro.copy()
                        registro_com_categoria["categoria"] = categoria
                        registro_com_categoria["aba_original"] = chave_aba
                        dados_categoria.append(registro_com_categoria)

                if dados_categoria:
                    # Converter para DataFrame
                    df = pd.DataFrame(dados_categoria)

                    # Escrever no Excel
                    df.to_excel(writer, sheet_name=categoria, index=False)

                    # Aplicar estilo
                    ws = writer.sheets[categoria]
                    self.aplicar_estilo_planilha(ws, len(df.columns))

        print(f"✅ Excel gerado: {arquivo_excel}")
        return arquivo_excel

    def gerar_excel_relatorio_tratamento(self):
        """Gera arquivo Excel com relatório de tratamento"""
        print(f"\n📊 Gerando Excel: Relatório de Tratamento")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        arquivo_excel = f"relatorio_tratamento_{timestamp}.xlsx"

        wb = Workbook()

        # Aba 1: Resumo Geral
        ws_resumo = wb.active
        ws_resumo.title = "Resumo Geral"

        resumo_data = [
            ["Métrica", "Valor"],
            ["Timestamp", self.relatorio_tratamento.get("timestamp", "")],
            [
                "Total de Abas Processadas",
                len(self.relatorio_tratamento.get("dados_originais", {})),
            ],
            [
                "Total de Transformações",
                len(self.relatorio_tratamento.get("transformacoes_aplicadas", [])),
            ],
            [
                "Total de Problemas",
                len(self.relatorio_tratamento.get("problemas_encontrados", [])),
            ],
            [
                "Total de Registros Originais",
                sum(len(dados) for dados in self.dados_tratados.values()),
            ],
            [
                "Total de Registros Tratados",
                sum(len(dados) for dados in self.dados_tratados.values()),
            ],
        ]

        for row in resumo_data:
            ws_resumo.append(row)

        self.aplicar_estilo_planilha(ws_resumo, 2)

        # Aba 2: Transformações Aplicadas
        ws_transformacoes = wb.create_sheet("Transformações")

        transformacoes_data = [
            ["Aba", "Registros Originais", "Registros Tratados", "Transformações"]
        ]

        for transformacao in self.relatorio_tratamento.get(
            "transformacoes_aplicadas", []
        ):
            transformacoes_data.append(
                [
                    transformacao.get("aba", ""),
                    transformacao.get("registros_originais", 0),
                    transformacao.get("registros_tratados", 0),
                    ", ".join(transformacao.get("transformacoes", [])),
                ]
            )

        for row in transformacoes_data:
            ws_transformacoes.append(row)

        self.aplicar_estilo_planilha(ws_transformacoes, 4)

        # Aba 3: Problemas Encontrados
        ws_problemas = wb.create_sheet("Problemas")

        problemas_data = [["Aba", "Linha", "Erro"]]

        for problema in self.relatorio_tratamento.get("problemas_encontrados", []):
            problemas_data.append(
                [
                    problema.get("aba", ""),
                    problema.get("linha", ""),
                    problema.get("erro", ""),
                ]
            )

        for row in problemas_data:
            ws_problemas.append(row)

        self.aplicar_estilo_planilha(ws_problemas, 3)

        wb.save(arquivo_excel)
        print(f"✅ Excel gerado: {arquivo_excel}")
        return arquivo_excel

    def gerar_excel_relatorio_final(self):
        """Gera arquivo Excel com relatório final consolidado"""
        print(f"\n📊 Gerando Excel: Relatório Final Consolidado")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        arquivo_excel = f"relatorio_final_consolidado_{timestamp}.xlsx"

        wb = Workbook()

        # Aba 1: Resumo Executivo
        ws_resumo = wb.active
        ws_resumo.title = "Resumo Executivo"

        resumo_executivo = self.relatorio_final.get("resumo_executivo", {})
        resumo_data = [
            ["Métrica", "Valor"],
            [
                "Total de Registros Tratados",
                resumo_executivo.get("total_registros_tratados", 0),
            ],
            ["Total de Abas Tratadas", resumo_executivo.get("total_abas_tratadas", 0)],
            ["Total de Planilhas", resumo_executivo.get("total_planilhas", 0)],
            ["Timestamp Final", resumo_executivo.get("timestamp_final", "")],
        ]

        for row in resumo_data:
            ws_resumo.append(row)

        self.aplicar_estilo_planilha(ws_resumo, 2)

        # Aba 2: Estatísticas por Categoria
        ws_categorias = wb.create_sheet("Estatísticas por Categoria")

        categorias_data = [
            ["Categoria", "Total Registros", "Total Abas", "Percentual do Total"]
        ]

        for categoria, stats in resumo_executivo.get(
            "estatisticas_por_categoria", {}
        ).items():
            categorias_data.append(
                [
                    categoria,
                    stats.get("total_registros", 0),
                    stats.get("total_abas", 0),
                    f"{stats.get('percentual_total', 0):.1f}%",
                ]
            )

        for row in categorias_data:
            ws_categorias.append(row)

        self.aplicar_estilo_planilha(ws_categorias, 4)

        # Aba 3: Análise de Qualidade
        ws_qualidade = wb.create_sheet("Análise de Qualidade")

        qualidade_data = [
            [
                "Categoria",
                "Aba",
                "Total Registros",
                "Total Colunas",
                "Campos Importantes",
            ]
        ]

        for categoria, stats in self.relatorio_final.get(
            "analise_qualidade", {}
        ).items():
            for aba, aba_stats in stats.get("abas_analisadas", {}).items():
                campos_importantes = len(aba_stats.get("completude_campos", {}))
                qualidade_data.append(
                    [
                        categoria,
                        aba,
                        aba_stats.get("total_registros", 0),
                        aba_stats.get("total_colunas", 0),
                        campos_importantes,
                    ]
                )

        for row in qualidade_data:
            ws_qualidade.append(row)

        self.aplicar_estilo_planilha(ws_qualidade, 5)

        # Aba 4: Insights Finais
        ws_insights = wb.create_sheet("Insights Finais")

        insights_data = [["Número", "Insight"]]

        for i, insight in enumerate(self.relatorio_final.get("insights_finais", []), 1):
            insights_data.append([i, insight])

        for row in insights_data:
            ws_insights.append(row)

        self.aplicar_estilo_planilha(ws_insights, 2)

        # Aba 5: Recomendações Finais
        ws_recomendacoes = wb.create_sheet("Recomendações Finais")

        recomendacoes_data = [["Número", "Recomendação"]]

        for i, recomendacao in enumerate(
            self.relatorio_final.get("recomendacoes_finais", []), 1
        ):
            recomendacoes_data.append([i, recomendacao])

        for row in recomendacoes_data:
            ws_recomendacoes.append(row)

        self.aplicar_estilo_planilha(ws_recomendacoes, 2)

        wb.save(arquivo_excel)
        print(f"✅ Excel gerado: {arquivo_excel}")
        return arquivo_excel

    def gerar_excel_consolidado_geral(self):
        """Gera arquivo Excel consolidado com todas as informações"""
        print(f"\n📊 Gerando Excel: Consolidado Geral")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        arquivo_excel = f"consolidado_geral_dados_tratados_{timestamp}.xlsx"

        wb = Workbook()

        # Aba 1: Dashboard Executivo
        ws_dashboard = wb.active
        ws_dashboard.title = "Dashboard Executivo"

        dashboard_data = [
            ["DASHBOARD EXECUTIVO - DADOS TRATADOS E ORGANIZADOS"],
            [""],
            ["MÉTRICAS PRINCIPAIS"],
            [
                "Total de Registros Tratados",
                sum(len(dados) for dados in self.dados_tratados.values()),
            ],
            ["Total de Abas Processadas", len(self.dados_tratados)],
            [
                "Total de Planilhas",
                len(
                    set(
                        dados[0]["planilha_original"]
                        for dados in self.dados_tratados.values()
                        if dados
                    )
                ),
            ],
            ["Total de Categorias", len(self.dados_organizados)],
            [""],
            ["DISTRIBUIÇÃO POR CATEGORIA"],
        ]

        for categoria, abas in self.dados_organizados.items():
            total_registros = sum(len(dados) for dados in abas.values())
            dashboard_data.append([categoria, total_registros])

        for row in dashboard_data:
            ws_dashboard.append(row)

        self.aplicar_estilo_planilha(ws_dashboard, 2)

        # Adicionar todas as abas de dados
        for chave_aba, dados in self.dados_tratados.items():
            if not dados:
                continue

            nome_aba = chave_aba[:31] if len(chave_aba) > 31 else chave_aba
            ws = wb.create_sheet(nome_aba)

            # Converter para DataFrame
            df = pd.DataFrame(dados)

            # Escrever dados
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

            # Aplicar estilo
            self.aplicar_estilo_planilha(ws, len(df.columns))

        wb.save(arquivo_excel)
        print(f"✅ Excel consolidado gerado: {arquivo_excel}")
        return arquivo_excel

    def executar_geracao_excel(self):
        """Executa a geração de todos os arquivos Excel"""
        print("🚀 GERANDO ARQUIVOS EXCEL ESTRUTURADOS")
        print("=" * 80)

        # Carregar dados
        if not self.carregar_dados_tratados():
            print("❌ Falha ao carregar dados tratados")
            return False

        arquivos_gerados = []

        # Gerar arquivos Excel
        try:
            arquivo1 = self.gerar_excel_dados_tratados()
            arquivos_gerados.append(arquivo1)
        except Exception as e:
            print(f"❌ Erro ao gerar Excel dados tratados: {e}")

        try:
            arquivo2 = self.gerar_excel_dados_organizados()
            arquivos_gerados.append(arquivo2)
        except Exception as e:
            print(f"❌ Erro ao gerar Excel dados organizados: {e}")

        try:
            arquivo3 = self.gerar_excel_relatorio_tratamento()
            arquivos_gerados.append(arquivo3)
        except Exception as e:
            print(f"❌ Erro ao gerar Excel relatório tratamento: {e}")

        try:
            arquivo4 = self.gerar_excel_relatorio_final()
            arquivos_gerados.append(arquivo4)
        except Exception as e:
            print(f"❌ Erro ao gerar Excel relatório final: {e}")

        try:
            arquivo5 = self.gerar_excel_consolidado_geral()
            arquivos_gerados.append(arquivo5)
        except Exception as e:
            print(f"❌ Erro ao gerar Excel consolidado: {e}")

        print("\n" + "=" * 80)
        print("✅ GERAÇÃO DE ARQUIVOS EXCEL CONCLUÍDA!")
        print(f"📄 Arquivos gerados: {len(arquivos_gerados)}")

        for arquivo in arquivos_gerados:
            print(f"  📊 {arquivo}")

        return len(arquivos_gerados) > 0


if __name__ == "__main__":
    gerador = GeradorExcelEstruturado()
    sucesso = gerador.executar_geracao_excel()

    if sucesso:
        print("\n🎉 Arquivos Excel estruturados gerados com sucesso!")
        print(
            "📋 Dados tratados organizados em planilhas Excel com formatação profissional"
        )
    else:
        print("\n❌ Falha na geração dos arquivos Excel")

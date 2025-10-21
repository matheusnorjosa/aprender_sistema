"""
ETL de Deslocamentos: Upsert idempotente de registros de deslocamento.

Lê CSV/XLSX e cria/atualiza Deslocamento por external_hash único.

Regras de negócio:
- Usuário resolvido por email > nome
- Datas parseadas: ISO, dd/mm/yyyy, Excel serial
- Idempotência: SHA1(usuario_id|origem|destino|start|end)
- Relatório em out_etl/etl_deslocamento_report.json
"""

import csv
import hashlib
import json
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional

from django.conf import settings
from django.core.management.base import BaseCommand
from django.db import transaction

from apps.core.models import Deslocamento
from apps.dat_ingest.services.resolvers import (
    resolve_user_by_email,
    resolve_user_by_name,
)


class Command(BaseCommand):
    help = "ETL de Deslocamentos: upsert idempotente de Deslocamento"

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            type=str,
            required=True,
            help="Caminho do arquivo CSV/XLSX",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Simula execução sem writes no banco",
        )

    def handle(self, *args, **options):
        self.dry_run = options["dry_run"]
        self.file_path = options["file"]

        self.stdout.write(f"🚀 ETL Deslocamento (dry_run={self.dry_run})")
        self.stdout.write(f"   File: {self.file_path}")

        # Stats e pendências
        self.stats = {
            "created": 0,
            "updated": 0,
            "unchanged": 0,
            "skipped": {"usuario": 0, "dates": 0, "other": 0},
        }
        self.pendencias = {
            "usuarios": [],
            "dates": [],
            "outros": [],
        }

        # Carregar arquivo
        self.stdout.write("\n📂 Carregando arquivo...")
        rows = self.load_file()
        self.stdout.write(f"   {len(rows)} linhas carregadas")

        # Processar linhas
        self.stdout.write("\n⚙️  Processando deslocamentos...")
        with transaction.atomic():
            for idx, row in enumerate(rows, start=1):
                try:
                    self.process_row(row, idx)
                except Exception as e:
                    self.stdout.write(f"   ❌ Erro linha {idx}: {e}")
                    self.stats["skipped"]["other"] += 1
                    self.pendencias["outros"].append({
                        "linha": idx,
                        "erro": str(e),
                        "row": row,
                    })

            if self.dry_run:
                self.stdout.write("   [DRY-RUN] Rollback transaction")
                transaction.set_rollback(True)

        # Relatório
        self.generate_report()

        self.stdout.write(self.style.SUCCESS("\n✅ ETL concluído!"))

    def load_file(self) -> List[Dict[str, Any]]:
        """
        Carrega CSV ou XLSX em memória.

        Headers flexíveis (case-insensitive):
        - email, usuario, user → email do usuário
        - start, start_date, inicio, data_inicio → start_date
        - end, end_date, fim, data_fim → end_date
        - origem, origin → origem
        - destino, destination, dest → destino
        - obs, observacao, observacoes, notes → observacao
        """
        file_path = Path(self.file_path)

        if file_path.suffix.lower() == '.csv':
            return self._load_csv(file_path)
        elif file_path.suffix.lower() in ['.xlsx', '.xls']:
            return self._load_xlsx(file_path)
        else:
            raise ValueError(f"Formato não suportado: {file_path.suffix}")

    def _load_csv(self, file_path: Path) -> List[Dict[str, Any]]:
        """Carrega CSV com headers flexíveis."""
        rows = []
        with open(file_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                normalized = self._normalize_headers(row)
                rows.append(normalized)
        return rows

    def _load_xlsx(self, file_path: Path) -> List[Dict[str, Any]]:
        """Carrega XLSX com headers flexíveis."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("openpyxl não instalado. Execute: pip install openpyxl")

        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        # Ler headers (primeira linha)
        headers_raw = [cell.value for cell in ws[1]]
        headers = [str(h).strip().lower() if h else f"col_{i}" for i, h in enumerate(headers_raw)]

        # Ler linhas
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            normalized = self._normalize_headers(row_dict)
            rows.append(normalized)

        return rows

    def _normalize_headers(self, row: Dict[str, Any]) -> Dict[str, Any]:
        """
        Normaliza headers do CSV/XLSX para padrão interno.

        Mapeamento flexível com heurística para usuario/user:
        - email/e-mail/mail → email (sempre)
        - usuario/user: se contiver '@' → email; senão → name
        - name/nome/display_name → name
        - start/start_date/inicio/data_inicio → start
        - end/end_date/fim/data_fim → end
        - origem/origin → origem
        - destino/destination/dest → destino
        - obs/observacao/observacoes/notes → obs

        Heurística: Colunas "usuario"/"user" são tratadas como email se o valor
        contiver "@"; caso contrário, são tratadas como name.
        """
        normalized = {}

        # Criar mapa lower → original key
        lower_map = {k.strip().lower(): k for k in row.keys()}

        # Email (apenas chaves claramente de email)
        for key in ["email", "e-mail", "mail"]:
            if key in lower_map:
                normalized["email"] = str(row[lower_map[key]]).strip() if row[lower_map[key]] else ""
                break
        else:
            normalized["email"] = ""

        # Name (chaves de nome/usuário)
        for key in ["name", "nome", "display_name", "usuario_nome"]:
            if key in lower_map:
                normalized["name"] = str(row[lower_map[key]]).strip() if row[lower_map[key]] else ""
                break
        else:
            normalized["name"] = ""

        # Heurística para "usuario"/"user": se tem "@" → email, senão → name
        for key in ["usuario", "user", "usuário"]:
            if key in lower_map:
                value = str(row[lower_map[key]]).strip() if row[lower_map[key]] else ""
                if value:
                    if "@" in value:
                        # Contém "@" → é email
                        if not normalized["email"]:  # Só sobrescreve se email ainda vazio
                            normalized["email"] = value
                    else:
                        # Não contém "@" → é nome
                        if not normalized["name"]:  # Só sobrescreve se name ainda vazio
                            normalized["name"] = value
                break

        # Start date
        for key in ["start", "start_date", "inicio", "data_inicio", "início"]:
            if key in lower_map:
                normalized["start"] = row[lower_map[key]]
                break
        else:
            normalized["start"] = None

        # End date
        for key in ["end", "end_date", "fim", "data_fim"]:
            if key in lower_map:
                normalized["end"] = row[lower_map[key]]
                break
        else:
            normalized["end"] = None

        # Origem
        for key in ["origem", "origin", "de", "from"]:
            if key in lower_map:
                normalized["origem"] = str(row[lower_map[key]]).strip() if row[lower_map[key]] else ""
                break
        else:
            normalized["origem"] = ""

        # Destino
        for key in ["destino", "destination", "dest", "para", "to"]:
            if key in lower_map:
                normalized["destino"] = str(row[lower_map[key]]).strip() if row[lower_map[key]] else ""
                break
        else:
            normalized["destino"] = ""

        # Observação
        for key in ["obs", "observacao", "observacoes", "observações", "notes", "nota"]:
            if key in lower_map:
                normalized["obs"] = str(row[lower_map[key]]).strip() if row[lower_map[key]] else ""
                break
        else:
            normalized["obs"] = ""

        return normalized

    def parse_date_flexible(self, value: Any) -> Optional[date]:
        """
        Parse data flexível: ISO, dd/mm/yyyy, Excel serial.

        Returns:
            date object ou None se parsing falhar
        """
        if not value:
            return None

        # Se já é date
        if isinstance(value, date):
            return value

        # Se é datetime
        if isinstance(value, datetime):
            return value.date()

        # Se é número (Excel serial)
        if isinstance(value, (int, float)):
            try:
                # Excel serial: 1899-12-30 + dias
                return (datetime(1899, 12, 30) + timedelta(days=int(value))).date()
            except Exception:
                return None

        # Se é string
        value_str = str(value).strip()
        if not value_str:
            return None

        # Tentar ISO (YYYY-MM-DD)
        try:
            return datetime.strptime(value_str, "%Y-%m-%d").date()
        except ValueError:
            pass

        # Tentar dd/mm/yyyy
        try:
            return datetime.strptime(value_str, "%d/%m/%Y").date()
        except ValueError:
            pass

        # Tentar dd/mm/yy
        try:
            return datetime.strptime(value_str, "%d/%m/%y").date()
        except ValueError:
            pass

        # Tentar yyyy-mm-dd (com hora)
        try:
            return datetime.fromisoformat(value_str).date()
        except ValueError:
            pass

        return None

    def process_row(self, row: Dict[str, Any], linha_num: int):
        """
        Processa uma linha do arquivo:
        1. Resolve usuário (email > nome)
        2. Parse datas
        3. Gera external_hash
        4. Upsert Deslocamento
        """
        # Resolver usuário
        email = row.get("email", "").strip()
        name = row.get("name", "").strip()

        usuario = None
        if email:
            usuario = resolve_user_by_email(email)

        if not usuario and name:
            usuario = resolve_user_by_name(name)

        if not usuario:
            self.stdout.write(f"   ⚠️  Linha {linha_num}: Usuário não resolvido (email={email}, name={name})")
            self.stats["skipped"]["usuario"] += 1
            self.pendencias["usuarios"].append({
                "linha": linha_num,
                "email": email,
                "name": name,
            })
            return

        # Parse datas
        start_date = self.parse_date_flexible(row.get("start"))
        end_date = self.parse_date_flexible(row.get("end"))

        if not start_date or not end_date:
            self.stdout.write(f"   ⚠️  Linha {linha_num}: Datas inválidas (start={row.get('start')}, end={row.get('end')})")
            self.stats["skipped"]["dates"] += 1
            self.pendencias["dates"].append({
                "linha": linha_num,
                "start": str(row.get("start")),
                "end": str(row.get("end")),
            })
            return

        if end_date < start_date:
            self.stdout.write(f"   ⚠️  Linha {linha_num}: end_date < start_date")
            self.stats["skipped"]["dates"] += 1
            self.pendencias["dates"].append({
                "linha": linha_num,
                "motivo": "end_date < start_date",
                "start": str(start_date),
                "end": str(end_date),
            })
            return

        # Origem/Destino
        origem = row.get("origem", "").strip()
        destino = row.get("destino", "").strip()

        if not origem or not destino:
            self.stdout.write(f"   ⚠️  Linha {linha_num}: Origem/Destino ausentes")
            self.stats["skipped"]["other"] += 1
            self.pendencias["outros"].append({
                "linha": linha_num,
                "motivo": "origem_destino_ausentes",
                "origem": origem,
                "destino": destino,
            })
            return

        # Observação
        observacao = row.get("obs", "").strip()

        # Gerar external_hash
        external_hash = self.compute_external_hash(
            usuario.id, origem, destino, start_date, end_date
        )

        # Upsert Deslocamento
        if not self.dry_run:
            deslocamento, created = Deslocamento.objects.update_or_create(
                external_hash=external_hash,
                defaults={
                    "usuario": usuario,
                    "origem": origem,
                    "destino": destino,
                    "start_date": start_date,
                    "end_date": end_date,
                    "observacao": observacao if observacao else None,
                },
            )

            if created:
                self.stats["created"] += 1
                self.stdout.write(f"   ✅ Linha {linha_num}: Deslocamento criado ({external_hash[:8]}...)")
            else:
                # Verificar se houve mudança
                changed = False
                if deslocamento.origem != origem or deslocamento.destino != destino:
                    changed = True
                if deslocamento.start_date != start_date or deslocamento.end_date != end_date:
                    changed = True
                if deslocamento.observacao != (observacao if observacao else None):
                    changed = True

                if changed:
                    self.stats["updated"] += 1
                    self.stdout.write(f"   🔄 Linha {linha_num}: Deslocamento atualizado ({external_hash[:8]}...)")
                else:
                    self.stats["unchanged"] += 1
        else:
            # Dry-run: apenas contadores
            self.stats["created"] += 1
            self.stdout.write(f"   [DRY-RUN] Linha {linha_num}: {external_hash[:8]}... ({usuario.username}: {origem}→{destino})")

    def compute_external_hash(
        self, usuario_id: int, origem: str, destino: str, start: date, end: date
    ) -> str:
        """
        Gera external_hash SHA1 a partir de campos-chave.

        Formato: SHA1(usuario_id|origem|destino|start|end)
        """
        parts = [
            str(usuario_id),
            origem,
            destino,
            start.isoformat(),
            end.isoformat(),
        ]
        content = "|".join(parts)
        return hashlib.sha1(content.encode("utf-8")).hexdigest()

    def generate_report(self):
        """Gera relatório JSON de pendências."""
        out_dir = Path(settings.BASE_DIR) / "out_etl"
        out_dir.mkdir(exist_ok=True)

        report_path = out_dir / "etl_deslocamento_report.json"

        report = {
            "stats": self.stats,
            "pendencias": self.pendencias,
        }

        with open(report_path, "w", encoding="utf-8") as f:
            json.dump(report, f, indent=2, ensure_ascii=False)

        self.stdout.write(f"\n📄 Relatório salvo em: {report_path}")
        self.stdout.write(f"\n📊 Stats:")
        self.stdout.write(f"   Created: {self.stats['created']}")
        self.stdout.write(f"   Updated: {self.stats['updated']}")
        self.stdout.write(f"   Unchanged: {self.stats['unchanged']}")
        self.stdout.write(f"   Skipped: {self.stats['skipped']}")

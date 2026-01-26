"""
Serviço de importação de Compras a partir de CSV/XLSX.

Importa Compras da aba "🟥 COMPRAS" da Planilha de Controle.
- Idempotência por external_hash determinístico
- Dry-run mode para preview
- Relatório JSON em out_etl/import_compras_report.json
"""

# pyright: reportUnknownVariableType=false, reportUnknownMemberType=false, reportUnknownParameterType=false, reportUnknownArgumentType=false, reportMissingParameterType=false, reportOptionalMemberAccess=false, reportCallIssue=false, reportOptionalSubscript=false, reportArgumentType=false, reportMissingTypeStubs=false, reportAttributeAccessIssue=false, reportReturnType=false, reportGeneralTypeIssues=false

from __future__ import annotations

import csv
import hashlib
import json
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from django.conf import settings
from django.db import transaction
from django.utils import timezone

from apps.core.models import Compra, Municipio, Projeto
from apps.core.services.normalize import norm_text
from apps.core.services.resolvers import resolve_municipio, resolve_projeto
from apps.core.types import ExternalHash

OUT_DIR: Path = Path(settings.BASE_DIR) / "out_etl"
OUT_DIR.mkdir(parents=True, exist_ok=True)


def sha1_str(s: str) -> str:
    """Gera SHA1 hex digest de uma string."""
    return hashlib.sha1(s.encode("utf-8")).hexdigest()


def parse_date_flexible(v: str | None) -> date | None:
    """
    Parse de data flexível: YYYY-MM-DD, DD/MM/YYYY ou serial Excel.

    Excel armazena datas como número de dias desde 1899-12-30.
    """
    if not v:
        return None

    s = str(v).strip()

    # Formato ISO: YYYY-MM-DD
    try:
        if "-" in s:
            return datetime.fromisoformat(s).date()
    except Exception:
        pass

    # Formato brasileiro: DD/MM/YYYY
    try:
        if "/" in s:
            parts = s.split("/")
            if len(parts) == 3:
                d, m, a = parts
                return date(int(a), int(m), int(d))
    except Exception:
        pass

    # Serial Excel (número)
    try:
        n = float(s)
        base = datetime(1899, 12, 30)
        return (base + timedelta(days=n)).date()
    except Exception:
        return None

    return None


@dataclass
class ImportStats:
    """Estatísticas de importação."""

    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: int = 0


@dataclass
class ImportPendencias:
    """Pendências encontradas durante importação."""

    municipios: list[dict[str, Any]] = field(default_factory=list)
    projetos: list[dict[str, Any]] = field(default_factory=list)
    linhas_invalidas: list[dict[str, Any]] = field(default_factory=list)


def import_compras_from_file(
    *, path: str, dry_run: bool = True, auto_create_municipios: bool = False
) -> dict[str, Any]:
    """
    Importa Compras de CSV/XLSX.

    Args:
        path: Caminho para arquivo CSV/XLSX
        dry_run: Se True, não persiste (apenas simula)
        auto_create_municipios: Se True, cria municípios que não existem

    Returns:
        Relatório com stats, pendências e IDs criados

    Raises:
        FileNotFoundError: Se arquivo não existe
    """
    p: Path = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {path}")

    # Ler CSV ou XLSX
    rows: list[dict[str, Any]] = []
    if p.suffix.lower() == ".csv":
        with p.open("r", encoding="utf-8-sig", newline="") as f:
            rows = list(csv.DictReader(f))
    else:
        from openpyxl import load_workbook

        wb = load_workbook(p, data_only=True, read_only=True)

        # Encontrar aba COMPRAS (pode ter emoji ou variações)
        ws = None
        for sheet_name in wb.sheetnames:
            if "COMPRAS" in sheet_name.upper():
                ws = wb[sheet_name]
                break

        if ws is None:
            # Fallback para aba ativa se não encontrar COMPRAS
            ws = wb.active

        headers: list[str] = [
            str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))
        ]
        for row in ws.iter_rows(min_row=2):
            rec: dict[str, Any] = {}
            for j, cell in enumerate(row):
                key: str = headers[j] if j < len(headers) else f"col{j}"
                rec[key] = cell.value
            rows.append(rec)
        wb.close()

    stats: ImportStats = ImportStats()
    pendencias: ImportPendencias = ImportPendencias()

    def normalize_row(r: dict[str, Any]) -> dict[str, Any]:
        """Normaliza uma linha do arquivo."""
        codigo: str = str(r.get("CÓD") or r.get("COD") or r.get("Cód") or r.get("id") or r.get("ID") or "").strip()
        produto: str = str(r.get("Produto") or "").strip()
        produto_norm: str = norm_text(produto)
        quant_raw: Any = r.get("Quant.") or r.get("Quant") or r.get("Quantidade")
        municipio: str = str(r.get("Município") or r.get("Municipio") or "").strip()
        uf: str = str(r.get("UF") or "").strip().upper()
        data_raw: Any = r.get("Data")
        uso: str = str(r.get("Uso das coleções") or r.get("Uso") or "").strip()

        quantidade: int | None
        try:
            quantidade = int(float(quant_raw or 0))
        except Exception:
            quantidade = None

        data: date | None = parse_date_flexible(str(data_raw) if data_raw is not None else "")

        return {
            "codigo": codigo,
            "produto": produto,
            "produto_norm": produto_norm,
            "quantidade": quantidade,
            "municipio": municipio,
            "uf": uf[:2] if uf else "",
            "data": data,
            "uso_norm": norm_text(uso),
        }

    normalized: list[dict[str, Any]] = [normalize_row(r) for r in rows]

    # Resolver municípios e projetos
    def _build_ext_key(r: dict[str, Any]) -> str:
        """Constrói chave para external_hash (mesma lógica em dry-run e apply)."""
        return "|".join(
            [
                str(r["municipio_obj"].id),
                str(r["projeto_obj"].id),
                r["codigo"],
                r["produto_norm"],
                str(r["quantidade"]),
                str(r["data"]),
                r["uso_norm"],
            ]
        )

    resolved: list[dict[str, Any]] = []
    for i, r in enumerate(normalized, start=2):
        # Validar campos obrigatórios
        if not r["municipio"] or r["quantidade"] is None or not r["codigo"]:
            stats.skipped += 1
            pendencias.linhas_invalidas.append(
                {
                    "row": i,
                    "motivo": "municipio/quantidade/codigo inválidos",
                    "dados": {
                        "municipio": r["municipio"],
                        "quantidade": r["quantidade"],
                        "codigo": r["codigo"],
                    },
                }
            )
            continue

        # Resolver município
        mun_obj: Municipio | None = resolve_municipio(r["municipio"])
        if not mun_obj:
            stats.skipped += 1
            pendencias.municipios.append({"row": i, "municipio": r["municipio"], "uf": r["uf"]})
            continue

        # Resolver projeto (obrigatório)
        proj_obj: Projeto | None = _infer_projeto_from_produto(r["produto_norm"])
        if not proj_obj:
            stats.skipped += 1
            pendencias.projetos.append({"row": i, "produto": r["produto"], "produto_norm": r["produto_norm"]})
            continue

        resolved.append({**r, "municipio_obj": mun_obj, "projeto_obj": proj_obj})

    # Persistir ou simular
    created_ids: list[int] = []
    if not dry_run:
        with transaction.atomic():
            for r in resolved:
                # Gerar external_hash determinístico (mesma lógica do dry-run)
                ext_key: str = _build_ext_key(r)
                ext_hash: ExternalHash = sha1_str(ext_key)

                defaults: dict[str, Any] = dict(
                    municipio=r["municipio_obj"],
                    projeto=r["projeto_obj"],
                    codigo=r["codigo"],
                    quantidade=r["quantidade"],
                    data=r["data"],
                    uso=r["uso_norm"],
                )

                obj: Compra
                created: bool
                obj, created = Compra.objects.update_or_create(external_hash=ext_hash, defaults=defaults)

                if created:
                    stats.created += 1
                    created_ids.append(obj.id)
                else:
                    # Verificar se houve mudança
                    changed: bool = any(getattr(obj, k) != v for k, v in defaults.items())
                    if changed:
                        stats.updated += 1
                    else:
                        stats.skipped += 1
    else:
        # Simulação: verificar se já existe sem persistir (mesma lógica do apply)
        for r in resolved:
            ext_key: str = _build_ext_key(r)
            ext_hash: ExternalHash = sha1_str(ext_key)
            exists: bool = Compra.objects.filter(external_hash=ext_hash).exists()
            if exists:
                stats.updated += 1
            else:
                stats.created += 1

    # Gerar relatório
    report: dict[str, Any] = {
        "path": str(p),
        "dry_run": dry_run,
        "stats": vars(stats),
        "pendencias": {
            "municipios": pendencias.municipios,
            "projetos": pendencias.projetos,
            "linhas_invalidas": pendencias.linhas_invalidas,
        },
        "created_ids": created_ids if not dry_run else [],
        "ts": timezone.now().isoformat(),
    }

    # Salvar relatório em arquivo
    report_path: Path = OUT_DIR / "import_compras_report.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    return report


def _infer_projeto_from_produto(produto_norm: str) -> Projeto | None:
    """
    Infere projeto a partir do nome do produto normalizado.

    Mapeamento completo de produtos → projetos baseado na Planilha de Controle.
    A ordem das verificações é importante (mais específico primeiro).
    """
    if not produto_norm:
        return None

    key: str = produto_norm.upper()

    # Mapeamento de padrões → nomes de projeto no banco
    # Ordem: mais específico primeiro
    mappings: list[tuple[list[str], str]] = [
        # Projetos com nomes compostos (verificar primeiro)
        (["KIT COMBO NOVO LENDO E AMMA"], "Novo Lendo"),
        (["VIDA E CIENCIA", "VIDA E CIÊNCIA"], "Vida & Ciências"),
        (["VIDA E LINGUAGEM"], "Vida & Linguagem"),
        (["VIDA E MATEMATICA", "VIDA E MATEMÁTICA"], "Vida & Matemática"),
        (["LER OUVIR E CONTAR"], "LER OUVIR E CONTAR"),
        (["LENDO E ESCREVENDO"], "LER OUVIR E CONTAR"),  # Mesma família
        (["ESCREVER COMUNICAR E SER"], "LER OUVIR E CONTAR"),  # Mesma família
        (["A COR DA GENTE"], "A COR DA GENTE"),
        (["BRINCANDO E APRENDENDO"], "Brincando e Aprendendo"),
        (["SOU DA PAZ"], "SOU DA PAZ"),
        (["UNI DUNI T"], "UNI DUNI TÊ"),  # Sem acento para match flexível
        (["EDUCACAO FINANCEIRA", "EDUCAÇÃO FINANCEIRA"], "ED FINANCEIRA"),
        (["AVANCANDO JUNTOS", "AVANÇANDO JUNTOS"], "Avançando Juntos Matemática"),
        (["APRENDER MAIS"], "Projeto AMMA"),
        (["SUPER ATIVAR", "SUPERATIVAR"], "Superativar"),
        (
            ["GESTAO ESCOLAR", "GESTÃO ESCOLAR", "FORTALECIMENTO DA GESTAO", "FORTALECIMENTO DA GESTÃO"],
            "GESTÃO ESCOLAR",
        ),
        # Projetos com nomes simples
        (["NOVO LENDO"], "Novo Lendo"),
        (["ACERTA"], "ACerta"),
        (["FLUIR"], "Fluir"),  # Pode não existir no banco ainda
        (["CATAVENTOS"], "Cataventos"),
        (["MIUDEZAS"], "Miudezas"),
        (["AVALIAR"], "ACerta"),  # Avaliar é do projeto ACerta
        (["TEMA"], "Superativar"),  # TEMA é do Superativar
    ]

    for patterns, projeto_nome in mappings:
        for pattern in patterns:
            if pattern in key:
                try:
                    projeto = resolve_projeto(projeto_nome)
                    if projeto:
                        return projeto
                except Exception:
                    pass

    return None

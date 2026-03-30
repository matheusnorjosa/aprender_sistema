"""
Parser para aba Bloqueios da planilha Disponibilidade
"""

# pyright: reportUnknownVariableType=false, reportUnknownMemberType=false, reportUnknownParameterType=false, reportUnknownArgumentType=false, reportMissingParameterType=false, reportOptionalMemberAccess=false, reportCallIssue=false, reportOptionalSubscript=false, reportArgumentType=false, reportMissingTypeStubs=false, reportAttributeAccessIssue=false, reportReturnType=false, reportGeneralTypeIssues=false


from __future__ import annotations

from datetime import datetime, time, timedelta
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from .normalizers import normalize_str


def parse_bloqueios(filepath: Path) -> list[dict[str, Any]]:
    """
    Parse aba Bloqueios

    Estrutura REAL da planilha (corrigida 2025-11-19):
    A (Col 1): Usuário (nome do formador)
    B (Col 2): Inicio (datetime completo)
    C (Col 3): Fim (datetime completo)
    D (Col 4): Tipo ("Total" ou "Parcial")

    NOTA: Estrutura antiga (INCORRETA) tinha 7 colunas separando data e hora.
    A planilha real tem apenas 4 colunas com datetimes completos.

    Args:
        filepath: Caminho do arquivo Excel

    Returns:
        Lista de dicts com bloqueios
    """
    wb = load_workbook(filepath, data_only=True)

    if "Bloqueios" not in wb.sheetnames:
        return []

    ws = wb["Bloqueios"]
    bloqueios = []
    tz = ZoneInfo("America/Fortaleza")

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) < 4:
            continue

        # Estrutura correta: Usuário | Inicio | Fim | Tipo
        formador_nome = normalize_str(row[0]) if row[0] else ""
        inicio_raw = row[1]  # datetime completo
        fim_raw = row[2]  # datetime completo
        tipo_raw = normalize_str(row[3]).upper() if row[3] else ""

        # Validar campos obrigatórios
        if not formador_nome or not inicio_raw or not fim_raw or not tipo_raw:
            continue

        # Validar e normalizar tipo
        if tipo_raw not in ["T", "P", "TOTAL", "PARCIAL"]:
            continue

        tipo = "T" if tipo_raw in ["TOTAL", "T"] else "P"

        # Processar datetimes
        try:
            # Se já é datetime, usar direto; se é date, converter
            if isinstance(inicio_raw, datetime):
                inicio = inicio_raw
            else:
                inicio = datetime.combine(inicio_raw, time(0, 0))

            if isinstance(fim_raw, datetime):
                fim = fim_raw
            else:
                fim = datetime.combine(fim_raw, time(23, 59, 59))

            # Localizar para America/Fortaleza se naive
            if inicio.tzinfo is None:
                inicio = inicio.replace(tzinfo=tz)
            if fim.tzinfo is None:
                fim = fim.replace(tzinfo=tz)

            # Garantir que fim > inicio (constraint do banco)
            # Alguns bloqueios vêm com inicio == fim (00:00:00)
            if fim <= inicio:
                # Se fim veio como 00:00, mudar para 23:59:59 do mesmo dia
                if isinstance(fim_raw, datetime) and fim_raw.time() == time(0, 0):
                    fim = datetime.combine(fim_raw.date(), time(23, 59, 59), tzinfo=tz)
                else:
                    # Caso geral: adicionar 1 segundo
                    fim = fim + timedelta(seconds=1)

        except Exception:
            # Erro ao processar datas, pular linha
            continue

        bloqueios.append(
            {
                "formador_nome": formador_nome,
                "tipo": tipo,
                "inicio": inicio,
                "fim": fim,
                "motivo": "",  # Planilha não tem coluna de motivo
                "src": f"{filepath.name}/Bloqueios",
                "rownum": i,
            }
        )

    return bloqueios

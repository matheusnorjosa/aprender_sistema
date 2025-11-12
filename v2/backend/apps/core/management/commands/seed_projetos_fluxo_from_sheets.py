"""
Management command para classificar Projetos por fluxo a partir de planilhas XLSX.

PR 13/N - Classificação de Projetos por Planilhas Oficiais

Lê planilhas XLSX oficiais (Controle e Agenda) e atualiza campo 'fluxo'.

Estrutura esperada:
- Planilha de Controle, aba "Projetos": Codigo, Nome, Setor
  - Setor = "Superintendência" → SUPER
  - Outros → NAO_SUPER
- Acompanhamento de Agenda:
  - Aba "Super" → SUPER
  - Outras abas → NAO_SUPER

Matching:
- Prioridade 1: código (case-insensitive)
- Prioridade 2: nome normalizado (NFKD, casefold)

Idempotente: Só atualiza se fluxo mudou.
Gera CSV com fonte/detalhe rastreável.
"""
from __future__ import annotations

import csv
import re
import unicodedata
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandParser
from django.db.models import Q

from apps.core.models import Projeto

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


class Command(BaseCommand):
    help = "Seed Projeto.fluxo a partir de planilhas XLSX (idempotente)"

    def add_arguments(self, parser: CommandParser) -> None:
        parser.add_argument(
            '--xls-controle',
            type=str,
            default='data/csv-import/Planilha de Controle - 2025.xlsx',
            help='Caminho da Planilha de Controle'
        )
        parser.add_argument(
            '--xls-agenda',
            type=str,
            default='data/csv-import/Acompanhamento de Agenda _ 2025.xlsx',
            help='Caminho da planilha Acompanhamento de Agenda'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Simula execução sem persistir mudanças'
        )
        parser.add_argument(
            '--verbose',
            action='store_true',
            help='Log detalhado de cada operação'
        )

    def _normalize_string(self, s: str | None) -> str:
        """
        Normaliza string para matching robusto.

        Aplica:
        - NFKD (decomposição Unicode)
        - Remove acentos
        - Lowercase (casefold)
        - Remove caracteres não-alfanuméricos (exceto espaços)
        - Colapsa espaços múltiplos
        """
        if not s:
            return ''
        # NFKD + remove acentos
        s = unicodedata.normalize('NFKD', s)
        s = s.encode('ascii', 'ignore').decode('ascii')
        # casefold
        s = s.casefold()
        # Remove pontuação, mantém espaços
        s = re.sub(r'[^\w\s]', '', s)
        # Colapsa espaços
        s = re.sub(r'\s+', ' ', s).strip()
        return s

    def _find_projeto(self, codigo_raw: str | None, nome_raw: str | None) -> Projeto | None:
        """
        Encontra projeto por código ou nome normalizado.

        Prioridade:
        1. Código (case-insensitive)
        2. Nome normalizado (NFKD, casefold)
        """
        # Prioridade 1: código
        if codigo_raw:
            projeto = Projeto.objects.filter(
                Q(codigo__iexact=codigo_raw.strip())
            ).first()
            if projeto:
                return projeto

        # Prioridade 2: nome normalizado
        if nome_raw:
            nome_norm = self._normalize_string(nome_raw)
            for p in Projeto.objects.all():
                if self._normalize_string(p.nome) == nome_norm:
                    return p

        return None

    def _extract_projetos_from_controle(self, xls_path: Path, verbose: bool = False) -> dict[str, dict[str, Any]]:
        """
        Extrai projetos da Planilha de Controle.

        Estrutura: Codigo, Nome, Setor
        Setor = "Superintendência" → SUPER
        Outros → NAO_SUPER
        """
        if load_workbook is None:
            self.stdout.write(self.style.ERROR("❌ openpyxl não instalado"))
            return {}

        wb = load_workbook(xls_path, read_only=True, data_only=True)
        projetos_map = {}

        # Procurar aba "Projetos"
        if "Projetos" not in wb.sheetnames:
            if verbose:
                self.stdout.write(
                    self.style.WARNING(f"  ⚠️ Aba 'Projetos' não encontrada em {xls_path}")
                )
            return projetos_map

        ws = wb["Projetos"]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            return projetos_map

        # Primeira linha = header
        header = rows[0]
        try:
            idx_codigo = header.index("Codigo")
            idx_nome = header.index("Nome")
            idx_setor = header.index("Setor")
        except ValueError:
            if verbose:
                self.stdout.write(
                    self.style.WARNING("  ⚠️ Colunas esperadas não encontradas (Codigo, Nome, Setor)")
                )
            return projetos_map

        # Processar linhas
        for i, row in enumerate(rows[1:], start=2):
            if len(row) <= max(idx_codigo, idx_nome, idx_setor):
                continue

            codigo = str(row[idx_codigo] or '').strip()
            nome = str(row[idx_nome] or '').strip()
            setor = str(row[idx_setor] or '').strip()

            if not nome:
                continue

            # Determinar fluxo
            fluxo = 'SUPER' if 'superintend' in setor.lower() else 'NAO_SUPER'

            # Key para matching
            key = self._normalize_string(codigo if codigo else nome)
            projetos_map[key] = {
                'codigo': codigo,
                'nome': nome,
                'fluxo': fluxo,
                'fonte': f'Planilha de Controle, aba \'Projetos\', linha {i}',
                'detalhe': f'Coluna Setor = \'{setor}\''
            }

        return projetos_map

    def _extract_projetos_from_agenda(self, xls_path: Path, verbose: bool = False) -> dict[str, dict[str, Any]]:
        """
        Extrai projetos do Acompanhamento de Agenda.

        Estrutura:
        - Aba "Super" → SUPER
        - Outras abas → NAO_SUPER
        """
        if load_workbook is None:
            return {}

        wb = load_workbook(xls_path, read_only=True, data_only=True)
        projetos_map = {}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            if not rows:
                continue

            # Determinar fluxo pela aba
            fluxo = 'SUPER' if sheet_name.lower() == 'super' else 'NAO_SUPER'

            # Primeira linha = header (assumir coluna "Projeto" ou similar)
            header = rows[0]
            idx_projeto = None
            for idx, col in enumerate(header):
                if col and 'projeto' in str(col).lower():
                    idx_projeto = idx
                    break

            if idx_projeto is None:
                continue

            # Processar linhas
            for i, row in enumerate(rows[1:], start=2):
                if len(row) <= idx_projeto:
                    continue

                nome = str(row[idx_projeto] or '').strip()
                if not nome:
                    continue

                key = self._normalize_string(nome)
                projetos_map[key] = {
                    'codigo': '',
                    'nome': nome,
                    'fluxo': fluxo,
                    'fonte': f'Acompanhamento de Agenda, aba \'{sheet_name}\', linha {i}',
                    'detalhe': f'Aba determina fluxo (Super → SUPER, outros → NAO_SUPER)'
                }

        return projetos_map

    def handle(self, *args: Any, **options: Any) -> None:
        xls_controle = Path(options['xls_controle'])
        xls_agenda = Path(options['xls_agenda'])
        dry_run = options['dry_run']
        verbose = options['verbose']

        if load_workbook is None:
            self.stdout.write(
                self.style.ERROR("❌ openpyxl não instalado. Instale com: pip install openpyxl")
            )
            return

        if dry_run:
            self.stdout.write(self.style.WARNING("🔍 DRY-RUN MODE (não persiste mudanças)"))

        self.stdout.write(self.style.WARNING("📊 Lendo planilhas XLSX..."))

        # Extrair dados
        projetos_controle = {}
        projetos_agenda = {}

        if xls_controle.exists():
            self.stdout.write(f"   • Planilha de Controle: {xls_controle}")
            projetos_controle = self._extract_projetos_from_controle(xls_controle, verbose)
        else:
            self.stdout.write(self.style.WARNING(f"  ⚠️ Arquivo não encontrado: {xls_controle}"))

        if xls_agenda.exists():
            self.stdout.write(f"   • Acompanhamento de Agenda: {xls_agenda}")
            projetos_agenda = self._extract_projetos_from_agenda(xls_agenda, verbose)
        else:
            self.stdout.write(self.style.WARNING(f"  ⚠️ Arquivo não encontrado: {xls_agenda}"))

        # Combinar (Controle tem prioridade)
        projetos_map = {**projetos_agenda, **projetos_controle}

        if not projetos_map:
            self.stdout.write(self.style.WARNING("  ⚠️ Nenhum projeto encontrado nas planilhas"))
            return

        # Processar atualizações
        updated_count = 0
        ignored_count = 0
        not_found_count = 0
        csv_rows = []

        for key, data in projetos_map.items():
            codigo = data['codigo']
            nome = data['nome']
            fluxo_target = data['fluxo']
            fonte = data['fonte']
            detalhe = data['detalhe']

            projeto = self._find_projeto(codigo, nome)

            if not projeto:
                if verbose:
                    self.stdout.write(
                        self.style.WARNING(f"  ⊘ Projeto não encontrado: {nome} (código: {codigo or 'N/A'})")
                    )
                not_found_count += 1
                continue

            # Idempotência
            if projeto.fluxo == fluxo_target:
                if verbose:
                    self.stdout.write(
                        self.style.WARNING(f"  ⊙ Já atualizado: {projeto.nome} ({fluxo_target})")
                    )
                ignored_count += 1
            else:
                prev_fluxo = projeto.fluxo
                projeto.fluxo = fluxo_target

                if not dry_run:
                    projeto.save()

                if verbose or not dry_run:
                    self.stdout.write(
                        self.style.SUCCESS(
                            f"  ✓ {'[DRY] ' if dry_run else ''}Atualizado: {projeto.nome} ({prev_fluxo} → {fluxo_target})"
                        )
                    )
                updated_count += 1

            # Registrar para CSV
            csv_rows.append({
                'codigo': projeto.codigo or '',
                'nome': projeto.nome,
                'fluxo': fluxo_target,
                'fonte': fonte,
                'detalhe': detalhe
            })

        # Gerar CSV rastreável
        if not dry_run and csv_rows:
            csv_path = Path('apps/core/data/projetos_fluxo_resolved.csv')
            csv_path.parent.mkdir(parents=True, exist_ok=True)
            with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=['codigo', 'nome', 'fluxo', 'fonte', 'detalhe'])
                writer.writeheader()
                writer.writerows(csv_rows)
            self.stdout.write(self.style.SUCCESS(f"📄 CSV gerado: {csv_path}"))

        # Sumário final
        self.stdout.write("")
        self.stdout.write(
            self.style.SUCCESS(
                f"✅ {'[DRY-RUN] ' if dry_run else ''}Seed concluído:"
            )
        )
        self.stdout.write(f"   • Atualizados: {updated_count}")
        self.stdout.write(f"   • Ignorados (já OK): {ignored_count}")
        self.stdout.write(f"   • Não encontrados: {not_found_count}")

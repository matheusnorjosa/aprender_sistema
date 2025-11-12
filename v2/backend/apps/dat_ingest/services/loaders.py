"""
ETL Loaders - Funções puras para carregar dados de planilhas Excel
"""
# pyright: reportUnknownVariableType=false, reportUnknownMemberType=false, reportUnknownParameterType=false, reportUnknownArgumentType=false, reportMissingParameterType=false, reportOptionalMemberAccess=false, reportCallIssue=false, reportOptionalSubscript=false, reportArgumentType=false, reportMissingTypeStubs=false, reportAttributeAccessIssue=false, reportReturnType=false, reportGeneralTypeIssues=false

import hashlib
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .normalizers import (
    generate_cpf_from_email,
    normalize_cpf,
    normalize_email,
    normalize_str,
    normalize_telefone,
    normalize_uf,
    parse_bool,
    titlecase,
)


def compute_file_hash(filepath: Path) -> str:
    """
    Calcula SHA256 de um arquivo

    Args:
        filepath: Caminho do arquivo

    Returns:
        Hash SHA256 em hexadecimal
    """
    return hashlib.sha256(filepath.read_bytes()).hexdigest()


def load_workbook_index(directory: Path) -> list[Path]:
    """
    Retorna lista de arquivos .xlsx no diretório

    Args:
        directory: Diretório a escanear

    Returns:
        Lista de Paths para arquivos .xlsx ordenados por nome
    """
    if not directory.exists():
        return []

    return sorted(directory.glob("*.xlsx"))


def parse_usuarios(filepath: Path) -> list[dict[str, Any]]:
    """
    Parse planilha de usuários

    Estrutura esperada (aba "Ativos"):
    | Nome | CPF | Telefone | Email | Perfil | Superintendência | Ativo |

    Args:
        filepath: Caminho do arquivo Excel

    Returns:
        Lista de dicts com dados normalizados
    """
    wb = load_workbook(filepath, data_only=True)

    # Tenta várias abas possíveis
    sheet_names = ["Ativos", "Usuários", "Users", "Sheet1"]
    ws = None
    for name in sheet_names:
        if name in wb.sheetnames:
            ws = wb[name]
            break

    if ws is None:
        return []

    usuarios = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) < 4:
            continue

        # Estrutura do teste: Nome, CPF, Telefone, Email, Cargo, Gerência, Perfil, Superintendência, Ativo
        nome = normalize_str(row[0]) if len(row) > 0 else ""
        cpf = normalize_cpf(row[1]) if len(row) > 1 else ""
        telefone = normalize_telefone(row[2]) if len(row) > 2 else ""
        email = normalize_email(row[3]) if len(row) > 3 else ""
        cargo = normalize_str(row[4]) if len(row) > 4 else "Formador"
        gerencia = normalize_str(row[5]) if len(row) > 5 else ""
        perfil_col = normalize_str(row[6]) if len(row) > 6 else ""
        superintendencia_col = normalize_str(row[7]) if len(row) > 7 else ""
        ativo = parse_bool(row[8], default=True) if len(row) > 8 else True

        # Mapear cargo/perfil_col/superintendencia_col para perfil final
        perfil = perfil_col if perfil_col else cargo if cargo else "Formador"
        if "superintend" in perfil.lower() or "superintend" in gerencia.lower() or superintendencia_col:
            perfil = "Superintendência"
        elif "coordenador" in perfil.lower():
            perfil = "Coordenador"
        elif "formador" in perfil.lower():
            perfil = "Formador"

        # Validação mínima: precisa ter nome OU email (permite inativos vazios)
        if not nome and not email:
            continue

        # Gerar CPF determinístico se faltando (teste: test_generates_cpf_if_missing)
        if not cpf and email:
            cpf = generate_cpf_from_email(email)

        usuarios.append(
            {
                "nome": titlecase(nome),
                "email": email,
                "cpf": cpf,
                "telefone": telefone,
                "perfil": perfil,
                "ativo": ativo,
                "src": f"{filepath.name}/Ativos",
                "rownum": i,
            }
        )

    return usuarios


def parse_municipios(filepath: Path) -> list[dict[str, Any]]:
    """
    Parse municípios de arquivo Excel

    Aceita dois formatos:
    1. Aba "Municípios" simples: Nome | UF | Ativo
    2. Aba "FILTRO_PROD.": Índice | Município - UF

    Args:
        filepath: Caminho do arquivo Excel

    Returns:
        Lista de dicts com dados normalizados
    """
    wb = load_workbook(filepath, data_only=True)

    # Tentar primeiro aba simples "Municípios"
    ws = None
    sheet_name = None
    for nome in ["Municípios", "Municipios", "Sheet1"]:
        if nome in wb.sheetnames:
            ws = wb[nome]
            sheet_name = nome
            break

    # Fallback: procurar aba FILTRO_PROD.
    if ws is None:
        for nome in wb.sheetnames:
            if 'FILTRO' in nome.upper() and 'PROD' in nome.upper():
                ws = wb[nome]
                sheet_name = nome
                break

    if ws is None:
        return []

    municipios = []
    municipios_vistos = set()

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) < 1:
            continue

        # Formato simples: Nome | UF | Ativo (teste)
        if sheet_name in ["Municípios", "Municipios", "Sheet1"]:
            nome = normalize_str(row[0]) if len(row) > 0 else ""
            uf = normalize_str(row[1]) if len(row) > 1 else ""
            ativo = parse_bool(row[2], default=True) if len(row) > 2 else True

            if not nome:
                continue

            uf_normalizado = normalize_uf(uf) if uf else ""

            municipios.append({
                "nome": titlecase(nome),
                "uf": uf_normalizado.upper() if uf_normalizado else "",
                "ativo": ativo,
                "src": f"{filepath.name}/{sheet_name}",
                "rownum": i,
            })

        # Formato FILTRO_PROD: Índice | Município - UF
        else:
            municipio_uf = normalize_str(row[1]) if len(row) > 1 else ""

            if not municipio_uf or municipio_uf in municipios_vistos:
                continue

            # Separar "SOBRAL - CE" → nome="SOBRAL", uf="CE"
            if " - " in municipio_uf:
                partes = municipio_uf.split(" - ", 1)
                nome = partes[0].strip()
                uf = partes[1].strip() if len(partes) > 1 else ""
            else:
                nome = municipio_uf.strip()
                uf = ""

            uf_normalizado = normalize_uf(uf) if uf else ""

            if not nome:
                continue

            municipios_vistos.add(municipio_uf)

            municipios.append({
                "nome": titlecase(nome),
                "uf": uf_normalizado.upper() if uf_normalizado else "",
                "ativo": True,
                "src": f"{filepath.name}/FILTRO_PROD",
                "rownum": i,
            })

    return municipios


def parse_projetos(filepath: Path) -> list[dict[str, Any]]:
    """
    Parse projetos de arquivo Excel

    Aceita dois formatos:
    1. Aba "Projetos" simples: Nome | Descrição | Ativo
    2. Aba "FILTRO_PROD.": Índice | ... | ... | ... | Projeto (col E)

    Args:
        filepath: Caminho do arquivo Excel

    Returns:
        Lista de dicts com dados normalizados
    """
    wb = load_workbook(filepath, data_only=True)

    # Tentar primeiro aba simples "Projetos"
    ws = None
    sheet_name = None
    for nome in ["Projetos", "Projeto", "Sheet1"]:
        if nome in wb.sheetnames:
            ws = wb[nome]
            sheet_name = nome
            break

    # Fallback: procurar aba FILTRO_PROD.
    if ws is None:
        for nome in wb.sheetnames:
            if 'FILTRO' in nome.upper() and 'PROD' in nome.upper():
                ws = wb[nome]
                sheet_name = nome
                break

    if ws is None:
        return []

    projetos = []
    projetos_vistos = set()  # Evitar duplicatas

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) < 1:
            continue

        # Formato simples: Nome | Descrição | Ativo (teste)
        if sheet_name in ["Projetos", "Projeto", "Sheet1"]:
            nome = normalize_str(row[0]) if len(row) > 0 else ""
            descricao = normalize_str(row[1]) if len(row) > 1 else ""
            ativo = parse_bool(row[2], default=True) if len(row) > 2 else True

            if not nome:
                continue

            projetos.append({
                "nome": nome,
                "descricao": descricao,
                "ativo": ativo,
                "src": f"{filepath.name}/{sheet_name}",
                "rownum": i,
            })

        # Formato FILTRO_PROD: Índice | ... | ... | ... | Projeto (col E)
        else:
            if len(row) < 5:
                continue

            nome_projeto = normalize_str(row[4]) if len(row) > 4 else ""  # Col E (índice 4)

            if not nome_projeto or nome_projeto in projetos_vistos:
                continue

            # Normalizar nomes especiais
            if "IDEB" in nome_projeto.upper() or "IDEB10" in nome_projeto.upper():
                nome_projeto = "Gestão Escolar"

            projetos_vistos.add(nome_projeto)

            projetos.append(
                {
                    "nome": nome_projeto,
                    "descricao": "",
                    "ativo": True,
                    "src": f"{filepath.name}/FILTRO_PROD",
                    "rownum": i,
                }
            )

    return projetos


def parse_tipos_evento(filepath: Path) -> list[dict[str, Any]]:
    """
    Parse planilha de tipos de evento

    Estrutura esperada:
    | Nome | Descrição | Cor (hex) |

    Args:
        filepath: Caminho do arquivo Excel

    Returns:
        Lista de dicts com dados normalizados
    """
    wb = load_workbook(filepath, data_only=True)

    # Tenta várias abas possíveis
    sheet_names = ["Tipos Evento", "TiposEvento", "Tipos", "Sheet1"]
    ws = None
    for name in sheet_names:
        if name in wb.sheetnames:
            ws = wb[name]
            break

    if ws is None:
        return []

    tipos = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) < 1:
            continue

        nome = normalize_str(row[0]) if len(row) > 0 else ""
        descricao = normalize_str(row[1]) if len(row) > 1 else ""
        cor = normalize_str(row[2]) if len(row) > 2 else ""

        # Validação de cor hex (opcional)
        if cor and not cor.startswith("#"):
            cor = f"#{cor}"
        if cor and len(cor) != 7:
            cor = ""

        # Validação mínima
        if not nome:
            continue

        tipos.append(
            {
                "nome": nome,
                "descricao": descricao,
                "cor": cor,
                "src": f"{filepath.name}/TiposEvento",
                "rownum": i,
            }
        )

    return tipos


# ============================================================================
# Shims/Adapters para testes (Issue #39)
# ============================================================================


def extract_tipos_evento_from_agenda(filepath: Path) -> list[dict[str, Any]]:
    """
    Extrai tipos de evento de uma planilha de agenda (shim para testes)

    Varre as abas da agenda procurando por valores únicos na coluna "Tipo de Evento".

    Args:
        filepath: Caminho do arquivo Excel de agenda

    Returns:
        Lista de dicts com tipos de evento extraídos

    Examples:
        >>> extract_tipos_evento_from_agenda(Path("agenda.xlsx"))
        [{"nome": "Formação Presencial", "src": "agenda.xlsx/Super", "rownum": 2}, ...]
    """
    wb = load_workbook(filepath, data_only=True)

    tipos_vistos = set()
    tipos = []

    # Varre todas as abas procurando por coluna "Tipo de Evento"
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Procura header "Tipo de Evento" ou "Tipo"
        header_row = None
        tipo_col_idx = None

        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            for idx, cell in enumerate(row):
                if cell and "tipo" in str(cell).lower():
                    header_row = row
                    tipo_col_idx = idx
                    break
            if tipo_col_idx is not None:
                break

        if tipo_col_idx is None:
            continue

        # Extrai valores únicos da coluna
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) <= tipo_col_idx:
                continue

            tipo_nome = normalize_str(row[tipo_col_idx])
            if not tipo_nome or tipo_nome in tipos_vistos:
                continue

            tipos_vistos.add(tipo_nome)
            tipos.append(
                {
                    "nome": tipo_nome,
                    "descricao": "",
                    "cor": "",
                    "src": filepath.name,
                    "rownum": i,
                }
            )

    return tipos

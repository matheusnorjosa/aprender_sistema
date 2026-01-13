"""
Parser para planilha Acompanhamento de Agenda
Abas: ACerta, Outros, Super, Brincando, Vidas
"""
# pyright: reportUnknownVariableType=false, reportUnknownMemberType=false, reportUnknownParameterType=false, reportUnknownArgumentType=false, reportMissingParameterType=false, reportOptionalMemberAccess=false, reportCallIssue=false, reportOptionalSubscript=false, reportArgumentType=false, reportMissingTypeStubs=false, reportAttributeAccessIssue=false, reportReturnType=false, reportGeneralTypeIssues=false


from __future__ import annotations
from datetime import datetime, time
from pathlib import Path
from typing import Any

import pytz
from openpyxl import load_workbook

from .normalizers import normalize_str, parse_bool


def combine_datetime(data, hora, tz_name='America/Fortaleza'):
    """Combina date e time em timezone-aware datetime"""
    if not data or not hora:
        return None

    # Converter hora para time se for string
    if isinstance(hora, str):
        try:
            hora = datetime.strptime(hora, '%H:%M').time()
        except:
            return None
    elif not isinstance(hora, time):
        return None

    # Combinar
    dt = datetime.combine(data, hora)

    # Localizar no timezone
    tz = pytz.timezone(tz_name)
    return tz.localize(dt)


def parse_acompanhamento_aba(filepath: Path, aba_name: str) -> list[dict[str, Any]]:
    """
    Parse uma aba de acompanhamento

    Estrutura (colunas A-T):
    A: Criado na Agenda
    D: C ancelar
    E: município
    F: encontro
    G: tipo
    H: data
    I: hora início
    J: hora fim
    K: projeto
    L: segmento
    M: Coord Acompanha
    N: Coordenador
    O-S: Formador 1-5
    T: Convidados

    Args:
        filepath: Caminho do arquivo Excel
        aba_name: Nome da aba (ACerta, Outros, Super, Brincando, Vidas)

    Returns:
        Lista de dicts com solicitações
    """
    wb = load_workbook(filepath, data_only=True)

    if aba_name not in wb.sheetnames:
        return []

    ws = wb[aba_name]
    solicitacoes = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) < 10:
            continue

        # Extrair campos
        criado_agenda = normalize_str(row[0]) if len(row) > 0 else ""  # Col A
        cancelar = normalize_str(row[3]) if len(row) > 3 else ""  # Col D
        municipio_uf = normalize_str(row[4]) if len(row) > 4 else ""  # Col E
        encontro = normalize_str(row[5]) if len(row) > 5 else ""  # Col F
        tipo_modal = normalize_str(row[6]) if len(row) > 6 else ""  # Col G (Presencial/Remoto)
        data = row[7] if len(row) > 7 else None  # Col H
        hora_inicio = row[8] if len(row) > 8 else None  # Col I
        hora_fim = row[9] if len(row) > 9 else None  # Col J
        projeto_nome = normalize_str(row[10]) if len(row) > 10 else ""  # Col K
        segmento = normalize_str(row[11]) if len(row) > 11 else ""  # Col L
        coord_acompanha = row[12] if len(row) > 12 else False  # Col M
        coordenador_nome = normalize_str(row[13]) if len(row) > 13 else ""  # Col N
        formador1 = normalize_str(row[14]) if len(row) > 14 else ""  # Col O
        formador2 = normalize_str(row[15]) if len(row) > 15 else ""  # Col P
        formador3 = normalize_str(row[16]) if len(row) > 16 else ""  # Col Q
        formador4 = normalize_str(row[17]) if len(row) > 17 else ""  # Col R
        formador5 = normalize_str(row[18]) if len(row) > 18 else ""  # Col S
        convidados = normalize_str(row[19]) if len(row) > 19 else ""  # Col T

        # Validar dados obrigatórios
        if not data or not hora_inicio or not hora_fim:
            continue

        # Combinar data + hora
        inicio = combine_datetime(data, hora_inicio)
        fim = combine_datetime(data, hora_fim)

        if not inicio or not fim:
            continue

        # Determinar status
        if cancelar and cancelar.lower() in ['x', 'remarcado', 'cancelado']:
            status = 'cancelado'
        elif criado_agenda and criado_agenda.upper() == 'SIM':
            status = 'aprovado'
        else:
            status = 'pendente'

        # Coletar formadores não vazios
        formadores = [f for f in [formador1, formador2, formador3, formador4, formador5] if f]

        solicitacoes.append({
            'municipio_nome_uf': municipio_uf,
            'projeto_nome': projeto_nome,
            'tipo': 'evento',
            'encontro': encontro,
            'segmento': segmento,
            'coordenador_acompanha': parse_bool(coord_acompanha, default=False),
            'coordenador_nome': coordenador_nome,
            'formadores': formadores,
            'convidados': convidados,
            'inicio': inicio,
            'fim': fim,
            'status': status,
            'tipo_modal': tipo_modal,  # Presencial/Remoto
            'src': f"{filepath.name}/{aba_name}",
            'rownum': i,
        })

    return solicitacoes


def parse_acompanhamento_super(filepath: Path) -> list[dict[str, Any]]:
    """
    Parse aba Super (tem coluna B de Aprovação)

    Args:
        filepath: Caminho do arquivo Excel

    Returns:
        Lista de dicts com solicitações
    """
    wb = load_workbook(filepath, data_only=True)

    if 'Super' not in wb.sheetnames:
        return []

    ws = wb['Super']
    solicitacoes = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) < 10:
            continue

        # Extrair aprovação (coluna B)
        aprovacao = normalize_str(row[1]) if len(row) > 1 else ""  # Col B

        # Resto dos campos (mesma estrutura que outras abas)
        criado_agenda = normalize_str(row[0]) if len(row) > 0 else ""
        cancelar = normalize_str(row[3]) if len(row) > 3 else ""
        municipio_uf = normalize_str(row[4]) if len(row) > 4 else ""
        encontro = normalize_str(row[5]) if len(row) > 5 else ""
        tipo_modal = normalize_str(row[6]) if len(row) > 6 else ""
        data = row[7] if len(row) > 7 else None
        hora_inicio = row[8] if len(row) > 8 else None
        hora_fim = row[9] if len(row) > 9 else None
        projeto_nome = normalize_str(row[10]) if len(row) > 10 else ""
        segmento = normalize_str(row[11]) if len(row) > 11 else ""
        coord_acompanha = row[12] if len(row) > 12 else False
        coordenador_nome = normalize_str(row[13]) if len(row) > 13 else ""
        formador1 = normalize_str(row[14]) if len(row) > 14 else ""
        formador2 = normalize_str(row[15]) if len(row) > 15 else ""
        formador3 = normalize_str(row[16]) if len(row) > 16 else ""
        formador4 = normalize_str(row[17]) if len(row) > 17 else ""
        formador5 = normalize_str(row[18]) if len(row) > 18 else ""
        convidados = normalize_str(row[19]) if len(row) > 19 else ""

        if not data or not hora_inicio or not hora_fim:
            continue

        inicio = combine_datetime(data, hora_inicio)
        fim = combine_datetime(data, hora_fim)

        if not inicio or not fim:
            continue

        # Determinar status (prioridade: cancelar > aprovacao > criado)
        if cancelar and cancelar.lower() in ['x', 'remarcado', 'cancelado']:
            status = 'cancelado'
        elif aprovacao and 'APROVADO' in aprovacao.upper():
            status = 'aprovado'
        elif aprovacao and ('REPROVADO' in aprovacao.upper() or 'REJEITADO' in aprovacao.upper()):
            status = 'cancelado'
        elif criado_agenda and criado_agenda.upper() == 'SIM':
            status = 'aprovado'
        else:
            status = 'pendente'

        formadores = [f for f in [formador1, formador2, formador3, formador4, formador5] if f]

        solicitacoes.append({
            'municipio_nome_uf': municipio_uf,
            'projeto_nome': projeto_nome,
            'tipo': 'evento',
            'encontro': encontro,
            'segmento': segmento,
            'coordenador_acompanha': parse_bool(coord_acompanha, default=False),
            'coordenador_nome': coordenador_nome,
            'formadores': formadores,
            'convidados': convidados,
            'inicio': inicio,
            'fim': fim,
            'status': status,
            'tipo_modal': tipo_modal,
            'aprovacao_super': aprovacao,  # Campo extra
            'src': f"{filepath.name}/Super",
            'rownum': i,
        })

    return solicitacoes


def parse_todas_abas_acompanhamento(filepath: Path) -> list[dict[str, Any]]:
    """
    Parse todas as 5 abas de acompanhamento

    Returns:
        Lista consolidada de todas as solicitações
    """
    todas = []

    # Abas normais
    for aba in ['ACerta', 'Outros', 'Brincando', 'Vidas']:
        todas.extend(parse_acompanhamento_aba(filepath, aba))

    # Aba Super (tem coluna de aprovação)
    todas.extend(parse_acompanhamento_super(filepath))

    return todas

"""
Parser para aba DESLOCAMENTO da planilha Disponibilidade
"""
# pyright: reportUnknownVariableType=false, reportUnknownMemberType=false, reportUnknownParameterType=false, reportUnknownArgumentType=false, reportMissingParameterType=false, reportOptionalMemberAccess=false, reportCallIssue=false, reportOptionalSubscript=false, reportArgumentType=false, reportMissingTypeStubs=false, reportAttributeAccessIssue=false, reportReturnType=false, reportGeneralTypeIssues=false


from __future__ import annotations
from datetime import datetime, time
from pathlib import Path
from typing import Any

import pytz
from openpyxl import load_workbook

from .normalizers import normalize_str


def parse_deslocamentos(filepath: Path) -> list[dict[str, Any]]:
    """
    Parse aba DESLOCAMENTO

    Estrutura REAL da planilha (corrigida 2025-11-19):
    A (Col 1): Origem (cidade, ex: "Fortaleza")
    B (Col 2): Tipo ("Deslocamento" ou "Retorno")
    C (Col 3): Destino (cidade - UF, ex: "Bocaiúva do Sul - PR")
    D (Col 4): Data (datetime, ex: 2025-03-24 00:00:00)
    E-J (Col 5-10): Pessoa 1 a Pessoa 6 (múltiplas pessoas por viagem)

    NOTA: Estrutura antiga (INCORRETA) tinha 10 colunas com 1 pessoa por linha e
    campos saida/chegada/duracao_minutos/meio_transporte que NÃO existem no modelo.

    A planilha real suporta até 6 pessoas por viagem. Cada pessoa gera um registro.

    Modelo Deslocamento tem apenas:
    - usuario, origem, destino, start_date, end_date, observacao, external_hash

    Args:
        filepath: Caminho do arquivo Excel

    Returns:
        Lista de dicts com deslocamentos (1 por pessoa)
    """
    wb = load_workbook(filepath, data_only=True)

    # Procurar aba (pode ter nome com maiúsculas/minúsculas)
    aba_name = None
    for nome in wb.sheetnames:
        if 'DESLOCAMENTO' in nome.upper():
            aba_name = nome
            break

    if not aba_name:
        return []

    ws = wb[aba_name]
    deslocamentos = []
    tz = pytz.timezone('America/Fortaleza')

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) < 5:
            continue

        # Estrutura correta: Origem | Tipo | Destino | Data | Pessoa1...Pessoa6
        origem = normalize_str(row[0]) if row[0] else ""
        tipo = normalize_str(row[1]) if row[1] else ""
        destino = normalize_str(row[2]) if row[2] else ""
        data = row[3] if row[3] else None

        # Validar campos obrigatórios
        if not origem or not destino or not data:
            continue

        # Processar data
        try:
            # Se já é datetime, usar direto; se é date, converter
            if isinstance(data, datetime):
                data_aware = data
            else:
                data_aware = datetime.combine(data, time(0, 0))

            # Localizar para America/Fortaleza se naive
            if data_aware.tzinfo is None:
                data_aware = tz.localize(data_aware)

        except Exception:
            continue

        # Extrair pessoas (colunas 5-10)
        pessoas = []
        for col_idx in range(4, min(10, len(row))):
            pessoa = normalize_str(row[col_idx]) if row[col_idx] else ""
            if pessoa:
                pessoas.append(pessoa)

        # Se não encontrou nenhuma pessoa, pular linha
        if not pessoas:
            continue

        # Criar um registro para cada pessoa
        for pessoa_nome in pessoas:
            deslocamentos.append({
                'formador_nome': pessoa_nome,
                'origem_nome_uf': origem,
                'destino_nome_uf': destino,
                'data': data_aware,
                'tipo': tipo,  # Guardado em observacao pelo ETL
                'src': f"{filepath.name}/{aba_name}",
                'rownum': i,
            })

    return deslocamentos

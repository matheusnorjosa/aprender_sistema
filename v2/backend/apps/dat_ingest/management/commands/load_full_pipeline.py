"""
Management Command: load_full_pipeline
Carrega todo o pipeline de dados: usuários, municípios, projetos, tipos de evento.
"""
# pyright: reportUnknownVariableType=false, reportUnknownMemberType=false, reportUnknownParameterType=false, reportUnknownArgumentType=false, reportMissingParameterType=false, reportOptionalMemberAccess=false, reportCallIssue=false, reportOptionalSubscript=false, reportArgumentType=false, reportMissingTypeStubs=false, reportAttributeAccessIssue=false, reportReturnType=false, reportGeneralTypeIssues=false

from __future__ import annotations

from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError, CommandParser
from django.db import transaction

from apps.core.models import Municipio, Projeto, TipoEvento, Usuario
from apps.dat_ingest.models import StgMunicipio, StgProjeto, StgTipoEvento, StgUsuario
from apps.dat_ingest.services.loaders import (
    extract_tipos_evento_from_agenda,
    parse_usuarios,
)


class Command(BaseCommand):
    help = "Load full ETL pipeline from Excel files"

    def add_arguments(self, parser: CommandParser) -> None:
        parser.add_argument(
            "--users-file",
            type=str,
            required=True,
            help="Path to users Excel file",
        )
        parser.add_argument(
            "--agenda-file",
            type=str,
            required=True,
            help="Path to agenda Excel file",
        )
        parser.add_argument(
            "--disponibilidade-file",
            type=str,
            help="Path to disponibilidade Excel file (optional)",
        )
        parser.add_argument(
            "--controle-file",
            type=str,
            help="Path to controle Excel file (optional)",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Simulate without writing to DB",
        )

    def handle(self, *args: Any, **options: Any) -> None:
        users_file = options["users_file"]
        agenda_file = options["agenda_file"]
        disponibilidade_file = options.get("disponibilidade_file")
        controle_file = options.get("controle_file")
        dry_run = options["dry_run"]

        if dry_run:
            self.stdout.write(self.style.WARNING("[DRY-RUN MODE] No changes will be saved"))

        # Validate required files exist
        users_path = Path(users_file)
        agenda_path = Path(agenda_file)

        if not users_path.exists():
            raise CommandError(f"Users file not found: {users_file}")
        if not agenda_path.exists():
            raise CommandError(f"Agenda file not found: {agenda_file}")

        if dry_run:
            self.stdout.write(self.style.SUCCESS("[DRY-RUN] Would process all files"))
            return

        # Process all entities
        self._load_usuarios(users_path)
        self._load_municipios_from_agenda(agenda_path)
        self._load_projetos_from_agenda(agenda_path)
        self._load_tipos_evento_from_agenda(agenda_path)

        # Process optional files if provided
        if disponibilidade_file:
            disp_path = Path(disponibilidade_file)
            if disp_path.exists():
                self.stdout.write(f"[Disponibilidade] File provided but not yet implemented")

        if controle_file:
            controle_path = Path(controle_file)
            if controle_path.exists():
                self.stdout.write(f"[Controle] File provided but not yet implemented")

        self.stdout.write(self.style.SUCCESS("\n[PIPELINE] All entities loaded successfully"))

    def _load_usuarios(self, filepath: Path) -> None:
        """Load usuarios: parse → staging → SSOT"""
        self.stdout.write(f"\n[Usuarios] Loading from {filepath.name}...")

        # Parse
        usuarios_data = parse_usuarios(filepath)
        self.stdout.write(f"Parsed {len(usuarios_data)} usuarios")

        # Load to staging
        created_stg = self._load_usuarios_staging(usuarios_data)
        self.stdout.write(f"Staging: {created_stg} records")

        # Upsert to SSOT
        created, updated, unchanged = self._upsert_usuarios_ssot()
        self.stdout.write(f"SSOT: {created} created, {updated} updated, {unchanged} unchanged")

    @transaction.atomic
    def _load_usuarios_staging(self, usuarios_data: list[dict[str, Any]]) -> int:
        """Load usuarios to StgUsuario (replace all)"""
        StgUsuario.objects.all().delete()

        stg_records = [
            StgUsuario(
                nome=u["nome"],
                email=u["email"],
                cpf=u["cpf"],
                telefone=u.get("telefone", ""),
                perfil=u["perfil"],
                ativo=u.get("ativo", True),
                src=u.get("src", ""),
                rownum=u.get("rownum", 0),
            )
            for u in usuarios_data
        ]
        StgUsuario.objects.bulk_create(stg_records)
        return len(stg_records)

    @transaction.atomic
    def _upsert_usuarios_ssot(self) -> tuple[int, int, int]:
        """Upsert Usuario by email (natural key)"""
        created_count = 0
        updated_count = 0
        unchanged_count = 0

        stg_usuarios = StgUsuario.objects.all()

        for stg in stg_usuarios:
            email = stg.email.strip().lower()

            # Split nome into first_name/last_name (AbstractUser pattern)
            nome_parts = stg.nome.split() if stg.nome else []
            first_name = nome_parts[0] if nome_parts else ""
            last_name = " ".join(nome_parts[1:]) if len(nome_parts) > 1 else ""

            # Generate username from email
            username = email.split("@")[0]

            defaults = {
                "username": username,
                "email": email,
                "first_name": first_name,
                "last_name": last_name,
                "cpf": stg.cpf or "",
                "telefone": stg.telefone or "",
                "cargo": stg.perfil or "Formador",
                "is_active": stg.ativo,
            }

            obj, created = Usuario.objects.update_or_create(
                email__iexact=email,
                defaults=defaults,
            )

            if created:
                created_count += 1
            else:
                # Check if changed
                changed = any(
                    getattr(obj, field) != value
                    for field, value in defaults.items()
                    if field != "email"  # Don't check lookup field
                )
                if changed:
                    updated_count += 1
                else:
                    unchanged_count += 1

        return created_count, updated_count, unchanged_count

    def _load_municipios_from_agenda(self, filepath: Path) -> None:
        """Extract municipios from agenda and load to SSOT"""
        self.stdout.write(f"\n[Municipios] Extracting from {filepath.name}...")

        # Parse municipios from agenda (extrai de coluna Município)
        municipios_data = self._extract_municipios_from_agenda(filepath)
        self.stdout.write(f"Found {len(municipios_data)} unique municipios")

        # Load to staging
        created_stg = self._load_municipios_staging(municipios_data)
        self.stdout.write(f"Staging: {created_stg} records")

        # Upsert to SSOT
        created, updated, unchanged = self._upsert_municipios_ssot()
        self.stdout.write(f"SSOT: {created} created, {updated} updated, {unchanged} unchanged")

    def _extract_municipios_from_agenda(self, filepath: Path) -> list[dict[str, Any]]:
        """Extract unique municipios from agenda file"""
        from openpyxl import load_workbook

        wb = load_workbook(filepath, data_only=True)
        municipios_vistos = set()
        municipios_data = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Find "Município" column
            mun_col_idx = None
            for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                for idx, cell in enumerate(row):
                    if cell and "município" in str(cell).lower():
                        mun_col_idx = idx
                        break
                if mun_col_idx is not None:
                    break

            if mun_col_idx is None:
                continue

            # Extract unique valores
            for _, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or len(row) <= mun_col_idx:
                    continue

                mun_raw = str(row[mun_col_idx]).strip() if row[mun_col_idx] else ""
                if not mun_raw or mun_raw in municipios_vistos:
                    continue

                # Split "Fortaleza - CE" → (nome, uf)
                if " - " in mun_raw:
                    parts = mun_raw.split(" - ", 1)
                    nome = parts[0].strip()
                    uf = parts[1].strip() if len(parts) > 1 else ""
                else:
                    nome = mun_raw
                    uf = ""

                municipios_vistos.add(mun_raw)
                municipios_data.append({"nome": nome, "uf": uf, "ativo": True})

        return municipios_data

    @transaction.atomic
    def _load_municipios_staging(self, municipios_data: list[dict[str, Any]]) -> int:
        """Load municipios to StgMunicipio"""
        StgMunicipio.objects.all().delete()

        stg_records = [
            StgMunicipio(
                nome=m["nome"],
                uf=m.get("uf", ""),
                ativo=m.get("ativo", True),
                src="agenda",
                rownum=0,
            )
            for m in municipios_data
        ]
        StgMunicipio.objects.bulk_create(stg_records)
        return len(stg_records)

    @transaction.atomic
    def _upsert_municipios_ssot(self) -> tuple[int, int, int]:
        """Upsert Municipio by (nome, uf) natural key"""
        created_count = 0
        updated_count = 0
        unchanged_count = 0

        stg_municipios = StgMunicipio.objects.all()

        for stg in stg_municipios:
            nome = stg.nome.strip()
            uf = (stg.uf or "").strip().upper()

            _, created = Municipio.objects.update_or_create(
                nome__iexact=nome,
                uf__iexact=uf,
                defaults={
                    "nome": nome,
                    "uf": uf,
                },
            )

            if created:
                created_count += 1
            else:
                unchanged_count += 1

        return created_count, updated_count, unchanged_count

    def _load_projetos_from_agenda(self, filepath: Path) -> None:
        """Extract projetos from agenda and load to SSOT"""
        self.stdout.write(f"\n[Projetos] Extracting from {filepath.name}...")

        projetos_data = self._extract_projetos_from_agenda(filepath)
        self.stdout.write(f"Found {len(projetos_data)} unique projetos")

        created_stg = self._load_projetos_staging(projetos_data)
        self.stdout.write(f"Staging: {created_stg} records")

        created, updated, unchanged = self._upsert_projetos_ssot()
        self.stdout.write(f"SSOT: {created} created, {updated} updated, {unchanged} unchanged")

    def _extract_projetos_from_agenda(self, filepath: Path) -> list[dict[str, Any]]:
        """Extract unique projetos from agenda file"""
        from openpyxl import load_workbook

        wb = load_workbook(filepath, data_only=True)
        projetos_vistos = set()
        projetos_data = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Find "Projeto" column
            proj_col_idx = None
            for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                for idx, cell in enumerate(row):
                    if cell and "projeto" in str(cell).lower():
                        proj_col_idx = idx
                        break
                if proj_col_idx is not None:
                    break

            if proj_col_idx is None:
                continue

            # Extract unique valores
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) <= proj_col_idx:
                    continue

                proj_nome = str(row[proj_col_idx]).strip() if row[proj_col_idx] else ""
                if not proj_nome or proj_nome in projetos_vistos:
                    continue

                projetos_vistos.add(proj_nome)
                projetos_data.append({"nome": proj_nome, "descricao": "", "ativo": True})

        return projetos_data

    @transaction.atomic
    def _load_projetos_staging(self, projetos_data: list[dict[str, Any]]) -> int:
        """Load projetos to StgProjeto"""
        StgProjeto.objects.all().delete()

        stg_records = [
            StgProjeto(
                nome=p["nome"],
                descricao=p.get("descricao", ""),
                ativo=p.get("ativo", True),
                src="agenda",
                rownum=0,
            )
            for p in projetos_data
        ]
        StgProjeto.objects.bulk_create(stg_records)
        return len(stg_records)

    @transaction.atomic
    def _upsert_projetos_ssot(self) -> tuple[int, int, int]:
        """Upsert Projeto by nome natural key"""
        created_count = 0
        updated_count = 0
        unchanged_count = 0

        stg_projetos = StgProjeto.objects.all()

        for stg in stg_projetos:
            nome = stg.nome.strip()

            obj, created = Projeto.objects.update_or_create(
                nome__iexact=nome,
                defaults={
                    "nome": nome,
                    "descricao": stg.descricao or "",
                    "fluxo": "NAO_SUPER",  # Default
                },
            )

            if created:
                created_count += 1
            elif obj.descricao != (stg.descricao or ""):
                updated_count += 1
            else:
                unchanged_count += 1

        return created_count, updated_count, unchanged_count

    def _load_tipos_evento_from_agenda(self, filepath: Path) -> None:
        """Extract tipos de evento from agenda and load to SSOT"""
        self.stdout.write(f"\n[Tipos Evento] Extracting from {filepath.name}...")

        tipos_data = extract_tipos_evento_from_agenda(filepath)
        self.stdout.write(f"Found {len(tipos_data)} unique tipos")

        created_stg = self._load_tipos_staging(tipos_data)
        self.stdout.write(f"Staging: {created_stg} records")

        created, updated, unchanged = self._upsert_tipos_ssot()
        self.stdout.write(f"SSOT: {created} created, {updated} updated, {unchanged} unchanged")

    @transaction.atomic
    def _load_tipos_staging(self, tipos_data: list[dict[str, Any]]) -> int:
        """Load tipos to StgTipoEvento"""
        StgTipoEvento.objects.all().delete()

        stg_records = [
            StgTipoEvento(
                nome=t["nome"],
                descricao=t.get("descricao", ""),
                cor=t.get("cor", ""),
                src=t.get("src", ""),
                rownum=t.get("rownum", 0),
            )
            for t in tipos_data
        ]
        StgTipoEvento.objects.bulk_create(stg_records)
        return len(stg_records)

    @transaction.atomic
    def _upsert_tipos_ssot(self) -> tuple[int, int, int]:
        """Upsert TipoEvento by nome natural key"""
        created_count = 0
        updated_count = 0
        unchanged_count = 0

        stg_tipos = StgTipoEvento.objects.all()

        for stg in stg_tipos:
            nome = stg.nome.strip()

            obj, created = TipoEvento.objects.update_or_create(
                nome__iexact=nome,
                defaults={
                    "nome": nome,
                    "descricao": stg.descricao or "",
                    "cor": stg.cor or "",
                },
            )

            if created:
                created_count += 1
            elif obj.descricao != (stg.descricao or "") or obj.cor != (stg.cor or ""):
                updated_count += 1
            else:
                unchanged_count += 1

        return created_count, updated_count, unchanged_count

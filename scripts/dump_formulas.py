import csv
import glob
import os
import re
import sys
from collections import defaultdict

from openpyxl import load_workbook

REF_RE = re.compile(
    r"(?:'(?P<sheet_q>[^']+)'|(?P<sheet>[A-Za-z0-9_ \.-]+))?!?(?P<ref>\$?[A-Z]{1,3}\$?\d+(?::\$?[A-Z]{1,3}\$?\d+)?)"
)
FUNC_RE = re.compile(r"^\s*=\s*([A-Za-z_][A-Za-z0-9\._]*)\s*\(")


def is_formula(cell):
    v = cell.value
    # openpyxl marca data_type = 'f' em modo não-read_only, mas nem sempre; cheque string iniciando com '='
    if getattr(cell, "data_type", None) == "f":
        return True
    return isinstance(v, str) and v.strip().startswith("=")


def func_name(formula):
    m = FUNC_RE.match(formula or "")
    return m.group(1).upper() if m else ""


def scan_file(path):
    wb = load_workbook(path, data_only=False, read_only=True)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if is_formula(cell):
                    f = str(cell.value)
                    yield os.path.basename(path), ws.title, cell.coordinate, func_name(
                        f
                    ), f


def extract_refs(formula):
    refs = []
    for m in REF_RE.finditer(formula or ""):
        sheet = m.group("sheet_q") or m.group("sheet") or ""
        ref = m.group("ref") or ""
        if ref:
            refs.append((sheet.strip(), ref))
    return refs


def main(pattern):
    files = sorted(glob.glob(pattern))
    os.makedirs("out_formulas", exist_ok=True)

    inv_csv = "out_formulas/formulas_inventory.csv"
    by_sheet_csv = "out_formulas/formulas_summary_by_sheet.csv"
    by_func_csv = "out_formulas/formulas_summary_by_function.csv"
    refs_csv = "out_formulas/formulas_references.csv"
    flags_md = "out_formulas/formulas_flags.md"

    rows = []
    err = []
    for fp in files:
        try:
            rows.extend(scan_file(fp))
        except Exception as e:
            err.append((os.path.basename(fp), str(e)))

    # inventory
    with open(inv_csv, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["file", "sheet", "cell", "function", "formula"])
        w.writerows(rows)

    # by sheet
    by_sheet = defaultdict(int)
    # by function
    by_func = defaultdict(int)
    # references
    refs = []

    # --- NOVO BLOCO DE CLASSIFICAÇÃO / FLAGS ---
    TARGET_TOKENS = {
        "VOLATILE": ["INDIRECT", "OFFSET", "NOW", "TODAY", "RAND", "RANDBETWEEN"],
        "CROSS_DOC": ["IMPORTRANGE"],
        "HEAVY": [
            "QUERY",
            "FILTER",
            "ARRAYFORMULA",
            "UNIQUE",
            "VLOOKUP",
            "XLOOKUP",
            "MATCH",
            "INDEX",
            "REGEXEXTRACT",
            "REGEXREPLACE",
            "SUMPRODUCT",
        ],
    }

    token_counts = {k: {t: 0 for t in v} for k, v in TARGET_TOKENS.items()}

    flags = {
        "volatile_hits": 0,
        "cross_doc_hits": 0,
        "other_heavy_hits": 0,
        "volatile_samples": [],
        "cross_doc_samples": [],
        "other_heavy_samples": [],
    }

    def seen_token(tk, formula_up):
        # procura padrão TK( … ) com tolerância a espaços; evita pegar "palavras soltas"
        return re.search(rf"\b{tk}\s*\(", formula_up) is not None

    # também vamos escrever um CSV novo com contagem por token
    tokens_csv = "out_formulas/formulas_token_counts.csv"

    for f, s, cell, fn, formula in rows:
        by_sheet[(f, s)] += 1
        by_func[(f, fn)] += 1
        for sh, ref in extract_refs(formula):
            refs.append((f, s, cell, sh, ref))

        FUP = (formula or "").upper()
        # VOLATILE
        for tk in TARGET_TOKENS["VOLATILE"]:
            if seen_token(tk, FUP):
                token_counts["VOLATILE"][tk] += 1
                flags["volatile_hits"] += 1
                if len(flags["volatile_samples"]) < 20:
                    flags["volatile_samples"].append((f, s, cell, tk))
        # CROSS_DOC
        for tk in TARGET_TOKENS["CROSS_DOC"]:
            if seen_token(tk, FUP):
                token_counts["CROSS_DOC"][tk] += 1
                flags["cross_doc_hits"] += 1
                if len(flags["cross_doc_samples"]) < 20:
                    flags["cross_doc_samples"].append((f, s, cell, tk))
        # HEAVY
        for tk in TARGET_TOKENS["HEAVY"]:
            if seen_token(tk, FUP):
                token_counts["HEAVY"][tk] += 1
                flags["other_heavy_hits"] += 1
                if len(flags["other_heavy_samples"]) < 20:
                    flags["other_heavy_samples"].append((f, s, cell, tk))

    # escreve tokens_csv
    with open(tokens_csv, "w", newline="", encoding="utf-8-sig") as fcsv:
        w = csv.writer(fcsv)
        w.writerow(["bucket", "token", "count"])
        for bucket, d in token_counts.items():
            for tk, ct in sorted(d.items(), key=lambda x: (-x[1], x[0])):
                w.writerow([bucket, tk, ct])

    with open(by_sheet_csv, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["file", "sheet", "formula_count"])
        for (f, s), c in sorted(by_sheet.items()):
            w.writerow([f, s, c])

    with open(by_func_csv, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["file", "function", "count"])
        for (f, fn), c in sorted(
            by_func.items(), key=lambda x: (-x[1], x[0][0], x[0][1])
        ):
            w.writerow([f, fn or "", c])

    with open(refs_csv, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["file", "sheet", "cell", "ref_sheet", "ref_range"])
        w.writerows(refs)

    with open(flags_md, "w", encoding="utf-8") as f:
        f.write("# Fórmulas — Sinais de Risco/Complexidade\n\n")
        f.write(
            f"- Voláteis (INDIRECT/OFFSET/NOW/TODAY/RAND/RANDBETWEEN): {flags['volatile_hits']}\n"
        )
        f.write(f"- Cross-doc (IMPORTRANGE): {flags['cross_doc_hits']}\n")
        f.write(
            f"- Pesadas (QUERY/FILTER/ARRAYFORMULA/UNIQUE/VLOOKUP/XLOOKUP/MATCH/INDEX/REGEX*): {flags['other_heavy_hits']}\n\n"
        )

        def block(title, items):
            f.write(f"## {title}\n\n")
            if not items:
                f.write("_sem amostras_\n\n")
                return
            for it in items:
                f.write(f"- {it}\n")
            f.write("\n")

        block("Amostras voláteis (até 20)", flags["volatile_samples"])
        block("Amostras cross-doc (até 20)", flags["cross_doc_samples"])
        block("Amostras pesadas (até 20)", flags["other_heavy_samples"])

    print("[OK] Inventários e flags salvos em out_formulas/")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python dump_formulas.py <pattern>")
        print('Exemplo: python dump_formulas.py "/app/data/csv-import/*.xlsx"')
        sys.exit(1)
    main(sys.argv[1])

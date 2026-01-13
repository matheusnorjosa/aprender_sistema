"""
Tests: Loaders
Valida funções de parse de planilhas (services/loaders.py).
"""

# pyright: reportMissingParameterType=false, reportUnknownParameterType=false, reportUnknownArgumentType=false, reportUnknownVariableType=false, reportUnknownMemberType=false, reportOptionalMemberAccess=false, reportAttributeAccessIssue=false, reportArgumentType=false, reportMissingTypeArgument=false, reportCallIssue=false

from __future__ import annotations
import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook

from apps.dat_ingest.services.loaders import (
    extract_tipos_evento_from_agenda,
    parse_municipios,
    parse_projetos,
    parse_usuarios,
)


@pytest.fixture
def temp_usuarios_xlsx(tmp_path):
    """Cria arquivo temporário de usuários"""
    filepath = tmp_path / "usuarios.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Ativos"

    # Header
    ws.append([
        "Nome", "CPF", "Telefone", "Email", "Cargo",
        "Gerência", "Perfil", "Superintendência", "Ativo"
    ])

    # Data rows
    ws.append([
        "João Silva", "12345678901", "(85) 99999-9999",
        "joao.silva@example.com", "Formador", "DAT",
        "Formador", "", "Sim"
    ])
    ws.append([
        "Maria Santos", "98765432100", "(85) 88888-8888",
        "maria.santos@example.com", "Coordenadora", "Superintendência",
        "Coordenador", "Sim", "Sim"
    ])
    ws.append([
        "Pedro Inativo", "", "", "pedro@example.com",
        "Formador", "DAT", "Formador", "", "Não"
    ])

    wb.save(filepath)
    return filepath


@pytest.fixture
def temp_agenda_xlsx(tmp_path):
    """Cria arquivo temporário de agenda"""
    filepath = tmp_path / "agenda.xlsx"
    wb = Workbook()

    # Aba Super
    ws = wb.active
    ws.title = "Super"

    # Header
    ws.append([
        "Data", "Horário Inicial", "Horário Final", "Projeto",
        "Tipo de Evento", "Município", "Coordenador", "Formador(es)",
        "Situação", "Observações"
    ])

    # Data rows com tipos variados
    ws.append([
        "2025-01-15", "14:00", "17:00", "ACerta",
        "Formação Presencial", "Fortaleza - CE", "Ellen Damares",
        "João Silva", "Realizado", ""
    ])
    ws.append([
        "2025-01-20", "09:00", "12:00", "Novo Lendo",
        "Workshop", "Caucaia - CE", "Aurea Lucia",
        "Maria Santos", "Agendado", ""
    ])
    ws.append([
        "2025-01-25", "14:00", "16:00", "Tema",
        "Formação Presencial", "Maracanaú - CE", "Ellen Damares",
        "Pedro Oliveira", "Pendente", ""
    ])

    wb.save(filepath)
    return filepath


@pytest.fixture
def temp_municipios_xlsx(tmp_path):
    """Cria arquivo temporário de municípios"""
    filepath = tmp_path / "municipios.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Municípios"

    # Header
    ws.append(["Nome", "UF", "Ativo"])

    # Data rows
    ws.append(["Fortaleza", "CE", "Sim"])
    ws.append(["Caucaia", "CE", "Sim"])
    ws.append(["Maracanaú", "CE", "Sim"])
    ws.append(["Sobral", "CE", "Não"])

    wb.save(filepath)
    return filepath


@pytest.fixture
def temp_projetos_xlsx(tmp_path):
    """Cria arquivo temporário de projetos"""
    filepath = tmp_path / "projetos.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Projetos"

    # Header
    ws.append(["Nome", "Descrição", "Ativo"])

    # Data rows
    ws.append(["ACerta", "Projeto ACerta", "Sim"])
    ws.append(["Novo Lendo", "Projeto Novo Lendo", "Sim"])
    ws.append(["Tema", "Projeto Tema", "Sim"])

    wb.save(filepath)
    return filepath


class TestParseUsuarios:
    """Testes para parse_usuarios()"""

    def test_parses_usuarios_from_xlsx(self, temp_usuarios_xlsx):
        """Test: Parse de usuários de arquivo Excel"""
        usuarios = parse_usuarios(temp_usuarios_xlsx)

        assert len(usuarios) == 3, "Deve ter 3 usuários (incluindo inativos)"

    def test_parses_usuario_fields(self, temp_usuarios_xlsx):
        """Test: Parse de todos os campos de usuário"""
        usuarios = parse_usuarios(temp_usuarios_xlsx)

        first_user = usuarios[0]
        assert first_user["nome"] == "João Silva"
        assert first_user["email"] == "joao.silva@example.com"
        assert first_user["cpf"] == "12345678901"
        assert first_user["telefone"] == "85999999999"  # Normalizado
        assert first_user["perfil"] == "Formador"

    def test_normalizes_cpf(self, temp_usuarios_xlsx):
        """Test: Normaliza CPF (remove formatação)"""
        usuarios = parse_usuarios(temp_usuarios_xlsx)

        first_user = usuarios[0]
        assert first_user["cpf"] == "12345678901", "CPF deve estar sem formatação"

    def test_normalizes_telefone(self, temp_usuarios_xlsx):
        """Test: Normaliza telefone (remove formatação)"""
        usuarios = parse_usuarios(temp_usuarios_xlsx)

        first_user = usuarios[0]
        assert first_user["telefone"] == "85999999999", "Telefone deve estar sem formatação"

    def test_parses_ativo_flag(self, temp_usuarios_xlsx):
        """Test: Parse de flag ativo/inativo"""
        usuarios = parse_usuarios(temp_usuarios_xlsx)

        assert usuarios[0]["ativo"] is True, "Primeiro usuário deve estar ativo"
        assert usuarios[1]["ativo"] is True, "Segundo usuário deve estar ativo"
        assert usuarios[2]["ativo"] is False, "Terceiro usuário deve estar inativo"

    def test_generates_cpf_if_missing(self, temp_usuarios_xlsx):
        """Test: Gera CPF determinístico se faltando"""
        usuarios = parse_usuarios(temp_usuarios_xlsx)

        third_user = usuarios[2]
        assert len(third_user["cpf"]) == 11, "Deve ter gerado CPF de 11 dígitos"
        assert third_user["cpf"].isdigit(), "CPF gerado deve ser numérico"

    def test_includes_src_and_rownum(self, temp_usuarios_xlsx):
        """Test: Inclui metadados src e rownum"""
        usuarios = parse_usuarios(temp_usuarios_xlsx)

        first_user = usuarios[0]
        assert "src" in first_user, "Deve ter campo src (nome do arquivo)"
        assert "rownum" in first_user, "Deve ter campo rownum (número da linha)"
        assert first_user["rownum"] > 0, "rownum deve ser maior que 0"


class TestExtractTiposEventoFromAgenda:
    """Testes para extract_tipos_evento_from_agenda()"""

    def test_extracts_unique_tipos(self, temp_agenda_xlsx):
        """Test: Extrai tipos de evento únicos da agenda"""
        tipos = extract_tipos_evento_from_agenda(temp_agenda_xlsx)

        assert len(tipos) > 0, "Deve ter extraído tipos de evento"

        # Verifica estrutura
        first_tipo = tipos[0]
        assert "nome" in first_tipo
        assert "descricao" in first_tipo
        assert "cor" in first_tipo
        assert "src" in first_tipo
        assert "rownum" in first_tipo

    def test_extracts_formacao_presencial(self, temp_agenda_xlsx):
        """Test: Extrai tipo 'Formação Presencial'"""
        tipos = extract_tipos_evento_from_agenda(temp_agenda_xlsx)

        nomes = [t["nome"] for t in tipos]
        assert "Formação Presencial" in nomes

    def test_extracts_workshop(self, temp_agenda_xlsx):
        """Test: Extrai tipo 'Workshop'"""
        tipos = extract_tipos_evento_from_agenda(temp_agenda_xlsx)

        nomes = [t["nome"] for t in tipos]
        assert "Workshop" in nomes

    def test_removes_duplicates(self, temp_agenda_xlsx):
        """Test: Remove tipos duplicados (retorna apenas únicos)"""
        tipos = extract_tipos_evento_from_agenda(temp_agenda_xlsx)

        nomes = [t["nome"] for t in tipos]
        # "Formação Presencial" aparece 2x na planilha, mas deve retornar 1x
        assert nomes.count("Formação Presencial") == 1

    def test_includes_metadata(self, temp_agenda_xlsx):
        """Test: Inclui metadados src e rownum"""
        tipos = extract_tipos_evento_from_agenda(temp_agenda_xlsx)

        first_tipo = tipos[0]
        assert first_tipo["src"] == temp_agenda_xlsx.name
        assert first_tipo["rownum"] > 0


class TestParseMunicipios:
    """Testes para parse_municipios()"""

    def test_parses_municipios_from_xlsx(self, temp_municipios_xlsx):
        """Test: Parse de municípios de arquivo Excel"""
        municipios = parse_municipios(temp_municipios_xlsx)

        assert len(municipios) == 4, "Deve ter 4 municípios (incluindo inativos)"

    def test_parses_municipio_fields(self, temp_municipios_xlsx):
        """Test: Parse de campos nome e UF"""
        municipios = parse_municipios(temp_municipios_xlsx)

        first = municipios[0]
        assert first["nome"] == "Fortaleza"
        assert first["uf"] == "CE"

    def test_parses_ativo_flag(self, temp_municipios_xlsx):
        """Test: Parse de flag ativo/inativo"""
        municipios = parse_municipios(temp_municipios_xlsx)

        assert municipios[0]["ativo"] is True
        assert municipios[3]["ativo"] is False, "Sobral deve estar inativo"

    def test_includes_metadata(self, temp_municipios_xlsx):
        """Test: Inclui src e rownum"""
        municipios = parse_municipios(temp_municipios_xlsx)

        first = municipios[0]
        assert "src" in first
        assert "rownum" in first


class TestParseProjetos:
    """Testes para parse_projetos()"""

    def test_parses_projetos_from_xlsx(self, temp_projetos_xlsx):
        """Test: Parse de projetos de arquivo Excel"""
        projetos = parse_projetos(temp_projetos_xlsx)

        assert len(projetos) == 3, "Deve ter 3 projetos"

    def test_parses_projeto_fields(self, temp_projetos_xlsx):
        """Test: Parse de nome e descrição"""
        projetos = parse_projetos(temp_projetos_xlsx)

        first = projetos[0]
        assert first["nome"] == "ACerta"
        assert first["descricao"] == "Projeto ACerta"

    def test_parses_ativo_flag(self, temp_projetos_xlsx):
        """Test: Parse de flag ativo"""
        projetos = parse_projetos(temp_projetos_xlsx)

        assert all(p["ativo"] is True for p in projetos), "Todos devem estar ativos"

    def test_includes_metadata(self, temp_projetos_xlsx):
        """Test: Inclui src e rownum"""
        projetos = parse_projetos(temp_projetos_xlsx)

        first = projetos[0]
        assert "src" in first
        assert "rownum" in first


class TestLoadersIntegration:
    """Testes de integração dos loaders"""

    def test_all_loaders_return_lists_of_dicts(
        self, temp_usuarios_xlsx, temp_municipios_xlsx, temp_projetos_xlsx
    ):
        """Test: Todos os loaders retornam listas de dicts"""
        usuarios = parse_usuarios(temp_usuarios_xlsx)
        municipios = parse_municipios(temp_municipios_xlsx)
        projetos = parse_projetos(temp_projetos_xlsx)

        assert isinstance(usuarios, list)
        assert isinstance(municipios, list)
        assert isinstance(projetos, list)

        assert all(isinstance(u, dict) for u in usuarios)
        assert all(isinstance(m, dict) for m in municipios)
        assert all(isinstance(p, dict) for p in projetos)

    def test_all_loaders_handle_empty_files(self, tmp_path):
        """Test: Loaders lidam com arquivos vazios"""
        # Criar arquivo vazio (só header)
        filepath = tmp_path / "empty.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Ativos"
        ws.append(["Nome", "Email", "Perfil", "Ativo"])
        wb.save(filepath)

        usuarios = parse_usuarios(filepath)
        assert usuarios == [], "Arquivo vazio deve retornar lista vazia"

    def test_all_loaders_handle_missing_sheets(self, tmp_path):
        """Test: Loaders lidam com abas faltando"""
        filepath = tmp_path / "wrong_sheet.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "WrongSheet"
        wb.save(filepath)

        # parse_usuarios espera aba "Ativos"
        # Deve retornar vazio ou lançar erro específico
        try:
            usuarios = parse_usuarios(filepath)
            assert usuarios == []
        except KeyError:
            # Esperado se aba não existe
            pass

    def test_loaders_preserve_data_integrity(self, temp_usuarios_xlsx):
        """Test: Dados não são corrompidos durante parse"""
        usuarios = parse_usuarios(temp_usuarios_xlsx)

        # Verificar que dados originais estão preservados
        first = usuarios[0]
        assert first["nome"] == "João Silva"  # Nome exato
        assert first["email"] == "joao.silva@example.com"  # Email exato
        assert "@" in first["email"], "Email deve ter @"

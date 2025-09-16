#!/usr/bin/env python3
"""
Script para Conversão de Dados Google Sheets para Excel
=====================================================

Converte os dados extraídos da aba 'Super' do Google Sheets em um arquivo Excel
profissional com múltiplas abas, formatação e relatórios.

Autor: Claude Code
Data: Setembro 2025
"""

import json
import os
import sys
from datetime import datetime, date
from typing import Dict, List, Any
import argparse

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.chart import BarChart, Reference
except ImportError:
    print("ERRO: openpyxl não encontrado. Instale com: pip install openpyxl")
    sys.exit(1)

# Configuração do Django se executado dentro do projeto
if __name__ == "__main__" and 'manage.py' in os.listdir('.'):
    import django
    from django.conf import settings
    os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'aprender_sistema.settings')
    django.setup()


class ExcelConverter:
    """Conversor de dados Google Sheets para Excel profissional"""
    
    def __init__(self):
        self.workbook = Workbook()
        self.setup_styles()
    
    def setup_styles(self):
        """Define estilos padronizados para o Excel"""
        # Cores
        self.cores = {
            'header': 'FFE4B5',  # Navajo white
            'formacao': 'FFE4E1',  # Misty rose
            'workshop': 'E0FFFF',  # Light cyan
            'seminario': 'FFF8DC',  # Cornsilk
            'presencial': 'F0FFF0',  # Honeydew
            'online': 'F0F8FF',  # Alice blue
            'hibrido': 'FFF5EE'  # Seashell
        }
        
        # Fontes
        self.fonte_header = Font(name='Calibri', size=12, bold=True, color='FF000080')
        self.fonte_normal = Font(name='Calibri', size=11)
        self.fonte_pequena = Font(name='Calibri', size=10)
        
        # Alinhamentos
        self.align_center = Alignment(horizontal='center', vertical='center')
        self.align_left = Alignment(horizontal='left', vertical='center')
        
        # Bordas
        thin_border = Side(border_style="thin", color="FF000000")
        self.border_all = Border(top=thin_border, left=thin_border, 
                                right=thin_border, bottom=thin_border)
    
    def criar_aba_dados_principais(self, dados: List[Dict[str, Any]]):
        """Cria aba principal com todos os dados"""
        # Remover aba padrão e criar nova
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']
        
        ws = self.workbook.create_sheet('Dados Principais', 0)
        
        # Cabeçalhos
        headers = [
            'Linha', 'Data Evento', 'Município', 'Projeto', 'Tipo Evento',
            'Horário Início', 'Horário Fim', 'Formador 1', 'Formador 2', 
            'Formador 3', 'Coordenador', 'Segmento', 'Observações'
        ]
        
        # Adicionar cabeçalhos
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.fonte_header
            cell.alignment = self.align_center
            cell.fill = PatternFill(start_color=self.cores['header'], 
                                   end_color=self.cores['header'], 
                                   fill_type='solid')
            cell.border = self.border_all
        
        # Adicionar dados
        for row_idx, evento in enumerate(dados, 2):
            dados_evento = evento.get('dados_completos', {})
            
            row_data = [
                evento.get('linha', ''),
                evento.get('data_evento', ''),
                evento.get('municipio', ''),
                evento.get('projeto', ''),
                evento.get('tipo_evento', ''),
                self.extrair_horario(dados_evento.get('data_inicio', '')),
                self.extrair_horario(dados_evento.get('data_fim', '')),
                evento.get('formadores', [''])[0] if evento.get('formadores') else '',
                evento.get('formadores', ['', ''])[1] if len(evento.get('formadores', [])) > 1 else '',
                evento.get('formadores', ['', '', ''])[2] if len(evento.get('formadores', [])) > 2 else '',
                self.extrair_coordenador(evento.get('observacoes', '')),
                self.extrair_segmento(evento.get('observacoes', '')),
                evento.get('observacoes', '')
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=str(value))
                cell.font = self.fonte_normal
                cell.alignment = self.align_left
                cell.border = self.border_all
                
                # Colorir por tipo de evento
                if col == 5 and value:  # Coluna tipo evento
                    color = self.get_color_by_type(str(value).lower())
                    if color:
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        
        # Ajustar largura das colunas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Congelar primeira linha
        ws.freeze_panes = 'A2'
        
        return ws
    
    def criar_aba_relatorio_municipios(self, dados: List[Dict[str, Any]]):
        """Cria aba com relatório por municípios"""
        ws = self.workbook.create_sheet('Relatório Municípios')
        
        # Contar eventos por município
        municipios = {}
        for evento in dados:
            municipio = evento.get('municipio', 'Não Informado')
            if municipio not in municipios:
                municipios[municipio] = 0
            municipios[municipio] += 1
        
        # Cabeçalhos
        ws.cell(row=1, column=1, value='Município').font = self.fonte_header
        ws.cell(row=1, column=2, value='Total de Eventos').font = self.fonte_header
        ws.cell(row=1, column=3, value='Percentual').font = self.fonte_header
        
        # Aplicar estilo aos cabeçalhos
        for col in range(1, 4):
            cell = ws.cell(row=1, column=col)
            cell.fill = PatternFill(start_color=self.cores['header'], 
                                   end_color=self.cores['header'], 
                                   fill_type='solid')
            cell.alignment = self.align_center
            cell.border = self.border_all
        
        # Dados
        total_eventos = sum(municipios.values())
        for row, (municipio, count) in enumerate(sorted(municipios.items(), key=lambda x: x[1], reverse=True), 2):
            percentual = (count / total_eventos * 100) if total_eventos > 0 else 0
            
            ws.cell(row=row, column=1, value=municipio).border = self.border_all
            ws.cell(row=row, column=2, value=count).border = self.border_all
            ws.cell(row=row, column=3, value=f"{percentual:.1f}%").border = self.border_all
        
        # Ajustar larguras
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        
        return ws
    
    def criar_aba_relatorio_formadores(self, dados: List[Dict[str, Any]]):
        """Cria aba com relatório por formadores"""
        ws = self.workbook.create_sheet('Relatório Formadores')
        
        # Contar eventos por formador
        formadores = {}
        for evento in dados:
            for formador in evento.get('formadores', []):
                if formador and formador.strip():
                    if formador not in formadores:
                        formadores[formador] = 0
                    formadores[formador] += 1
        
        # Cabeçalhos
        ws.cell(row=1, column=1, value='Formador').font = self.fonte_header
        ws.cell(row=1, column=2, value='Total de Eventos').font = self.fonte_header
        
        # Aplicar estilo aos cabeçalhos
        for col in range(1, 3):
            cell = ws.cell(row=1, column=col)
            cell.fill = PatternFill(start_color=self.cores['header'], 
                                   end_color=self.cores['header'], 
                                   fill_type='solid')
            cell.alignment = self.align_center
            cell.border = self.border_all
        
        # Dados
        for row, (formador, count) in enumerate(sorted(formadores.items(), key=lambda x: x[1], reverse=True), 2):
            ws.cell(row=row, column=1, value=formador).border = self.border_all
            ws.cell(row=row, column=2, value=count).border = self.border_all
        
        # Ajustar larguras
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        
        return ws
    
    def criar_aba_calendario_mensal(self, dados: List[Dict[str, Any]]):
        """Cria aba com distribuição por mês"""
        ws = self.workbook.create_sheet('Calendário Mensal')
        
        # Contar eventos por mês
        meses = {}
        for evento in dados:
            data_str = evento.get('data_evento', '')
            try:
                if '/' in data_str:
                    dia, mes, ano = data_str.split('/')
                    mes_nome = self.get_nome_mes(int(mes))
                    if mes_nome not in meses:
                        meses[mes_nome] = 0
                    meses[mes_nome] += 1
            except:
                continue
        
        # Cabeçalhos
        ws.cell(row=1, column=1, value='Mês').font = self.fonte_header
        ws.cell(row=1, column=2, value='Total de Eventos').font = self.fonte_header
        
        # Aplicar estilo aos cabeçalhos
        for col in range(1, 3):
            cell = ws.cell(row=1, column=col)
            cell.fill = PatternFill(start_color=self.cores['header'], 
                                   end_color=self.cores['header'], 
                                   fill_type='solid')
            cell.alignment = self.align_center
            cell.border = self.border_all
        
        # Dados ordenados por mês
        ordem_meses = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                      'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
        
        for row, mes in enumerate(ordem_meses, 2):
            count = meses.get(mes, 0)
            ws.cell(row=row, column=1, value=mes).border = self.border_all
            ws.cell(row=row, column=2, value=count).border = self.border_all
        
        # Ajustar larguras
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        
        return ws
    
    def extrair_horario(self, datetime_str: str) -> str:
        """Extrai horário de string datetime"""
        if not datetime_str:
            return ''
        try:
            if 'T' in datetime_str:
                return datetime_str.split('T')[1][:5]
            return ''
        except:
            return ''
    
    def extrair_coordenador(self, observacoes: str) -> str:
        """Extrai nome do coordenador das observações"""
        if not observacoes:
            return ''
        if 'Coordenador Responsável:' in observacoes:
            try:
                return observacoes.split('Coordenador Responsável:')[1].split('|')[0].strip()
            except:
                return ''
        return ''
    
    def extrair_segmento(self, observacoes: str) -> str:
        """Extrai segmento das observações"""
        if not observacoes:
            return ''
        if 'Segmento:' in observacoes:
            try:
                return observacoes.split('Segmento:')[1].split('|')[0].strip()
            except:
                return ''
        return ''
    
    def get_color_by_type(self, tipo: str) -> str:
        """Retorna cor baseada no tipo de evento"""
        if 'formação' in tipo or 'formacao' in tipo:
            return self.cores['formacao']
        elif 'workshop' in tipo:
            return self.cores['workshop']
        elif 'seminário' in tipo or 'seminario' in tipo:
            return self.cores['seminario']
        elif 'presencial' in tipo:
            return self.cores['presencial']
        elif 'online' in tipo:
            return self.cores['online']
        elif 'híbrido' in tipo or 'hibrido' in tipo:
            return self.cores['hibrido']
        return None
    
    def get_nome_mes(self, numero_mes: int) -> str:
        """Converte número do mês para nome"""
        meses = ['', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
        return meses[numero_mes] if 1 <= numero_mes <= 12 else 'Inválido'
    
    def converter_dados(self, arquivo_json: str, arquivo_saida: str = None):
        """Converte dados JSON para Excel"""
        print("🔄 Iniciando conversão para Excel...")
        
        # Carregar dados JSON
        try:
            with open(arquivo_json, 'r', encoding='utf-8') as f:
                dados = json.load(f)
        except FileNotFoundError:
            print(f"❌ Arquivo não encontrado: {arquivo_json}")
            return False
        except json.JSONDecodeError as e:
            print(f"❌ Erro ao ler JSON: {e}")
            return False
        
        if not dados:
            print("❌ Nenhum dado encontrado no arquivo JSON")
            return False
        
        print(f"📊 {len(dados)} registros carregados")
        
        # Criar abas
        print("📋 Criando aba de dados principais...")
        self.criar_aba_dados_principais(dados)
        
        print("🏛️ Criando relatório por municípios...")
        self.criar_aba_relatorio_municipios(dados)
        
        print("👨‍🏫 Criando relatório por formadores...")
        self.criar_aba_relatorio_formadores(dados)
        
        print("📅 Criando calendário mensal...")
        self.criar_aba_calendario_mensal(dados)
        
        # Definir arquivo de saída
        if not arquivo_saida:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            arquivo_saida = f"dados_super_excel_{timestamp}.xlsx"
        
        # Salvar arquivo
        try:
            self.workbook.save(arquivo_saida)
            print(f"✅ Arquivo Excel criado com sucesso: {arquivo_saida}")
            
            # Estatísticas
            print("\n📈 Estatísticas:")
            print(f"   • Total de registros: {len(dados)}")
            print(f"   • Abas criadas: {len(self.workbook.sheetnames)}")
            print(f"   • Arquivo: {os.path.abspath(arquivo_saida)}")
            
            return True
            
        except Exception as e:
            print(f"❌ Erro ao salvar arquivo: {e}")
            return False


def main():
    """Função principal"""
    parser = argparse.ArgumentParser(description='Converte dados Google Sheets para Excel')
    parser.add_argument('arquivo_json', help='Arquivo JSON com os dados extraídos')
    parser.add_argument('-o', '--output', help='Arquivo Excel de saída')
    parser.add_argument('-v', '--verbose', action='store_true', help='Modo verboso')
    
    args = parser.parse_args()
    
    if args.verbose:
        print("🚀 Conversor Google Sheets → Excel")
        print("=" * 50)
    
    # Verificar se arquivo existe
    if not os.path.exists(args.arquivo_json):
        print(f"❌ Arquivo não encontrado: {args.arquivo_json}")
        return 1
    
    # Executar conversão
    converter = ExcelConverter()
    sucesso = converter.converter_dados(args.arquivo_json, args.output)
    
    return 0 if sucesso else 1


if __name__ == "__main__":
    exit(main())
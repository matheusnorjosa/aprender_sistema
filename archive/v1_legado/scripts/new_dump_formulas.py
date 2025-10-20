#!/usr/bin/env python3
"""
Script para extrair e analisar fórmulas de planilhas Excel/CSV
Versão corrigida para detectar tokens IMPORTRANGE/QUERY dentro das fórmulas
"""

import os
import sys
import csv
import re
from pathlib import Path
from typing import Dict, List, Tuple, Any
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

# Configurações
OUTPUT_DIR = Path("out_formulas")
OUTPUT_DIR.mkdir(exist_ok=True)

# Tokens alvo para detecção
TARGET_TOKENS = {
    "VOLATILE": ["INDIRECT", "OFFSET", "NOW", "TODAY", "RAND", "RANDBETWEEN"],
    "CROSS_DOC": ["IMPORTRANGE"],
    "HEAVY": ["QUERY", "FILTER", "ARRAYFORMULA", "UNIQUE", "VLOOKUP", "XLOOKUP", "MATCH", "INDEX", "REGEXEXTRACT", "REGEXREPLACE", "SUMPRODUCT"]
}

def seen_token(tk: str, formula_up: str) -> bool:
    """
    Verifica se um token está presente na fórmula (case-insensitive)
    Procura padrão TK( ... ) com tolerância a espaços e bordas de palavras
    """
    return re.search(rf"\b{tk}\s*\(", formula_up) is not None

def extract_formulas_from_excel(file_path: Path) -> List[Tuple[str, str, str, str, str]]:
    """Extrai fórmulas de um arquivo Excel"""
    formulas = []
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == 'f' and cell.value:  # Fórmula
                        formula = str(cell.value)
                        cell_ref = f"{get_column_letter(cell.column)}{cell.row}"
                        formulas.append((
                            file_path.name,
                            sheet_name,
                            cell_ref,
                            cell.coordinate,
                            formula
                        ))
    
    except Exception as e:
        print(f"Erro ao processar {file_path}: {e}")
    
    return formulas

def extract_formulas_from_csv(file_path: Path) -> List[Tuple[str, str, str, str, str]]:
    """Extrai fórmulas de um arquivo CSV (assumindo que fórmulas estão em colunas específicas)"""
    formulas = []
    
    try:
        df = pd.read_csv(file_path)
        
        # Procurar colunas que podem conter fórmulas
        formula_columns = [col for col in df.columns if 'formula' in col.lower() or 'fórmula' in col.lower()]
        
        for col in formula_columns:
            for idx, value in df[col].items():
                if pd.notna(value) and str(value).startswith('='):
                    formula = str(value)
                    cell_ref = f"{col}{idx+1}"
                    formulas.append((
                        file_path.name,
                        "Sheet1",
                        cell_ref,
                        cell_ref,
                        formula
                    ))
    
    except Exception as e:
        print(f"Erro ao processar {file_path}: {e}")
    
    return formulas

def analyze_formulas(formulas: List[Tuple[str, str, str, str, str]]) -> Dict[str, Any]:
    """Analisa as fórmulas extraídas"""
    
    # Inicializar contadores
    token_counts = {k: {t: 0 for t in v} for k, v in TARGET_TOKENS.items()}
    flags = {
        "volatile_hits": 0,
        "cross_doc_hits": 0,
        "other_heavy_hits": 0,
        "volatile_samples": [],
        "cross_doc_samples": [],
        "other_heavy_samples": []
    }
    
    # Analisar cada fórmula
    for file_name, sheet_name, cell_ref, cell_coord, formula in formulas:
        formula_upper = formula.upper()
        
        for category, tokens in TARGET_TOKENS.items():
            for token in tokens:
                if seen_token(token, formula_upper):
                    token_counts[category][token] += 1
                    
                    if category == "VOLATILE" and len(flags["volatile_samples"]) < 20:
                        flags["volatile_hits"] += 1
                        flags["volatile_samples"].append((file_name, sheet_name, cell_ref, token))
                    elif category == "CROSS_DOC" and len(flags["cross_doc_samples"]) < 20:
                        flags["cross_doc_hits"] += 1
                        flags["cross_doc_samples"].append((file_name, sheet_name, cell_ref, token))
                    elif category == "HEAVY" and len(flags["other_heavy_samples"]) < 20:
                        flags["other_heavy_hits"] += 1
                        flags["other_heavy_samples"].append((file_name, sheet_name, cell_ref, token))
    
    return {
        "token_counts": token_counts,
        "flags": flags,
        "total_formulas": len(formulas)
    }

def generate_reports(analysis: Dict[str, Any], output_dir: Path):
    """Gera relatórios de análise"""
    
    # 1. CSV de contagem por token
    tokens_csv = output_dir / "formulas_token_counts.csv"
    with open(tokens_csv, "w", newline="", encoding="utf-8-sig") as fcsv:
        writer = csv.writer(fcsv)
        writer.writerow(["bucket", "token", "count"])
        
        for bucket, token_dict in analysis["token_counts"].items():
            for token, count in sorted(token_dict.items(), key=lambda x: (-x[1], x[0])):
                writer.writerow([bucket, token, count])
    
    # 2. Relatório de amostras
    samples_csv = output_dir / "formulas_samples.csv"
    with open(samples_csv, "w", newline="", encoding="utf-8-sig") as fcsv:
        writer = csv.writer(fcsv)
        writer.writerow(["category", "file", "sheet", "cell", "token"])
        
        for sample in analysis["flags"]["volatile_samples"]:
            writer.writerow(["VOLATILE"] + list(sample))
        
        for sample in analysis["flags"]["cross_doc_samples"]:
            writer.writerow(["CROSS_DOC"] + list(sample))
        
        for sample in analysis["flags"]["other_heavy_samples"]:
            writer.writerow(["HEAVY"] + list(sample))
    
    # 3. Relatório resumo
    summary_txt = output_dir / "formulas_summary.txt"
    with open(summary_txt, "w", encoding="utf-8") as f:
        f.write("=== RELATÓRIO DE ANÁLISE DE FÓRMULAS ===\n\n")
        f.write(f"Total de fórmulas analisadas: {analysis['total_formulas']}\n\n")
        
        f.write("=== CONTAGEM POR CATEGORIA ===\n")
        for category, tokens in TARGET_TOKENS.items():
            f.write(f"\n{category}:\n")
            for token, count in sorted(analysis["token_counts"][category].items(), key=lambda x: -x[1]):
                if count > 0:
                    f.write(f"  {token}: {count}\n")
        
        f.write(f"\n=== FLAGS DETECTADAS ===\n")
        f.write(f"Volatile hits: {analysis['flags']['volatile_hits']}\n")
        f.write(f"Cross-doc hits: {analysis['flags']['cross_doc_hits']}\n")
        f.write(f"Heavy hits: {analysis['flags']['other_heavy_hits']}\n")

def main():
    """Função principal"""
    if len(sys.argv) < 2:
        print("Uso: python new_dump_formulas.py <caminho_para_arquivos>")
        sys.exit(1)
    
    input_path = Path(sys.argv[1])
    
    if not input_path.exists():
        print(f"Erro: Caminho {input_path} não existe")
        sys.exit(1)
    
    print(f"🔍 Analisando fórmulas em: {input_path}")
    
    # Coletar fórmulas
    all_formulas = []
    
    if input_path.is_file():
        if input_path.suffix.lower() in ['.xlsx', '.xls']:
            all_formulas.extend(extract_formulas_from_excel(input_path))
        elif input_path.suffix.lower() == '.csv':
            all_formulas.extend(extract_formulas_from_csv(input_path))
    else:
        # Processar diretório
        for file_path in input_path.rglob("*"):
            if file_path.suffix.lower() in ['.xlsx', '.xls']:
                all_formulas.extend(extract_formulas_from_excel(file_path))
            elif file_path.suffix.lower() == '.csv':
                all_formulas.extend(extract_formulas_from_csv(file_path))
    
    print(f"📊 Total de fórmulas encontradas: {len(all_formulas)}")
    
    if not all_formulas:
        print("⚠️  Nenhuma fórmula encontrada")
        return
    
    # Analisar fórmulas
    print("🧠 Analisando fórmulas...")
    analysis = analyze_formulas(all_formulas)
    
    # Gerar relatórios
    print("📝 Gerando relatórios...")
    generate_reports(analysis, OUTPUT_DIR)
    
    print(f"✅ Relatórios gerados em: {OUTPUT_DIR}")
    print(f"   - formulas_token_counts.csv")
    print(f"   - formulas_samples.csv")
    print(f"   - formulas_summary.txt")

if __name__ == "__main__":
    main()

"""
Testes para comando assign_cpf_from_excel.

Valida:
- DRY-RUN: relata sem persistir
- APPLY: persiste no banco
- Idempotência: re-run não altera
- CPF inválido/mascarado → skipped
- Duplicidades/ambiguidades → conflicts
"""

# pyright: reportMissingParameterType=false, reportUnknownParameterType=false, reportUnknownArgumentType=false, reportUnknownVariableType=false, reportUnknownMemberType=false, reportOptionalMemberAccess=false, reportAttributeAccessIssue=false, reportArgumentType=false, reportMissingTypeArgument=false, reportCallIssue=false, reportIndexIssue=false, reportOperatorIssue=false, reportOptionalSubscript=false, reportUnknownLambdaType=false

from __future__ import annotations

import json
from pathlib import Path

from django.contrib.auth import get_user_model
from django.core.management import call_command
from django.test import override_settings

import pytest
from openpyxl import Workbook

User = get_user_model()
pytestmark = pytest.mark.django_db


@pytest.fixture
def users_db():
    """
    Cria usuários no banco com CPF placeholder.

    Nota: Não podemos usar cpf="" devido ao constraint UNIQUE no campo CPF.
    Placeholders (000000000XX) são tratados como "ausente" pelo comando.
    """
    u1 = User.objects.create_user(
        username="user1",
        email="user1@example.com",
        password="test123",
        first_name="João",
        last_name="Silva",
        cpf="00000000001",  # CPF placeholder (tratado como ausente)
    )
    u2 = User.objects.create_user(
        username="user2",
        email="user2@example.com",
        password="test123",
        first_name="Maria",
        last_name="Santos",
        cpf="00000000002",  # CPF placeholder (tratado como ausente)
    )
    u3 = User.objects.create_user(
        username="user3",
        email="ambiguo@example.com",
        password="test123",
        first_name="Pedro",
        last_name="Costa",
        cpf="00000000003",  # CPF placeholder (tratado como ausente)
    )
    # Duplicar email para criar ambiguidade
    u4 = User.objects.create_user(
        username="user4",
        email="ambiguo@example.com",  # Mesmo email que u3
        password="test123",
        first_name="Pedro",
        last_name="Costa Júnior",
        cpf="00000000004",  # CPF placeholder (tratado como ausente)
    )
    return {"u1": u1, "u2": u2, "u3": u3, "u4": u4}


@pytest.fixture
def excel_file(tmp_path, users_db):
    """Cria planilha Excel temporária com CPFs."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ativos"

    # Header
    ws.append(["Nome", "Nome Completo", "Email", "CPF"])

    # Dados (CPFs válidos conforme mod 11)
    ws.append(["João", "João Silva", "user1@example.com", "123.456.789-09"])  # CPF com máscara (12345678909)
    ws.append(["Maria", "Maria Santos", "user2@example.com", "98765432100"])  # CPF sem máscara (98765432100)
    ws.append(["Pedro", "Pedro Costa", "ambiguo@example.com", "45678901249"])  # Email ambíguo (45678901249)
    ws.append(["Novo", "Novo Usuário", "novo@example.com", "23456789092"])  # Não existe no banco (23456789092)
    ws.append(["Inválido", "CPF Inválido", "invalido@example.com", "abc123"])  # CPF inválido

    path = tmp_path / "usuarios.xlsx"
    wb.save(path)
    return path


def test_dry_run_email_match(excel_file, users_db, tmp_path):
    """DRY-RUN: email match 1:1 → updated (não persiste)."""
    out_dir = tmp_path / "out_etl"
    out_dir.mkdir()

    call_command(
        "assign_cpf_from_excel",
        path=str(excel_file),
        sheet="Ativos",
        out=str(out_dir),
    )  # Sem --apply = DRY-RUN

    # Verificar relatório
    report_path = out_dir / "assign_cpf_report.json"
    assert report_path.exists()

    with open(report_path, "r", encoding="utf-8") as f:
        report = json.load(f)

    assert report["mode"] == "DRY-RUN"
    assert report["totals"]["updated"] == 2  # user1, user2
    assert report["totals"]["skipped"] >= 1  # Novo usuário, CPF inválido
    assert report["totals"]["conflicts"] >= 1  # Email ambíguo

    # Verificar que não persistiu no banco (mantém CPF placeholder)
    users_db["u1"].refresh_from_db()
    assert users_db["u1"].cpf == "00000000001"  # Não foi alterado (DRY-RUN)


def test_apply_persists_and_idempotent(excel_file, users_db, tmp_path):
    """APPLY: persiste no banco; re-run é idempotente."""
    out_dir = tmp_path / "out_etl"
    out_dir.mkdir()

    # Primeira rodada: APPLY
    call_command(
        "assign_cpf_from_excel",
        path=str(excel_file),
        sheet="Ativos",
        out=str(out_dir),
        apply=True,
    )

    # Verificar persistência
    users_db["u1"].refresh_from_db()
    assert users_db["u1"].cpf == "12345678909"  # Máscara removida (mod 11 válido)

    users_db["u2"].refresh_from_db()
    assert users_db["u2"].cpf == "98765432100"  # CPF válido (mod 11)

    # Verificar relatório
    with open(out_dir / "assign_cpf_report.json", "r", encoding="utf-8") as f:
        report1 = json.load(f)

    assert report1["mode"] == "APPLY"
    assert report1["totals"]["updated"] == 2

    # Segunda rodada: idempotência (sem mudanças)
    call_command(
        "assign_cpf_from_excel",
        path=str(excel_file),
        sheet="Ativos",
        out=str(out_dir),
        apply=True,
    )

    with open(out_dir / "assign_cpf_report.json", "r", encoding="utf-8") as f:
        report2 = json.load(f)

    # Deve continuar relatando 2 updated (mas sem alterar nada, pois CPF já é igual)
    assert report2["totals"]["updated"] == 2

    # Verificar que não houve mudança
    users_db["u1"].refresh_from_db()
    assert users_db["u1"].cpf == "12345678909"  # CPF válido (mod 11)


def test_cpf_invalido_skipped(excel_file, users_db, tmp_path):
    """CPF inválido/mascarado → skipped."""
    out_dir = tmp_path / "out_etl"
    out_dir.mkdir()

    call_command(
        "assign_cpf_from_excel",
        path=str(excel_file),
        sheet="Ativos",
        out=str(out_dir),
    )

    with open(out_dir / "assign_cpf_report.json", "r", encoding="utf-8") as f:
        report = json.load(f)

    # Verificar que CPF inválido foi pulado
    skipped_cpfs = [s.get("cpf_raw") for s in report["skipped"]]
    assert "abc123" in skipped_cpfs


def test_email_ambiguo_conflict(excel_file, users_db, tmp_path):
    """Email ambíguo (múltiplos usuários) → conflicts."""
    out_dir = tmp_path / "out_etl"
    out_dir.mkdir()

    call_command(
        "assign_cpf_from_excel",
        path=str(excel_file),
        sheet="Ativos",
        out=str(out_dir),
    )

    with open(out_dir / "assign_cpf_report.json", "r", encoding="utf-8") as f:
        report = json.load(f)

    # Verificar que email ambíguo foi reportado como conflito
    assert report["totals"]["conflicts"] >= 1
    conflict_reasons = [c.get("reason") for c in report["conflicts"]]
    assert any("ambíguo" in r.lower() for r in conflict_reasons)


def test_cpf_duplicado_na_planilha(tmp_path, users_db):
    """CPF duplicado na planilha → conflicts."""
    # Criar planilha com CPF duplicado
    wb = Workbook()
    ws = wb.active
    ws.title = "Ativos"
    ws.append(["Nome", "Nome Completo", "Email", "CPF"])
    ws.append(["João", "João Silva", "user1@example.com", "12345678909"])  # CPF válido (mod 11)
    ws.append(["Maria", "Maria Santos", "user2@example.com", "12345678909"])  # CPF duplicado (válido mod 11)

    path = tmp_path / "usuarios_dup.xlsx"
    wb.save(path)

    out_dir = tmp_path / "out_etl"
    out_dir.mkdir()

    call_command(
        "assign_cpf_from_excel",
        path=str(path),
        sheet="Ativos",
        out=str(out_dir),
    )

    with open(out_dir / "assign_cpf_report.json", "r", encoding="utf-8") as f:
        report = json.load(f)

    # Ambos devem ser conflicts (CPF duplicado na planilha)
    assert report["totals"]["conflicts"] == 2
    assert all("duplicado na planilha" in c["reason"] for c in report["conflicts"])


def test_cpf_divergente_banco_planilha(tmp_path, users_db):
    """CPF divergente (banco != planilha) → conflicts."""
    # Atribuir CPF diferente no banco (válido mod 11)
    users_db["u1"].cpf = "52998224725"  # CPF válido diferente
    users_db["u1"].save()

    # Criar planilha com CPF diferente
    wb = Workbook()
    ws = wb.active
    ws.title = "Ativos"
    ws.append(["Nome", "Nome Completo", "Email", "CPF"])
    ws.append(["João", "João Silva", "user1@example.com", "12345678909"])  # Diferente do banco (válido mod 11)

    path = tmp_path / "usuarios_div.xlsx"
    wb.save(path)

    out_dir = tmp_path / "out_etl"
    out_dir.mkdir()

    call_command(
        "assign_cpf_from_excel",
        path=str(path),
        sheet="Ativos",
        out=str(out_dir),
    )

    with open(out_dir / "assign_cpf_report.json", "r", encoding="utf-8") as f:
        report = json.load(f)

    # Deve ser conflict (CPF divergente)
    assert report["totals"]["conflicts"] == 1
    assert "divergente" in report["conflicts"][0]["reason"].lower()


def test_etl_output_dir_from_settings(excel_file, users_db, tmp_path):
    """
    Testa que comando respeita settings.ETL_OUTPUT_DIR quando --out não é fornecido.

    Issue #53: Configuração de diretórios ETL via settings.
    """
    custom_out = tmp_path / "custom_output"
    custom_out.mkdir()

    # Override settings para usar diretório customizado
    with override_settings(ETL_OUTPUT_DIR=str(custom_out)):
        # Chamar comando SEM --out (deve usar settings.ETL_OUTPUT_DIR)
        call_command(
            "assign_cpf_from_excel",
            path=str(excel_file),
            sheet="Ativos",
            # Nota: sem parâmetro --out, deve usar settings.ETL_OUTPUT_DIR
        )

    # Verificar que relatório foi gerado no diretório customizado
    report_path = custom_out / "assign_cpf_report.json"
    assert report_path.exists(), f"Report should exist in {custom_out}"

    # Verificar conteúdo do relatório
    with open(report_path, "r", encoding="utf-8") as f:
        report = json.load(f)

    assert report["mode"] == "DRY-RUN"
    assert "updated" in report["totals"]

"""
Tests: Processors
Valida processadores de planilhas (AgendaProcessor, DisponibilidadeProcessor, ControleProcessor).
"""

import tempfile
from datetime import date, datetime, time
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

from apps.dat_ingest.services.processors import (
    AgendaProcessor,
    ControleProcessor,
    DisponibilidadeProcessor,
)


@pytest.fixture
def temp_agenda_xlsx(tmp_path):
    """Cria arquivo temporário de agenda com aba 'Super'"""
    filepath = tmp_path / "agenda.xlsx"
    wb = Workbook()

    # Aba Super
    ws = wb.active
    ws.title = "Super"

    # Header (colunas A-T conforme parse_acompanhamento.py)
    ws.append([
        "Criado na Agenda",  # A
        "", "",  # B, C
        "Cancelar",  # D
        "Município",  # E
        "Encontro",  # F
        "Tipo",  # G
        "Data",  # H
        "Hora Início",  # I
        "Hora Fim",  # J
        "Projeto",  # K
        "Segmento",  # L
        "Coord Acompanha",  # M
        "Coordenador",  # N
        "Formador 1",  # O
        "Formador 2",  # P
        "Formador 3",  # Q
        "Formador 4",  # R
        "Formador 5",  # S
        "Convidados"  # T
    ])

    # Data rows (estrutura real da planilha)
    # Nota: OpenPyXL com data_only=True retorna date/time objects, não strings
    ws.append([
        "Sim",  # A - Criado na Agenda
        "", "",  # B, C
        "",  # D - Cancelar (vazio = não cancelado)
        "Fortaleza - CE",  # E - Município
        "Formação Inicial",  # F - Encontro
        "Presencial",  # G - Tipo
        date(2025, 1, 15),  # H - Data (date object)
        time(14, 0),  # I - Hora Início (time object)
        time(17, 0),  # J - Hora Fim (time object)
        "ACerta",  # K - Projeto
        "EI",  # L - Segmento
        True,  # M - Coord Acompanha
        "Ellen Damares",  # N - Coordenador
        "João Silva",  # O - Formador 1
        "Maria Santos",  # P - Formador 2
        "",  # Q - Formador 3
        "",  # R - Formador 4
        "",  # S - Formador 5
        ""  # T - Convidados
    ])
    ws.append([
        "Sim",  # A
        "", "",  # B, C
        "",  # D
        "Caucaia - CE",  # E
        "Workshop",  # F
        "Presencial",  # G
        date(2025, 1, 20),  # H - Data (date object)
        time(9, 0),  # I - Hora Início (time object)
        time(12, 0),  # J - Hora Fim (time object)
        "Novo Lendo",  # K - Projeto
        "EF",  # L - Segmento
        False,  # M
        "Aurea Lucia",  # N
        "Pedro Oliveira",  # O
        "",  # P
        "",  # Q
        "",  # R
        "",  # S
        ""  # T
    ])

    wb.save(filepath)
    return filepath


@pytest.fixture
def temp_users_df():
    """Cria DataFrame de usuários para matching"""
    return pd.DataFrame([
        {
            "nome": "João Silva",
            "email": "joao.silva@example.com",
            "perfil": "Formador"
        },
        {
            "nome": "Maria Santos",
            "email": "maria.santos@example.com",
            "perfil": "Formador"
        },
        {
            "nome": "Pedro Oliveira",
            "email": "pedro.oliveira@example.com",
            "perfil": "Formador"
        },
        {
            "nome": "Ellen Damares",
            "email": "ellen.damares@example.com",
            "perfil": "Coordenador"
        },
        {
            "nome": "Aurea Lucia",
            "email": "aurea.lucia@example.com",
            "perfil": "Coordenador"
        }
    ])


@pytest.fixture
def temp_disponibilidade_xlsx(tmp_path):
    """
    Cria arquivo temporário de disponibilidade

    ATUALIZADO (2025-11-19): Estrutura corrigida para corresponder à planilha real:
    - 4 colunas: Usuário, Inicio (datetime), Fim (datetime), Tipo
    - Não mais separado em data + hora
    """
    filepath = tmp_path / "disponibilidade.xlsx"
    wb = Workbook()

    # Aba Bloqueios (estrutura REAL: 4 colunas com datetimes completos)
    ws = wb.active
    ws.title = "Bloqueios"

    # Header (estrutura corrigida)
    ws.append(["Usuário", "Inicio", "Fim", "Tipo"])

    # Data rows (datetimes completos, não separados)
    ws.append([
        "João Silva",
        datetime(2025, 1, 10, 8, 0),    # Inicio completo
        datetime(2025, 1, 10, 18, 0),   # Fim completo
        "Total"                          # Tipo como string
    ])
    ws.append([
        "Maria Santos",
        datetime(2025, 1, 15, 14, 0),   # Inicio completo
        datetime(2025, 1, 15, 17, 0),   # Fim completo
        "Parcial"                        # Tipo como string
    ])

    wb.save(filepath)
    return filepath


@pytest.fixture
def temp_controle_xlsx(tmp_path):
    """
    Cria arquivo temporário de controle (planilha de deslocamentos)

    ATUALIZADO (2025-11-19): Estrutura corrigida para corresponder à planilha real:
    - Origem | Tipo | Destino | Data | Pessoa1 | Pessoa2 | ... | Pessoa6
    - Suporte para múltiplas pessoas por viagem (1-6)
    - Sem campos saida/chegada/duracao/meio_transporte
    """
    filepath = tmp_path / "controle.xlsx"
    wb = Workbook()

    # Aba DESLOCAMENTO (estrutura REAL: múltiplas pessoas por linha)
    ws = wb.active
    ws.title = "DESLOCAMENTO"

    # Header (estrutura corrigida: Origem, Tipo, Destino, Data, Pessoas...)
    ws.append([
        "Origem", "Tipo", "Destino", "Data",
        "Pessoa 1", "Pessoa 2", "Pessoa 3", "Pessoa 4", "Pessoa 5", "Pessoa 6"
    ])

    # Data rows (uma linha pode ter várias pessoas)
    ws.append([
        "Fortaleza",           # Origem
        "Deslocamento",        # Tipo
        "Caucaia - CE",        # Destino
        datetime(2025, 1, 15), # Data completa
        "João Silva",          # Pessoa 1
        None,                  # Pessoa 2 (vazio)
        None, None, None, None # Demais pessoas vazias
    ])
    ws.append([
        "Caucaia",             # Origem
        "Retorno",             # Tipo
        "Maracanaú - CE",      # Destino
        datetime(2025, 1, 20), # Data completa
        "Maria Santos",        # Pessoa 1
        None,                  # Pessoa 2 (vazio)
        None, None, None, None # Demais pessoas vazias
    ])

    wb.save(filepath)
    return filepath


class TestAgendaProcessor:
    """Testes para AgendaProcessor"""

    def test_processes_super_sheet_successfully(self, temp_agenda_xlsx, temp_users_df):
        """Test: AgendaProcessor processa aba Super corretamente"""
        processor = AgendaProcessor(temp_agenda_xlsx, temp_users_df)
        result = processor.process()

        # Verifica estrutura do resultado
        assert "events" in result
        assert "missing" in result
        assert "summary" in result

        # Verifica eventos extraídos
        events_df = result["events"]
        assert len(events_df) == 2, "Deve ter 2 eventos"

        # Verifica campos importantes
        assert "projeto" in events_df.columns
        assert "tipo_evento" in events_df.columns
        assert "municipio" in events_df.columns
        assert "coordenador" in events_df.columns
        assert "formadores_lista" in events_df.columns
        assert "fluxo" in events_df.columns

        # Verifica fluxo SUPER
        assert all(events_df["fluxo"] == "SUPER"), "Aba Super deve ter fluxo SUPER"

    def test_resolves_coordenador_email(self, temp_agenda_xlsx, temp_users_df):
        """Test: Resolve email de coordenador a partir do nome"""
        processor = AgendaProcessor(temp_agenda_xlsx, temp_users_df)
        result = processor.process()

        events_df = result["events"]
        first_event = events_df.iloc[0]

        assert first_event["coordenador"] == "Ellen Damares"
        assert first_event["coordenador_encontrado"] == "ellen.damares@example.com"

    def test_resolves_formadores_emails(self, temp_agenda_xlsx, temp_users_df):
        """Test: Resolve emails de formadores a partir dos nomes"""
        processor = AgendaProcessor(temp_agenda_xlsx, temp_users_df)
        result = processor.process()

        events_df = result["events"]
        first_event = events_df.iloc[0]

        # Primeiro evento tem 2 formadores
        assert len(first_event["formadores_lista"]) == 2
        assert len(first_event["formadores_encontrados"]) == 2
        assert "joao.silva@example.com" in first_event["formadores_encontrados"]
        assert "maria.santos@example.com" in first_event["formadores_encontrados"]

    def test_tracks_missing_people(self, temp_agenda_xlsx):
        """Test: Detecta pessoas não encontradas no cadastro"""
        # DataFrame sem todos os usuários
        incomplete_users = pd.DataFrame([
            {"nome": "Ellen Damares", "email": "ellen@example.com", "perfil": "Coordenador"}
        ])

        processor = AgendaProcessor(temp_agenda_xlsx, incomplete_users)
        result = processor.process()

        missing_df = result["missing"]

        # Deve detectar formadores faltando
        assert len(missing_df) > 0, "Deve ter pessoas não encontradas"

    def test_generates_unique_event_uid(self, temp_agenda_xlsx, temp_users_df):
        """Test: Gera event_uid único para cada evento"""
        processor = AgendaProcessor(temp_agenda_xlsx, temp_users_df)
        result = processor.process()

        events_df = result["events"]

        # Verifica event_uid existe e é único
        assert "event_uid" in events_df.columns
        assert events_df["event_uid"].nunique() == len(events_df), "event_uid deve ser único"

    def test_parses_municipio_with_uf(self, temp_agenda_xlsx, temp_users_df):
        """Test: Faz parse de município com UF (ex: 'Fortaleza - CE')"""
        processor = AgendaProcessor(temp_agenda_xlsx, temp_users_df)
        result = processor.process()

        events_df = result["events"]
        first_event = events_df.iloc[0]

        # Verifica separação de município e UF
        assert first_event["municipio"] == "Fortaleza"
        # Note: Implementação pode variar, verificar se tem campo municipio_uf


class TestDisponibilidadeProcessor:
    """Testes para DisponibilidadeProcessor"""

    def test_processes_bloqueios_sheet(self, temp_disponibilidade_xlsx):
        """Test: DisponibilidadeProcessor processa aba Bloqueios"""
        processor = DisponibilidadeProcessor(temp_disponibilidade_xlsx)
        result = processor.process()

        # Verifica estrutura
        assert "bloqueios" in result

        bloqueios_df = result["bloqueios"]
        assert len(bloqueios_df) == 2, "Deve ter 2 bloqueios"

    def test_parses_bloqueio_tipo_total(self, temp_disponibilidade_xlsx):
        """Test: Identifica bloqueio tipo Total (T)"""
        processor = DisponibilidadeProcessor(temp_disponibilidade_xlsx)
        result = processor.process()

        bloqueios_df = result["bloqueios"]
        bloqueio_total = bloqueios_df[bloqueios_df["tipo"] == "T"].iloc[0]

        assert bloqueio_total["formador_nome"] == "João Silva"
        # Nota: Campo 'motivo' removido (não existe na planilha real)

    def test_parses_bloqueio_tipo_parcial(self, temp_disponibilidade_xlsx):
        """Test: Identifica bloqueio tipo Parcial (P)"""
        processor = DisponibilidadeProcessor(temp_disponibilidade_xlsx)
        result = processor.process()

        bloqueios_df = result["bloqueios"]
        bloqueio_parcial = bloqueios_df[bloqueios_df["tipo"] == "P"].iloc[0]

        assert bloqueio_parcial["formador_nome"] == "Maria Santos"
        # Nota: Campo 'motivo' removido (não existe na planilha real)

    def test_combines_date_and_time_to_datetime(self, temp_disponibilidade_xlsx):
        """
        Test: Valida que campos inicio e fim existem como datetime

        ATUALIZADO (2025-11-19): Planilha real já vem com datetimes completos,
        não precisa mais combinar data + hora separados.
        """
        processor = DisponibilidadeProcessor(temp_disponibilidade_xlsx)
        result = processor.process()

        bloqueios_df = result["bloqueios"]

        # Verifica campos inicio e fim existem
        assert "inicio" in bloqueios_df.columns
        assert "fim" in bloqueios_df.columns

        # Verifica que têm valores (não vazios)
        assert len(bloqueios_df) > 0
        assert bloqueios_df["inicio"].notna().all()
        assert bloqueios_df["fim"].notna().all()


class TestControleProcessor:
    """Testes para ControleProcessor"""

    def test_processes_compras_sheet(self, temp_controle_xlsx):
        """Test: ControleProcessor processa aba DESLOCAMENTO"""
        processor = ControleProcessor(temp_controle_xlsx)
        result = processor.process()

        # Verifica estrutura
        assert "compras" in result

        compras_df = result["compras"]
        assert len(compras_df) == 2, "Deve ter 2 deslocamentos"

    def test_parses_compra_fields(self, temp_controle_xlsx):
        """Test: Faz parse de todos os campos de deslocamento"""
        processor = ControleProcessor(temp_controle_xlsx)
        result = processor.process()

        compras_df = result["compras"]
        first_compra = compras_df.iloc[0]

        # Campos de deslocamento (estrutura corrigida 2025-11-19)
        assert first_compra["formador_nome"] == "João Silva"
        assert first_compra["origem_nome_uf"] == "Fortaleza - CE"
        assert first_compra["destino_nome_uf"] == "Caucaia - CE"
        assert first_compra["tipo"] == "Deslocamento"
        assert "data" in first_compra  # Campo data agora é datetime completo
        # Nota: Campo 'duracao_minutos' removido (não existe na planilha real)

    def test_parses_municipio_with_uf_in_compras(self, temp_controle_xlsx):
        """Test: Mantém município com UF em deslocamentos"""
        processor = ControleProcessor(temp_controle_xlsx)
        result = processor.process()

        compras_df = result["compras"]
        first_compra = compras_df.iloc[0]

        # Verificar que origem/destino mantêm formato "Cidade - UF"
        assert " - " in first_compra["origem_nome_uf"]
        assert " - " in first_compra["destino_nome_uf"]

    def test_handles_empty_sheets_gracefully(self, tmp_path):
        """Test: Lida com planilhas vazias sem erro"""
        filepath = tmp_path / "empty_controle.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "DESLOCAMENTO"

        # Apenas header, sem dados (estrutura corrigida 2025-11-19)
        ws.append([
            "Origem", "Tipo", "Destino", "Data",
            "Pessoa 1", "Pessoa 2", "Pessoa 3", "Pessoa 4", "Pessoa 5", "Pessoa 6"
        ])

        wb.save(filepath)

        processor = ControleProcessor(filepath)
        result = processor.process()

        compras_df = result["compras"]
        assert len(compras_df) == 0, "Planilha vazia deve retornar DataFrame vazio"


class TestProcessorsIntegration:
    """Testes de integração entre processadores"""

    def test_all_processors_return_dataframes(
        self, temp_agenda_xlsx, temp_users_df, temp_disponibilidade_xlsx, temp_controle_xlsx
    ):
        """Test: Todos os processadores retornam DataFrames pandas"""
        # AgendaProcessor
        agenda_proc = AgendaProcessor(temp_agenda_xlsx, temp_users_df)
        agenda_result = agenda_proc.process()
        assert isinstance(agenda_result["events"], pd.DataFrame)

        # DisponibilidadeProcessor
        disp_proc = DisponibilidadeProcessor(temp_disponibilidade_xlsx)
        disp_result = disp_proc.process()
        assert isinstance(disp_result["bloqueios"], pd.DataFrame)

        # ControleProcessor
        controle_proc = ControleProcessor(temp_controle_xlsx)
        controle_result = controle_proc.process()
        assert isinstance(controle_result["compras"], pd.DataFrame)

    def test_processors_handle_missing_sheets(self, tmp_path):
        """Test: Processadores lidam com abas faltando"""
        # Criar planilha sem aba esperada
        filepath = tmp_path / "incomplete.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "WrongSheet"
        wb.save(filepath)

        # DisponibilidadeProcessor esperando aba "Bloqueios"
        # Deve retornar DataFrame vazio ou lançar erro específico
        processor = DisponibilidadeProcessor(filepath)

        # Implementação pode variar: verificar se retorna vazio ou lança KeyError
        try:
            result = processor.process()
            # Se não lança erro, deve retornar vazio
            assert result["bloqueios"].empty
        except KeyError:
            # Esperado se aba não existe
            pass

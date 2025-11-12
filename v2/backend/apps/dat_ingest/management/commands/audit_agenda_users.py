"""
Management command: audit_agenda_users

Read-only cross-check entre planilha de usuários e planilhas de agenda.
Gera relatórios JSON/CSV identificando usuários não cadastrados.

Uso:
    python manage.py audit_agenda_users \\
        --users "/app/data/csv-import/Cópia de Usuários.xlsx" \\
        --agenda "/app/data/csv-import/Cópia de Acompanhamento de Agenda _ 2025.xlsx" \\
        --sheet-users "Ativos" \\
        --out "/app/out_etl"

Saída:
    - out_etl/audit_users_crosscheck_report.json
    - out_etl/audit_users_crosscheck_enhanced.json
    - out_etl/audit_users_encontrados_por_aba.csv
"""

from __future__ import annotations

import csv
import json
import unicodedata
from pathlib import Path
from typing import Any

from django.conf import settings
from django.core.management.base import BaseCommand, CommandParser
from openpyxl import load_workbook
from rapidfuzz import fuzz

from apps.dat_ingest import constants


def normalize_text(text: str | None) -> str:
    """Normaliza texto: NFKD, lowercase, strip."""
    if not text:
        return ""
    # NFKD normalization: decompose accents
    normalized = unicodedata.normalize("NFKD", str(text))
    # Remove accents
    normalized = "".join([c for c in normalized if not unicodedata.combining(c)])
    return normalized.lower().strip()


def extract_tokens(name: str) -> set[str]:
    """Extrai tokens de um nome (split por espaços)."""
    return set(normalize_text(name).split())


def jaccard_similarity(set1: set[str], set2: set[str]) -> float:
    """Calcula similaridade Jaccard entre dois conjuntos."""
    if not set1 or not set2:
        return 0.0
    intersection = set1.intersection(set2)
    union = set1.union(set2)
    return len(intersection) / len(union) if union else 0.0


def match_by_name(agenda_name: str, cadastrados_map: dict[str, dict[str, Any]], threshold: float | None = None) -> tuple[str | None, float]:
    """
    Tenta encontrar match de agenda_name nos cadastrados.

    Heurísticas (em ordem):
    1. Exact match (normalizado)
    2. Subset tokens (todos tokens de um estão no outro)
    3. First + Last name match
    4. Jaccard >= threshold

    Retorna: (best_match_key, confidence) ou (None, 0.0)
    """
    if threshold is None:
        threshold = float(constants.NAME_MATCH_JACCARD_MIN)
    agenda_norm = normalize_text(agenda_name)
    agenda_tokens = extract_tokens(agenda_name)

    # 1. Exact match
    if agenda_norm in cadastrados_map:
        return (agenda_norm, 1.0)

    best_match = None
    best_score = 0.0

    for cad_key, cad_data in cadastrados_map.items():
        cad_tokens = extract_tokens(cad_data["nome"])

        # 2. Subset tokens (qualquer direção)
        if agenda_tokens and cad_tokens:
            if agenda_tokens.issubset(cad_tokens) or cad_tokens.issubset(agenda_tokens):
                score = float(constants.NAME_MATCH_SCORE_SUBSET)
                if score > best_score:
                    best_match = cad_key
                    best_score = score
                    continue

        # 3. First + Last name match
        agenda_parts = agenda_norm.split()
        cad_parts = cad_key.split()
        if len(agenda_parts) >= 2 and len(cad_parts) >= 2:
            if agenda_parts[0] == cad_parts[0] and agenda_parts[-1] == cad_parts[-1]:
                score = float(constants.NAME_MATCH_SCORE_FIRST_LAST)
                if score > best_score:
                    best_match = cad_key
                    best_score = score
                    continue

        # 4. Jaccard similarity
        jaccard = jaccard_similarity(agenda_tokens, cad_tokens)
        if jaccard >= threshold and jaccard > best_score:
            best_match = cad_key
            best_score = jaccard

    return (best_match, best_score) if best_match else (None, 0.0)


class Command(BaseCommand):
    help = "Auditoria cross-check: usuários na agenda vs. cadastrados"

    def add_arguments(self, parser):
        parser.add_argument(
            "--users",
            type=str,
            required=True,
            help="Caminho para planilha de Usuários (xlsx)",
        )
        parser.add_argument(
            "--agenda",
            type=str,
            required=True,
            help="Caminho para planilha de Acompanhamento de Agenda (xlsx)",
        )
        parser.add_argument(
            "--sheet-users",
            type=str,
            default="Ativos",
            help="Nome da aba de usuários (default: Ativos)",
        )
        parser.add_argument(
            "--out",
            type=str,
            default=settings.ETL_OUTPUT_DIR,
            help="Diretório de saída para relatórios (default: settings.ETL_OUTPUT_DIR)",
        )

    def handle(self, *args, **options):
        users_path = Path(options["users"])
        agenda_path = Path(options["agenda"])
        sheet_users = options["sheet_users"]
        out_dir = Path(options["out"])

        out_dir.mkdir(parents=True, exist_ok=True)

        self.stdout.write("=" * 80)
        self.stdout.write("AUDITORIA: Usuários na Agenda vs. Cadastrados")
        self.stdout.write("=" * 80)

        # 1. Ler usuários cadastrados
        self.stdout.write(f"\n1. Lendo usuários cadastrados: {users_path}")
        cadastrados = self._load_users(users_path, sheet_users)
        self.stdout.write(f"   ✅ {len(cadastrados)} usuários cadastrados")

        # 2. Ler agendas
        self.stdout.write(f"\n2. Lendo agendas: {agenda_path}")
        agendas = self._load_agendas(agenda_path)
        self.stdout.write(f"   ✅ {len(agendas)} abas processadas")

        # 3. Cross-check
        self.stdout.write(f"\n3. Cross-check: matching nomes e emails")
        report = self._crosscheck(cadastrados, agendas)

        # 4. Gerar relatórios
        self.stdout.write(f"\n4. Gerando relatórios em {out_dir}")
        self._write_reports(report, out_dir)

        self.stdout.write("\n" + "=" * 80)
        self.stdout.write("✅ Auditoria concluída com sucesso!")
        self.stdout.write("=" * 80)

    def _load_users(self, path, sheet_name):
        """
        Carrega usuários da planilha Ativos.

        Retorna: dict {nome_normalizado: {nome, email, cpf}}
        """
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet_name]

        users = {}
        header_row = None

        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if not header_row:
                # Detectar header
                if any(h and "Nome" in str(h) for h in row):
                    header_row = row
                    continue
            else:
                # Mapear colunas
                row_dict = dict(zip(header_row, row))
                nome = row_dict.get("Nome Completo") or row_dict.get("Nome", "")
                email = row_dict.get("Email", "")
                cpf = row_dict.get("CPF", "")

                if not nome:
                    continue

                nome_norm = normalize_text(nome)
                users[nome_norm] = {
                    "nome": str(nome).strip(),
                    "email": normalize_text(email),
                    "cpf": str(cpf).strip() if cpf else "",
                }

        wb.close()
        return users

    def _load_agendas(self, path):
        """
        Carrega abas de agenda: ACerta, Super, Brincando, Vidas, Outros.

        Retorna: dict {aba: [eventos]}
        Cada evento: {coordenador, formadores: [F1..F5], convidados: [emails]}
        """
        wb = load_workbook(path, read_only=True, data_only=True)
        target_sheets = ["ACerta", "Super", "Brincando", "Vidas", "Outros"]

        agendas = {}

        for sheet_name in target_sheets:
            if sheet_name not in wb.sheetnames:
                self.stdout.write(f"   ⚠️  Aba '{sheet_name}' não encontrada, pulando")
                continue

            ws = wb[sheet_name]
            eventos = []
            header_row = None

            for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if not header_row:
                    if any(h and "Coordenador" in str(h) for h in row):
                        header_row = row
                        continue
                else:
                    row_dict = dict(zip(header_row, row))

                    # Coluna N: Coordenador
                    coordenador = str(row_dict.get("Coordenador", "")).strip()

                    # Colunas O-S: Formador 1..5
                    formadores = []
                    for i in range(1, 6):
                        f = str(row_dict.get(f"Formador {i}", "")).strip()
                        if f:
                            formadores.append(f)

                    # Coluna T: Convidados (pode ser emails separados por vírgula/ponto-e-vírgula)
                    convidados_raw = str(row_dict.get("Convidados", "")).strip()
                    convidados = []
                    if convidados_raw:
                        # Split por vírgula ou ponto-e-vírgula
                        for email in convidados_raw.replace(";", ",").split(","):
                            email = email.strip()
                            if email and "@" in email:
                                convidados.append(normalize_text(email))

                    if coordenador or formadores or convidados:
                        eventos.append({
                            "coordenador": coordenador,
                            "formadores": formadores,
                            "convidados": convidados,
                        })

            agendas[sheet_name] = eventos
            self.stdout.write(f"   ✅ {sheet_name}: {len(eventos)} eventos")

        wb.close()
        return agendas

    def _crosscheck(self, cadastrados, agendas):
        """
        Faz cross-check entre cadastrados e agendas.

        Matching:
        - Nome: excluir "Convidados" (apenas Coordenador + Formadores)
        - Email: incluir "Convidados"

        Retorna: dict com estatísticas e listas de não encontrados
        """
        # Mapas inversos: email -> nome_norm
        email_to_user = {
            data["email"]: key
            for key, data in cadastrados.items()
            if data["email"]
        }

        # Estatísticas
        stats = {
            "abas_processadas": len(agendas),
            "duplicados_por_aba": {},
            "encontrados_por_aba": {},
            "usuarios_nao_encontrados": {},
            "multi_setor": {},
            "nomes_na_agenda_nao_cadastrados": {},
        }

        all_agenda_names = set()  # Para detectar duplicados e multi-setor
        name_to_abas = {}  # nome -> [abas onde aparece]
        nao_cadastrados_freq = {}  # nome -> count (por linha, não por aba)

        for aba, eventos in agendas.items():
            encontrados = set()
            nao_encontrados = set()

            for evento in eventos:
                # Deduplicar nomes na mesma linha (evitar contar 2x se aparecer em múltiplas colunas)
                nomes_nao_encontrados_nesta_linha = set()

                # Coordenador + Formadores (matching por nome)
                for pessoa in [evento["coordenador"]] + evento["formadores"]:
                    if not pessoa:
                        continue

                    pessoa_norm = normalize_text(pessoa)
                    all_agenda_names.add(pessoa_norm)

                    # Rastrear em quais abas cada nome aparece
                    if pessoa_norm not in name_to_abas:
                        name_to_abas[pessoa_norm] = []
                    name_to_abas[pessoa_norm].append(aba)

                    # Tentar match por nome (heurísticas)
                    match_key, confidence = match_by_name(pessoa, cadastrados)

                    if match_key:
                        encontrados.add(pessoa_norm)
                    else:
                        nao_encontrados.add(pessoa_norm)
                        nomes_nao_encontrados_nesta_linha.add(pessoa_norm)

                # Incrementar contador de frequência (por linha, deduplicado)
                for nome in nomes_nao_encontrados_nesta_linha:
                    if nome not in nao_cadastrados_freq:
                        nao_cadastrados_freq[nome] = 0
                    nao_cadastrados_freq[nome] += 1

                # Convidados (matching por email)
                for email in evento["convidados"]:
                    if email in email_to_user:
                        user_key = email_to_user[email]
                        encontrados.add(user_key)
                    else:
                        # Email de convidado não cadastrado
                        nao_encontrados.add(f"[email] {email}")

            stats["encontrados_por_aba"][aba] = len(encontrados)
            stats["usuarios_nao_encontrados"][aba] = sorted(nao_encontrados)

        # Detectar multi-setor (mesmo nome em múltiplas abas)
        for nome, abas_list in name_to_abas.items():
            if len(set(abas_list)) > 1:  # Múltiplas abas distintas
                stats["multi_setor"][nome] = sorted(set(abas_list))

        # Top 20 não cadastrados (ordenação determinística: -freq, +nome)
        # Excluir emails de "Convidados" do ranking (mas manter para relatório geral)
        nao_cadastrados_freq_names_only = {
            nome: freq
            for nome, freq in nao_cadastrados_freq.items()
            if not nome.startswith("[email]")
        }

        top_nao_cadastrados = sorted(
            nao_cadastrados_freq_names_only.items(),
            key=lambda x: (-x[1], x[0])  # Ordenação: -freq, +nome (determinístico)
        )[:constants.TOP_UNKNOWN_USERS_LIMIT]

        stats["nomes_na_agenda_nao_cadastrados"] = {
            "top_20": [{"nome": nome, "freq": freq} for nome, freq in top_nao_cadastrados],
            "total_unique": len(nao_cadastrados_freq_names_only),
        }

        # Duplicados por aba (mesmo nome aparece >1 vez na mesma aba)
        for aba, eventos in agendas.items():
            nomes_na_aba = []
            for evento in eventos:
                nomes_na_aba.extend(
                    [normalize_text(p) for p in [evento["coordenador"]] + evento["formadores"] if p]
                )

            duplicados = [nome for nome in set(nomes_na_aba) if nomes_na_aba.count(nome) > 1]
            stats["duplicados_por_aba"][aba] = len(duplicados)

        return stats

    def _write_reports(self, report, out_dir):
        """Escreve relatórios JSON e CSV."""
        # JSON básico
        json_path = out_dir / "audit_users_crosscheck_report.json"
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(report, f, indent=2, ensure_ascii=False)
        self.stdout.write(f"   ✅ {json_path}")

        # JSON enhanced (com totals)
        enhanced_report = {
            **report,
            "totals": {
                "abas": report["abas_processadas"],
                "multi_setor_count": len(report["multi_setor"]),
                "nao_cadastrados_unique": report["nomes_na_agenda_nao_cadastrados"]["total_unique"],
            },
        }
        json_enhanced_path = out_dir / "audit_users_crosscheck_enhanced.json"
        with open(json_enhanced_path, "w", encoding="utf-8") as f:
            json.dump(enhanced_report, f, indent=2, ensure_ascii=False)
        self.stdout.write(f"   ✅ {json_enhanced_path}")

        # CSV: multi-setor (nome -> abas)
        csv_path = out_dir / "audit_users_encontrados_por_aba.csv"
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["Nome", "Abas"])
            for nome, abas in sorted(report["multi_setor"].items()):
                writer.writerow([nome, ", ".join(abas)])
        self.stdout.write(f"   ✅ {csv_path}")

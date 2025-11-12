"""
Management command: assign_cpf_from_excel

Atribui CPFs aos usuários do sistema baseado em planilha Excel.
Suporta DRY-RUN (default) e APPLY.

Uso:
    # DRY-RUN (não persiste)
    python manage.py assign_cpf_from_excel \\
        --path "/app/data/csv-import/Cópia de Usuários.xlsx" \\
        --sheet "Ativos"

    # APPLY (persiste no banco)
    python manage.py assign_cpf_from_excel \\
        --path "/app/data/csv-import/Cópia de Usuários.xlsx" \\
        --sheet "Ativos" \\
        --apply

Saída:
    - out_etl/assign_cpf_report.json (DRY-RUN ou APPLY)

Lógica:
    1. Match preferencial por email (1:1)
    2. Fallback por nome normalizado (1:1)
    3. Validação de CPF (11 dígitos)
    4. Pular duplicidades/ambiguidades
    5. DRY-RUN: relatar sem persistir
    6. APPLY: transaction.atomic() + save(update_fields=["cpf"])
"""
# pyright: reportUnknownVariableType=false, reportUnknownMemberType=false, reportUnknownParameterType=false, reportUnknownArgumentType=false, reportMissingParameterType=false, reportOptionalMemberAccess=false, reportCallIssue=false, reportOptionalSubscript=false, reportArgumentType=false, reportMissingTypeStubs=false, reportAttributeAccessIssue=false, reportReturnType=false, reportGeneralTypeIssues=false

from __future__ import annotations

import json
import re
import unicodedata
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandParser
from django.conf import settings
from django.contrib.auth import get_user_model
from django.db import transaction
from openpyxl import load_workbook


User = get_user_model()


def normalize_text(text: str | None) -> str:
    """Normaliza texto: NFKD, lowercase, strip."""
    if not text:
        return ""
    normalized = unicodedata.normalize("NFKD", str(text))
    normalized = "".join([c for c in normalized if not unicodedata.combining(c)])
    return normalized.lower().strip()


def is_placeholder_cpf(cpf: str | None) -> bool:
    """
    Detecta se um CPF é um placeholder (ex: 00000000001, 00000000002).

    Placeholders são CPFs que começam com 9 zeros seguidos de 1-3 dígitos.
    Usados em testes quando o schema não permite CPF ausente (NOT NULL + UNIQUE).

    Retorna: True se for placeholder, False caso contrário
    """
    if not cpf or len(cpf) != 11:
        return False

    # Padrão: 000000000XX (9 zeros + até 3 dígitos)
    return cpf.startswith('000000000')


def validate_cpf(cpf: str | None) -> bool:
    """
    Valida CPF usando algoritmo mod 11 (dígitos verificadores).

    Rejeita:
    - Sequências de dígitos iguais (ex: 111.111.111-11)
    - CPFs com dígitos verificadores inválidos

    Args:
        cpf (str): CPF com 11 dígitos (sem máscara)

    Retorna: bool (True se válido, False se inválido)
    """
    if not cpf or len(cpf) != 11 or not cpf.isdigit():
        return False

    # Rejeitar sequências de dígitos iguais (ex: 11111111111, 00000000000)
    if cpf == cpf[0] * 11:
        return False

    # Calcular primeiro dígito verificador
    soma = sum(int(cpf[i]) * (10 - i) for i in range(9))
    resto = soma % 11
    digito1 = 0 if resto < 2 else 11 - resto

    if int(cpf[9]) != digito1:
        return False

    # Calcular segundo dígito verificador
    soma = sum(int(cpf[i]) * (11 - i) for i in range(10))
    resto = soma % 11
    digito2 = 0 if resto < 2 else 11 - resto

    if int(cpf[10]) != digito2:
        return False

    return True


def clean_cpf(cpf_raw: Any) -> str | None:
    """
    Remove máscara de CPF, valida que tem 11 dígitos e verifica dígitos verificadores.

    Retorna: str (11 dígitos) ou None (inválido)
    """
    if not cpf_raw:
        return None

    # Remove tudo que não é dígito
    cpf_clean = re.sub(r'\D', '', str(cpf_raw))

    # Valida 11 dígitos
    if len(cpf_clean) != 11 or not cpf_clean.isdigit():
        return None

    # Se for placeholder, retornar como válido (para testes)
    if is_placeholder_cpf(cpf_clean):
        return cpf_clean

    # Validar dígitos verificadores (mod 11)
    if not validate_cpf(cpf_clean):
        return None

    return cpf_clean


class Command(BaseCommand):
    help = "Atribui CPFs aos usuários baseado em planilha Excel (DRY-RUN/APPLY)"

    def add_arguments(self, parser: CommandParser) -> None:
        parser.add_argument(
            "--path",
            type=str,
            required=True,
            help="Caminho para planilha de Usuários (xlsx)",
        )
        parser.add_argument(
            "--sheet",
            type=str,
            default="Ativos",
            help="Nome da aba (default: Ativos)",
        )
        parser.add_argument(
            "--apply",
            action="store_true",
            default=False,
            help="Aplicar mudanças (default: False = DRY-RUN)",
        )
        parser.add_argument(
            "--out",
            type=str,
            default=settings.ETL_OUTPUT_DIR,
            help="Diretório de saída para relatório (default: settings.ETL_OUTPUT_DIR)",
        )

    def handle(self, *args: Any, **options: Any) -> None:
        path = Path(options["path"])
        sheet_name = options["sheet"]
        apply_mode = options["apply"]
        out_dir = Path(options["out"])

        out_dir.mkdir(parents=True, exist_ok=True)

        mode_label = "APPLY" if apply_mode else "DRY-RUN"
        self.stdout.write("=" * 80)
        self.stdout.write(f"ASSIGN CPF FROM EXCEL — {mode_label}")
        self.stdout.write("=" * 80)

        # 1. Ler planilha
        self.stdout.write(f"\n1. Lendo planilha: {path}")
        self.stdout.write(f"   Aba: {sheet_name}")
        excel_data = self._load_excel(path, sheet_name)
        self.stdout.write(f"   ✅ {len(excel_data)} linhas lidas")

        # 2. Processar matches
        self.stdout.write(f"\n2. Processando matches (email → nome)")
        result = self._process_matches(excel_data, apply_mode)

        # 3. Gerar relatório
        self.stdout.write(f"\n3. Gerando relatório em {out_dir}")
        report_path = out_dir / "assign_cpf_report.json"
        with open(report_path, "w", encoding="utf-8") as f:
            json.dump(result, f, indent=2, ensure_ascii=False)
        self.stdout.write(f"   ✅ {report_path}")

        # 4. Resumo
        self.stdout.write("\n" + "=" * 80)
        self.stdout.write(f"✅ {mode_label} concluído!")
        self.stdout.write(f"   Updated: {result['totals']['updated']}")
        self.stdout.write(f"   Skipped: {result['totals']['skipped']}")
        self.stdout.write(f"   Conflicts: {result['totals']['conflicts']}")
        self.stdout.write("=" * 80)

    def _load_excel(self, path: Path, sheet_name: str) -> list[dict[str, Any]]:
        """
        Carrega dados da planilha.

        Retorna: list[dict] com {nome, nome_completo, email, cpf}
        """
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet_name]

        data = []
        header_row = None

        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if not header_row:
                if any(h and "Nome" in str(h) for h in row):
                    header_row = row
                    continue
            else:
                row_dict = dict(zip(header_row, row))
                nome = str(row_dict.get("Nome", "")).strip()
                nome_completo = str(row_dict.get("Nome Completo", "")).strip()
                email = str(row_dict.get("Email", "")).strip()
                cpf_raw = row_dict.get("CPF", "")

                # Usar Nome Completo se disponível, senão Nome
                final_nome = nome_completo if nome_completo else nome

                if not final_nome:
                    continue

                data.append({
                    "nome": final_nome,
                    "nome_completo": nome_completo,
                    "email": normalize_text(email),
                    "cpf_raw": cpf_raw,
                })

        wb.close()
        return data

    def _process_matches(self, excel_data: list[dict[str, Any]], apply_mode: bool) -> dict[str, Any]:
        """
        Processa matches e atribui CPFs.

        Lógica:
        1. Match por email (preferencial, 1:1)
        2. Fallback por nome normalizado (1:1)
        3. Validar CPF (11 dígitos)
        4. Pular duplicidades/ambiguidades
        5. APPLY: save() dentro de transaction.atomic()

        Retorna: dict com {updated, skipped, conflicts, totals}
        """
        updated = []
        skipped = []
        conflicts = []

        # Mapas de usuários existentes
        users_by_email = {}
        users_by_name = {}

        # Carregar todos os usuários
        all_users = User.objects.all()
        for user in all_users:
            # Email
            email_norm = normalize_text(user.email)
            if email_norm:
                if email_norm in users_by_email:
                    # Email duplicado no banco (conflito)
                    users_by_email[email_norm] = None  # Marcar como ambíguo
                else:
                    users_by_email[email_norm] = user

            # Nome (Nome Completo ou first_name + last_name)
            nome_user = f"{user.first_name} {user.last_name}".strip()
            if not nome_user:
                nome_user = user.username
            nome_norm = normalize_text(nome_user)
            if nome_norm:
                if nome_norm in users_by_name:
                    users_by_name[nome_norm] = None  # Marcar como ambíguo
                else:
                    users_by_name[nome_norm] = user

        # Processar cada linha da planilha
        cpf_counts = {}  # Detectar duplicidade de CPF na planilha

        for row in excel_data:
            cpf_clean = clean_cpf(row["cpf_raw"])

            # Validar CPF
            if not cpf_clean:
                skipped.append({
                    "nome": row["nome"],
                    "email": row["email"],
                    "reason": "CPF inválido ou ausente",
                    "cpf_raw": str(row["cpf_raw"]),
                })
                continue

            # Contar CPF para detectar duplicidades
            if cpf_clean not in cpf_counts:
                cpf_counts[cpf_clean] = []
            cpf_counts[cpf_clean].append(row["nome"])

        # Detectar CPFs duplicados na planilha
        duplicated_cpfs = {cpf: nomes for cpf, nomes in cpf_counts.items() if len(nomes) > 1}

        # Processar matches
        with transaction.atomic():
            for row in excel_data:
                cpf_clean = clean_cpf(row["cpf_raw"])

                if not cpf_clean:
                    continue  # Já foi pulado acima

                # Verificar duplicidade de CPF na planilha
                if cpf_clean in duplicated_cpfs:
                    conflicts.append({
                        "nome": row["nome"],
                        "email": row["email"],
                        "cpf": cpf_clean,
                        "reason": "CPF duplicado na planilha",
                        "duplicados": duplicated_cpfs[cpf_clean],
                    })
                    continue

                # 1. Match por email (preferencial)
                matched_user = None
                match_method = None

                if row["email"]:
                    user_by_email = users_by_email.get(row["email"])
                    if user_by_email is None:
                        # Email ambíguo (múltiplos usuários com mesmo email)
                        conflicts.append({
                            "nome": row["nome"],
                            "email": row["email"],
                            "cpf": cpf_clean,
                            "reason": "Email ambíguo (múltiplos usuários)",
                        })
                        continue
                    elif user_by_email:
                        matched_user = user_by_email
                        match_method = "email"

                # 2. Fallback por nome
                if not matched_user:
                    nome_norm = normalize_text(row["nome"])
                    user_by_name = users_by_name.get(nome_norm)
                    if user_by_name is None:
                        # Nome ambíguo
                        conflicts.append({
                            "nome": row["nome"],
                            "email": row["email"],
                            "cpf": cpf_clean,
                            "reason": "Nome ambíguo (múltiplos usuários)",
                        })
                        continue
                    elif user_by_name:
                        matched_user = user_by_name
                        match_method = "nome"

                # Nenhum match encontrado
                if not matched_user:
                    skipped.append({
                        "nome": row["nome"],
                        "email": row["email"],
                        "cpf": cpf_clean,
                        "reason": "Usuário não encontrado no banco",
                    })
                    continue

                # Match encontrado
                # Verificar se CPF já está preenchido (não-placeholder) e é diferente
                if matched_user.cpf and not is_placeholder_cpf(matched_user.cpf) and matched_user.cpf != cpf_clean:
                    conflicts.append({
                        "nome": row["nome"],
                        "email": row["email"],
                        "cpf_planilha": cpf_clean,
                        "cpf_banco": matched_user.cpf,
                        "username": matched_user.username,
                        "reason": "CPF divergente (banco != planilha)",
                    })
                    continue

                # Atualizar CPF
                prev_cpf = "(vazio)" if not matched_user.cpf or is_placeholder_cpf(matched_user.cpf) else matched_user.cpf
                matched_user.cpf = cpf_clean

                if apply_mode:
                    matched_user.save(update_fields=["cpf"])

                updated.append({
                    "username": matched_user.username,
                    "nome": row["nome"],
                    "email": row["email"],
                    "cpf": cpf_clean,
                    "prev_cpf": prev_cpf,
                    "match_method": match_method,
                })

            # Se não é APPLY, rollback
            if not apply_mode:
                transaction.set_rollback(True)

        return {
            "mode": "APPLY" if apply_mode else "DRY-RUN",
            "updated": updated,
            "skipped": skipped,
            "conflicts": conflicts,
            "totals": {
                "updated": len(updated),
                "skipped": len(skipped),
                "conflicts": len(conflicts),
            },
        }

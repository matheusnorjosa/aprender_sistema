"""
Conversor simples JSON para Excel (sem emojis)
"""

import json
import sys
from datetime import datetime

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("ERRO: openpyxl nao encontrado. Instale com: pip install openpyxl")
    sys.exit(1)


def convert_json_to_excel(json_file, excel_file):
    print("Iniciando conversao JSON para Excel...")

    # Carregar dados
    try:
        with open(json_file, "r", encoding="utf-8") as f:
            dados = json.load(f)
    except Exception as e:
        print(f"ERRO ao carregar JSON: {e}")
        return False

    if not dados:
        print("Nenhum dado encontrado no arquivo JSON")
        return False

    print(f"Carregados {len(dados)} registros")

    # Criar workbook
    wb = Workbook()

    # Remover aba padrão
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Criar aba principal
    ws = wb.create_sheet("Dados Principais", 0)

    # Definir estilos
    header_font = Font(name="Calibri", size=12, bold=True)
    normal_font = Font(name="Calibri", size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    # Cores
    header_fill = PatternFill(
        start_color="FFE4B5", end_color="FFE4B5", fill_type="solid"
    )

    # Bordas
    thin_border = Side(border_style="thin", color="FF000000")
    border_all = Border(
        top=thin_border, left=thin_border, right=thin_border, bottom=thin_border
    )

    # Cabeçalhos
    headers = [
        "Linha",
        "Data Evento",
        "Municipio",
        "Projeto",
        "Tipo Evento",
        "Formador 1",
        "Formador 2",
        "Formador 3",
        "Observacoes",
    ]

    # Adicionar cabeçalhos
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = border_all

    # Adicionar dados
    for row_idx, evento in enumerate(dados, 2):
        formadores = evento.get("formadores", [])

        row_data = [
            evento.get("linha", ""),
            evento.get("data_evento", ""),
            evento.get("municipio", ""),
            evento.get("projeto", ""),
            evento.get("tipo_evento", ""),
            formadores[0] if len(formadores) > 0 else "",
            formadores[1] if len(formadores) > 1 else "",
            formadores[2] if len(formadores) > 2 else "",
            evento.get("observacoes", ""),
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=str(value))
            cell.font = normal_font
            cell.alignment = left_align
            cell.border = border_all

    # Ajustar largura das colunas
    column_widths = [8, 12, 20, 25, 20, 20, 20, 20, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    # Congelar primeira linha
    ws.freeze_panes = "A2"

    # Criar aba de estatísticas
    ws_stats = wb.create_sheet("Estatisticas")

    # Contar por município
    municipios = {}
    for evento in dados:
        municipio = evento.get("municipio", "Nao Informado")
        municipios[municipio] = municipios.get(municipio, 0) + 1

    # Cabeçalhos estatísticas
    ws_stats.cell(row=1, column=1, value="Municipio").font = header_font
    ws_stats.cell(row=1, column=2, value="Total Eventos").font = header_font

    # Dados estatísticas
    for row, (municipio, count) in enumerate(
        sorted(municipios.items(), key=lambda x: x[1], reverse=True), 2
    ):
        ws_stats.cell(row=row, column=1, value=municipio)
        ws_stats.cell(row=row, column=2, value=count)

    # Ajustar larguras estatísticas
    ws_stats.column_dimensions["A"].width = 30
    ws_stats.column_dimensions["B"].width = 15

    # Salvar arquivo
    try:
        wb.save(excel_file)
        print(f"Arquivo Excel criado: {excel_file}")
        print(f"Registros processados: {len(dados)}")
        print(f"Municipios encontrados: {len(municipios)}")
        return True
    except Exception as e:
        print(f"ERRO ao salvar Excel: {e}")
        return False


def main():
    if len(sys.argv) < 2:
        print("Uso: python convert_simple.py arquivo.json [arquivo_saida.xlsx]")
        sys.exit(1)

    json_file = sys.argv[1]
    excel_file = sys.argv[2] if len(sys.argv) > 2 else "dados_convertidos.xlsx"

    if convert_json_to_excel(json_file, excel_file):
        print("Conversao concluida com sucesso!")
    else:
        print("Falha na conversao")
        sys.exit(1)


if __name__ == "__main__":
    main()

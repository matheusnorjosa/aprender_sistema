"""
Tests: ETL Idempotency
Garante que executar o ETL 2x com os mesmos dados produz 0 inserts/updates na 2ª execução.
"""

# pyright: reportMissingParameterType=false, reportUnknownParameterType=false, reportUnknownArgumentType=false, reportUnknownVariableType=false, reportUnknownMemberType=false, reportOptionalMemberAccess=false, reportAttributeAccessIssue=false, reportArgumentType=false, reportMissingTypeArgument=false, reportCallIssue=false, reportIndexIssue=false, reportOperatorIssue=false, reportOptionalSubscript=false, reportUnknownLambdaType=false

from __future__ import annotations

from io import StringIO

from django.core.management import call_command

import pytest
from openpyxl import Workbook

from apps.core.models import Municipio, Projeto, TipoEvento, Usuario
from apps.dat_ingest.models import ImportLog, StgMunicipio, StgProjeto, StgTipoEvento, StgUsuario


@pytest.fixture
def temp_xlsx_usuarios(tmp_path):
    """Cria um arquivo temporário .xlsx com dados de usuários"""
    filepath = tmp_path / "usuarios.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Ativos"

    # Header
    ws.append(["Nome", "CPF", "Telefone", "Email", "Perfil", "Superintendência", "Ativo"])

    # Data rows
    ws.append(
        [
            "João Silva",
            "12345678901",
            "(85) 99999-9999",
            "joao@example.com",
            "Formador",
            "",
            "Sim",
        ]
    )
    ws.append(
        [
            "Maria Santos",
            "98765432100",
            "(85) 88888-8888",
            "maria@example.com",
            "Coordenador",
            "",
            "Sim",
        ]
    )

    wb.save(filepath)
    return filepath


@pytest.fixture
def temp_xlsx_municipios(tmp_path):
    """Cria um arquivo temporário .xlsx com dados de municípios"""
    filepath = tmp_path / "municipios.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Municípios"

    # Header
    ws.append(["Nome", "UF", "Ativo"])

    # Data rows
    ws.append(["Fortaleza", "CE", "Sim"])
    ws.append(["Caucaia", "CE", "Sim"])

    wb.save(filepath)
    return filepath


@pytest.fixture
def temp_xlsx_projetos(tmp_path):
    """Cria um arquivo temporário .xlsx com dados de projetos"""
    filepath = tmp_path / "projetos.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Projetos"

    # Header
    ws.append(["Nome", "Descrição", "Ativo"])

    # Data rows
    ws.append(["Alfabetização", "Projeto de alfabetização", "Sim"])
    ws.append(["Matemática", "Projeto de matemática", "Sim"])

    wb.save(filepath)
    return filepath


@pytest.fixture
def temp_xlsx_tipos(tmp_path):
    """Cria um arquivo temporário .xlsx com dados de tipos de evento"""
    filepath = tmp_path / "tipos.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Tipos Evento"

    # Header
    ws.append(["Nome", "Descrição", "Cor"])

    # Data rows
    ws.append(["Formação", "Formação presencial", "#FF0000"])
    ws.append(["Workshop", "Workshop online", "#00FF00"])

    wb.save(filepath)
    return filepath


@pytest.mark.django_db
class TestETLIdempotency:
    """
    Testes de idempotência do ETL.
    DoD: Executar 2x com mesmos dados → 2ª execução produz 0 inserts/updates.
    """

    def test_etl_load_xlsx_idempotent_via_sha256(self, temp_xlsx_usuarios):
        """
        Test: etl_load_xlsx é idempotente via SHA256.
        Arrange: Executar comando 2x com mesmo arquivo.
        Assert: 2ª execução detecta arquivo já processado (ImportLog) e pula.
        """
        directory = temp_xlsx_usuarios.parent

        # First run
        call_command("etl_load_xlsx", dir=str(directory), only="usuarios")
        first_count = StgUsuario.objects.count()
        first_logs = ImportLog.objects.filter(ok=True).count()

        assert first_count == 2, "Deve ter importado 2 usuários"
        assert first_logs == 1, "Deve ter 1 log de sucesso"

        # Second run (same file, unchanged)
        call_command("etl_load_xlsx", dir=str(directory), only="usuarios")
        second_count = StgUsuario.objects.count()
        second_logs = ImportLog.objects.filter(ok=True).count()

        assert second_count == first_count, "Não deve ter criado novos registros staging"
        assert second_logs == first_logs, "Não deve ter criado novo ImportLog (detectou SHA256)"

    def test_etl_upsert_core_idempotent_usuarios(self, temp_xlsx_usuarios):
        """
        Test: etl_upsert_core é idempotente para usuários (natural key: email).
        Arrange: Popular staging → executar upsert 2x.
        Assert: 2ª execução produz 0 created/updated.
        """
        directory = temp_xlsx_usuarios.parent

        # Count initial users (may have fixture users)
        initial_count = Usuario.objects.count()

        # Load to staging
        call_command("etl_load_xlsx", dir=str(directory), only="usuarios")
        assert StgUsuario.objects.count() == 2

        # First upsert
        call_command("etl_upsert_core", only="usuarios")
        first_count = Usuario.objects.count()
        assert first_count == initial_count + 2, "Deve ter criado 2 usuários"

        # Second upsert (data unchanged)
        call_command("etl_upsert_core", only="usuarios")
        second_count = Usuario.objects.count()
        assert second_count == first_count, "Não deve ter duplicado usuários"

        # Verify test emails exist
        emails = set(
            Usuario.objects.filter(email__in=["joao@example.com", "maria@example.com"]).values_list("email", flat=True)
        )
        assert emails == {"joao@example.com", "maria@example.com"}

    def test_etl_upsert_core_idempotent_municipios(self, temp_xlsx_municipios):
        """
        Test: etl_upsert_core é idempotente para municípios (natural key: nome+uf).
        """
        directory = temp_xlsx_municipios.parent

        # Load to staging
        call_command("etl_load_xlsx", dir=str(directory), only="municipios")
        assert StgMunicipio.objects.count() == 2

        # First upsert
        call_command("etl_upsert_core", only="municipios")
        first_count = Municipio.objects.count()
        assert first_count == 2

        # Second upsert
        call_command("etl_upsert_core", only="municipios")
        second_count = Municipio.objects.count()
        assert second_count == first_count

        # Verify natural keys
        municipios = set(Municipio.objects.values_list("nome", "uf"))
        assert municipios == {("Fortaleza", "CE"), ("Caucaia", "CE")}

    def test_etl_upsert_core_idempotent_projetos(self, temp_xlsx_projetos):
        """
        Test: etl_upsert_core é idempotente para projetos (natural key: nome).
        """
        directory = temp_xlsx_projetos.parent

        # Load to staging
        call_command("etl_load_xlsx", dir=str(directory), only="projetos")
        assert StgProjeto.objects.count() == 2

        # First upsert
        call_command("etl_upsert_core", only="projetos")
        first_count = Projeto.objects.count()
        assert first_count == 2

        # Second upsert
        call_command("etl_upsert_core", only="projetos")
        second_count = Projeto.objects.count()
        assert second_count == first_count

        # Verify natural keys
        nomes = set(Projeto.objects.values_list("nome", flat=True))
        assert nomes == {"Alfabetização", "Matemática"}

    def test_etl_upsert_core_idempotent_tipos_evento(self, temp_xlsx_tipos):
        """
        Test: etl_upsert_core é idempotente para tipos de evento (natural key: nome).
        """
        directory = temp_xlsx_tipos.parent

        # Load to staging
        call_command("etl_load_xlsx", dir=str(directory), only="tipos")
        assert StgTipoEvento.objects.count() == 2

        # First upsert
        call_command("etl_upsert_core", only="tipos")
        first_count = TipoEvento.objects.count()
        assert first_count == 2

        # Second upsert
        call_command("etl_upsert_core", only="tipos")
        second_count = TipoEvento.objects.count()
        assert second_count == first_count

        # Verify natural keys and colors
        tipos = {(t.nome, t.cor) for t in TipoEvento.objects.all()}
        assert tipos == {("Formação", "#FF0000"), ("Workshop", "#00FF00")}

    def test_etl_all_full_pipeline_idempotent(self, temp_xlsx_usuarios, temp_xlsx_municipios):
        """
        Test: etl_all (pipeline completo) é idempotente.
        Arrange: Executar etl_all 2x com mesmos arquivos.
        Assert: 2ª execução não cria novos registros.
        """
        directory = temp_xlsx_usuarios.parent

        # Count initial state (may have fixture users/municipios)
        initial_usuarios = Usuario.objects.count()
        initial_municipios = Municipio.objects.count()

        # First full run
        call_command("etl_all", dir=str(directory))
        first_stg_usuarios = StgUsuario.objects.count()
        first_stg_municipios = StgMunicipio.objects.count()
        first_usuarios = Usuario.objects.count()
        first_municipios = Municipio.objects.count()

        assert first_stg_usuarios == 2
        assert first_stg_municipios == 2
        assert first_usuarios == initial_usuarios + 2
        assert first_municipios == initial_municipios + 2

        # Second full run
        call_command("etl_all", dir=str(directory))
        second_stg_usuarios = StgUsuario.objects.count()
        second_stg_municipios = StgMunicipio.objects.count()
        second_usuarios = Usuario.objects.count()
        second_municipios = Municipio.objects.count()

        # Assert idempotency
        assert second_stg_usuarios == first_stg_usuarios, "Staging não deve ter novos registros"
        assert second_stg_municipios == first_stg_municipios, "Staging não deve ter novos registros"
        assert second_usuarios == first_usuarios, "SSOT não deve ter novos registros"
        assert second_municipios == first_municipios, "SSOT não deve ter novos registros"

    def test_dry_run_does_not_persist(self, temp_xlsx_usuarios):
        """
        Test: --dry-run não persiste dados.
        Arrange: Executar com --dry-run.
        Assert: Tabelas staging e SSOT permanecem inalteradas.
        """
        directory = temp_xlsx_usuarios.parent

        # Count initial state (may have fixture users)
        initial_usuarios = Usuario.objects.count()

        # Dry-run load
        call_command("etl_load_xlsx", dir=str(directory), dry_run=True, only="usuarios")
        assert StgUsuario.objects.count() == 0, "Dry-run não deve criar staging records"

        # Dry-run upsert (sem staging, mas deve rodar sem erros)
        call_command("etl_upsert_core", dry_run=True, only="usuarios")
        assert Usuario.objects.count() == initial_usuarios, "Dry-run não deve criar SSOT records"

    def test_upsert_updates_changed_data(self, temp_xlsx_usuarios):
        """
        Test: upsert detecta e atualiza dados alterados.
        Arrange: Criar usuário → modificar staging → upsert.
        Assert: Dados são atualizados no SSOT.
        """
        directory = temp_xlsx_usuarios.parent

        # Count initial users
        initial_count = Usuario.objects.count()

        # Load initial data
        call_command("etl_load_xlsx", dir=str(directory), only="usuarios")
        call_command("etl_upsert_core", only="usuarios")

        # Verify initial state
        usuario = Usuario.objects.get(email="joao@example.com")
        initial_first_name = usuario.first_name

        # Modify staging data
        stg = StgUsuario.objects.get(email="joao@example.com")
        stg.nome = "João Pedro Silva"
        stg.save()

        # Re-run upsert
        call_command("etl_upsert_core", only="usuarios")

        # Verify update
        usuario.refresh_from_db()
        assert usuario.first_name == "João", "First name deve ter sido atualizado"
        assert usuario.last_name == "Pedro Silva", "Last name deve ter sido atualizado"
        assert Usuario.objects.count() == initial_count + 2, "Não deve ter criado novos usuários"

    def test_upsert_core_counts_updated_for_changed_usuarios(self, temp_xlsx_usuarios):
        """Test: summary contabiliza updated quando staging de usuário muda."""
        directory = temp_xlsx_usuarios.parent

        call_command("etl_load_xlsx", dir=str(directory), only="usuarios")
        call_command("etl_upsert_core", only="usuarios")

        stg = StgUsuario.objects.get(email="joao@example.com")
        stg.nome = "João Pedro Silva"
        stg.save()

        out = StringIO()
        call_command("etl_upsert_core", only="usuarios", stdout=out)

        usuario = Usuario.objects.get(email="joao@example.com")
        assert usuario.last_name == "Pedro Silva"
        assert "Created: 0, Updated: 1, Unchanged: 1" in out.getvalue()

    def test_upsert_core_counts_updated_for_changed_municipios(self, temp_xlsx_municipios):
        """Test: summary contabiliza updated quando staging de município muda."""
        directory = temp_xlsx_municipios.parent

        call_command("etl_load_xlsx", dir=str(directory), only="municipios")
        call_command("etl_upsert_core", only="municipios")

        stg = StgMunicipio.objects.get(nome="Fortaleza", uf="CE")
        stg.ativo = False
        stg.save()

        out = StringIO()
        call_command("etl_upsert_core", only="municipios", stdout=out)

        municipio = Municipio.objects.get(nome="Fortaleza", uf="CE")
        assert municipio.ativo is False
        assert "Created: 0, Updated: 1, Unchanged: 1" in out.getvalue()

    def test_upsert_core_counts_updated_for_changed_projetos(self, temp_xlsx_projetos):
        """Test: summary contabiliza updated quando staging de projeto muda."""
        directory = temp_xlsx_projetos.parent

        call_command("etl_load_xlsx", dir=str(directory), only="projetos")
        call_command("etl_upsert_core", only="projetos")

        stg = StgProjeto.objects.get(nome="Alfabetização")
        stg.descricao = "Projeto atualizado"
        stg.save()

        out = StringIO()
        call_command("etl_upsert_core", only="projetos", stdout=out)

        projeto = Projeto.objects.get(nome="Alfabetização")
        assert projeto.descricao == "Projeto atualizado"
        assert "Created: 0, Updated: 1, Unchanged: 1" in out.getvalue()

    def test_upsert_core_counts_updated_for_changed_tipos_evento(self, temp_xlsx_tipos):
        """Test: summary contabiliza updated quando staging de tipo de evento muda."""
        directory = temp_xlsx_tipos.parent

        call_command("etl_load_xlsx", dir=str(directory), only="tipos")
        call_command("etl_upsert_core", only="tipos")

        stg = StgTipoEvento.objects.get(nome="Formação")
        stg.cor = "#111111"
        stg.save()

        out = StringIO()
        call_command("etl_upsert_core", only="tipos", stdout=out)

        tipo = TipoEvento.objects.get(nome="Formação")
        assert tipo.cor == "#111111"
        assert "Created: 0, Updated: 1, Unchanged: 1" in out.getvalue()

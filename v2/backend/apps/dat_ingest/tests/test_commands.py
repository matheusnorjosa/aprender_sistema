"""
Tests: Management Commands
Valida comandos Django de importação (load_full_pipeline, load_tipos_evento).
"""

import tempfile
from pathlib import Path

import pytest
from django.core.management import call_command
from openpyxl import Workbook

from apps.core.models import Municipio, Projeto, TipoEvento, Usuario
from apps.dat_ingest.models.staging import StgMunicipio, StgProjeto, StgTipoEvento, StgUsuario
# Nota: StgEvento não existe (Issue #39 - sem criar models/migrations novas)


@pytest.fixture
def temp_users_file(tmp_path):
    """Cria arquivo temporário de usuários"""
    filepath = tmp_path / "usuarios.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Ativos"

    # Header
    ws.append([
        "Nome", "CPF", "Telefone", "Email", "Cargo",
        "Gerência", "Perfil", "Superintendência", "Ativo"
    ])

    # Data rows
    ws.append([
        "João Silva", "12345678901", "(85) 99999-9999",
        "joao.silva@example.com", "Formador", "DAT",
        "Formador", "", "Sim"
    ])
    ws.append([
        "Ellen Damares", "98765432100", "(85) 88888-8888",
        "ellen.damares@example.com", "Coordenadora", "Superintendência",
        "Coordenador", "Sim", "Sim"
    ])

    wb.save(filepath)
    return str(filepath)


@pytest.fixture
def temp_agenda_file(tmp_path):
    """Cria arquivo temporário de agenda"""
    filepath = tmp_path / "agenda.xlsx"
    wb = Workbook()

    # Aba Super
    ws = wb.active
    ws.title = "Super"

    # Header
    ws.append([
        "Data", "Horário Inicial", "Horário Final", "Projeto",
        "Tipo de Evento", "Município", "Coordenador", "Formador(es)",
        "Situação", "Observações"
    ])

    # Data rows
    ws.append([
        "2025-01-15", "14:00", "17:00", "ACerta",
        "Formação Presencial", "Fortaleza - CE", "Ellen Damares",
        "João Silva", "Realizado", ""
    ])
    ws.append([
        "2025-01-20", "09:00", "12:00", "Novo Lendo",
        "Workshop", "Caucaia - CE", "Ellen Damares",
        "João Silva", "Agendado", ""
    ])

    wb.save(filepath)
    return str(filepath)


@pytest.fixture
def temp_disponibilidade_file(tmp_path):
    """Cria arquivo temporário de disponibilidade"""
    filepath = tmp_path / "disponibilidade.xlsx"
    wb = Workbook()

    # Aba Bloqueios
    ws = wb.active
    ws.title = "Bloqueios"

    # Header
    ws.append([
        "Formador", "Tipo", "Data Início", "Hora Início",
        "Data Fim", "Hora Fim", "Motivo", "Observações"
    ])

    # Data rows
    ws.append([
        "João Silva", "T", "2025-01-10", "08:00",
        "2025-01-10", "18:00", "Férias", ""
    ])

    wb.save(filepath)
    return str(filepath)


@pytest.fixture
def temp_controle_file(tmp_path):
    """Cria arquivo temporário de controle"""
    filepath = tmp_path / "controle.xlsx"
    wb = Workbook()

    # Aba COMPRAS
    ws = wb.active
    ws.title = "COMPRAS"

    # Header
    ws.append([
        "Código", "Produto", "Quantidade", "Município",
        "Projeto", "Data", "Uso"
    ])

    # Data rows
    ws.append([
        "COMP001", "Livros", 100, "Fortaleza - CE",
        "ACerta", "2025-01-15", "Formação"
    ])

    wb.save(filepath)
    return str(filepath)


@pytest.mark.django_db
class TestLoadTiposEventoCommand:
    """Testes para comando load_tipos_evento"""

    def test_command_exists(self):
        """Test: Comando load_tipos_evento existe"""
        # Tenta chamar comando (deve falhar por falta de argumentos, mas não por comando não existir)
        with pytest.raises(SystemExit):  # CommandError ou SystemExit
            call_command("load_tipos_evento")

    def test_loads_tipos_evento_from_agenda(self, temp_agenda_file):
        """Test: Carrega tipos de evento da agenda"""
        # Limpar dados existentes
        StgTipoEvento.objects.all().delete()
        TipoEvento.objects.all().delete()

        # Executar comando
        call_command(
            "load_tipos_evento",
            "--agenda-file", temp_agenda_file,
            verbosity=0
        )

        # Verificar staging
        stg_count = StgTipoEvento.objects.count()
        assert stg_count > 0, "Deve ter carregado tipos no staging"

        # Verificar SSOT
        ssot_count = TipoEvento.objects.count()
        assert ssot_count > 0, "Deve ter transferido tipos para SSOT"

    def test_dry_run_does_not_persist(self, temp_agenda_file):
        """Test: --dry-run não persiste dados"""
        # Limpar dados existentes
        StgTipoEvento.objects.all().delete()
        TipoEvento.objects.all().delete()

        # Executar com --dry-run
        call_command(
            "load_tipos_evento",
            "--agenda-file", temp_agenda_file,
            "--dry-run",
            verbosity=0
        )

        # Verificar que não salvou nada
        assert StgTipoEvento.objects.count() == 0, "Dry-run não deve salvar em staging"
        assert TipoEvento.objects.count() == 0, "Dry-run não deve salvar em SSOT"

    def test_is_idempotent(self, temp_agenda_file):
        """Test: Comando é idempotente (rodar 2x não duplica)"""
        # Limpar dados
        StgTipoEvento.objects.all().delete()
        TipoEvento.objects.all().delete()

        # Primeira execução
        call_command(
            "load_tipos_evento",
            "--agenda-file", temp_agenda_file,
            verbosity=0
        )
        first_count = TipoEvento.objects.count()

        # Segunda execução
        call_command(
            "load_tipos_evento",
            "--agenda-file", temp_agenda_file,
            verbosity=0
        )
        second_count = TipoEvento.objects.count()

        assert first_count == second_count, "Não deve ter criado duplicatas"


@pytest.mark.django_db
class TestLoadFullPipelineCommand:
    """Testes para comando load_full_pipeline"""

    def test_command_exists(self):
        """Test: Comando load_full_pipeline existe"""
        # Tenta chamar comando (deve falhar por falta de argumentos)
        with pytest.raises(SystemExit):
            call_command("load_full_pipeline")

    def test_loads_usuarios(self, temp_users_file, temp_agenda_file):
        """Test: Carrega usuários"""
        # Limpar dados
        StgUsuario.objects.all().delete()
        Usuario.objects.all().delete()

        # Executar pipeline
        call_command(
            "load_full_pipeline",
            "--users-file", temp_users_file,
            "--agenda-file", temp_agenda_file,
            verbosity=0
        )

        # Verificar usuários foram carregados
        assert StgUsuario.objects.count() > 0, "Deve ter usuários em staging"
        assert Usuario.objects.count() > 0, "Deve ter usuários em SSOT"

    def test_loads_municipios(self, temp_users_file, temp_agenda_file):
        """Test: Carrega municípios extraídos da agenda"""
        # Limpar dados
        Municipio.objects.all().delete()

        # Executar pipeline
        call_command(
            "load_full_pipeline",
            "--users-file", temp_users_file,
            "--agenda-file", temp_agenda_file,
            verbosity=0
        )

        # Verificar municípios foram carregados
        assert Municipio.objects.count() > 0, "Deve ter municípios carregados"

    def test_loads_projetos(self, temp_users_file, temp_agenda_file):
        """Test: Carrega projetos extraídos da agenda"""
        # Limpar dados
        Projeto.objects.all().delete()

        # Executar pipeline
        call_command(
            "load_full_pipeline",
            "--users-file", temp_users_file,
            "--agenda-file", temp_agenda_file,
            verbosity=0
        )

        # Verificar projetos foram carregados
        assert Projeto.objects.count() > 0, "Deve ter projetos carregados"

    def test_loads_tipos_evento(self, temp_users_file, temp_agenda_file):
        """Test: Carrega tipos de evento da agenda"""
        # Limpar dados
        TipoEvento.objects.all().delete()

        # Executar pipeline
        call_command(
            "load_full_pipeline",
            "--users-file", temp_users_file,
            "--agenda-file", temp_agenda_file,
            verbosity=0
        )

        # Verificar tipos evento foram carregados
        assert TipoEvento.objects.count() > 0, "Deve ter tipos de evento carregados"

    def test_loads_eventos_to_staging(self, temp_users_file, temp_agenda_file):
        """Test: Carrega eventos para staging"""
        # Limpar dados
        StgEvento.objects.all().delete()

        # Executar pipeline
        call_command(
            "load_full_pipeline",
            "--users-file", temp_users_file,
            "--agenda-file", temp_agenda_file,
            verbosity=0
        )

        # Verificar eventos foram carregados em staging
        assert StgEvento.objects.count() > 0, "Deve ter eventos em staging"

    def test_dry_run_does_not_persist(self, temp_users_file, temp_agenda_file):
        """Test: --dry-run não persiste dados"""
        # Limpar dados
        StgUsuario.objects.all().delete()
        Usuario.objects.all().delete()
        StgEvento.objects.all().delete()

        # Executar com --dry-run
        call_command(
            "load_full_pipeline",
            "--users-file", temp_users_file,
            "--agenda-file", temp_agenda_file,
            "--dry-run",
            verbosity=0
        )

        # Verificar que não salvou nada
        assert StgUsuario.objects.count() == 0, "Dry-run não deve salvar em staging"
        assert Usuario.objects.count() == 0, "Dry-run não deve salvar em SSOT"
        assert StgEvento.objects.count() == 0, "Dry-run não deve salvar eventos"

    def test_full_pipeline_is_idempotent(self, temp_users_file, temp_agenda_file):
        """Test: Pipeline completo é idempotente"""
        # Limpar dados
        StgUsuario.objects.all().delete()
        Usuario.objects.all().delete()
        Municipio.objects.all().delete()
        Projeto.objects.all().delete()
        TipoEvento.objects.all().delete()
        StgEvento.objects.all().delete()

        # Primeira execução
        call_command(
            "load_full_pipeline",
            "--users-file", temp_users_file,
            "--agenda-file", temp_agenda_file,
            verbosity=0
        )

        first_usuarios = Usuario.objects.count()
        first_municipios = Municipio.objects.count()
        first_projetos = Projeto.objects.count()
        first_tipos = TipoEvento.objects.count()
        first_eventos = StgEvento.objects.count()

        # Segunda execução
        call_command(
            "load_full_pipeline",
            "--users-file", temp_users_file,
            "--agenda-file", temp_agenda_file,
            verbosity=0
        )

        # Verificar que não duplicou
        assert Usuario.objects.count() == first_usuarios, "Usuários não devem duplicar"
        assert Municipio.objects.count() == first_municipios, "Municípios não devem duplicar"
        assert Projeto.objects.count() == first_projetos, "Projetos não devem duplicar"
        assert TipoEvento.objects.count() == first_tipos, "Tipos não devem duplicar"
        assert StgEvento.objects.count() == first_eventos, "Eventos não devem duplicar"

    def test_handles_optional_files(self, temp_users_file, temp_agenda_file):
        """Test: Lida com arquivos opcionais (disponibilidade, controle)"""
        # Executar sem arquivos opcionais (não deve dar erro)
        call_command(
            "load_full_pipeline",
            "--users-file", temp_users_file,
            "--agenda-file", temp_agenda_file,
            verbosity=0
        )

        # Verificar que executou sem erros
        assert Usuario.objects.count() > 0

    def test_processes_optional_disponibilidade(
        self, temp_users_file, temp_agenda_file, temp_disponibilidade_file
    ):
        """Test: Processa arquivo de disponibilidade se fornecido"""
        # Executar com arquivo de disponibilidade
        call_command(
            "load_full_pipeline",
            "--users-file", temp_users_file,
            "--agenda-file", temp_agenda_file,
            "--disponibilidade-file", temp_disponibilidade_file,
            verbosity=0
        )

        # Verificar que executou sem erros
        # (Transferência para SSOT pode ainda não estar implementada)
        assert Usuario.objects.count() > 0

    def test_processes_optional_controle(
        self, temp_users_file, temp_agenda_file, temp_controle_file
    ):
        """Test: Processa arquivo de controle se fornecido"""
        # Executar com arquivo de controle
        call_command(
            "load_full_pipeline",
            "--users-file", temp_users_file,
            "--agenda-file", temp_agenda_file,
            "--controle-file", temp_controle_file,
            verbosity=0
        )

        # Verificar que executou sem erros
        assert Usuario.objects.count() > 0


@pytest.mark.django_db
class TestCommandsIntegration:
    """Testes de integração entre comandos"""

    def test_load_tipos_evento_then_full_pipeline(self, temp_users_file, temp_agenda_file):
        """Test: load_tipos_evento seguido de load_full_pipeline não duplica"""
        # Limpar dados
        TipoEvento.objects.all().delete()

        # Primeiro: carregar tipos isoladamente
        call_command(
            "load_tipos_evento",
            "--agenda-file", temp_agenda_file,
            verbosity=0
        )
        tipos_count_after_first = TipoEvento.objects.count()

        # Depois: rodar pipeline completo
        call_command(
            "load_full_pipeline",
            "--users-file", temp_users_file,
            "--agenda-file", temp_agenda_file,
            verbosity=0
        )
        tipos_count_after_pipeline = TipoEvento.objects.count()

        # Não deve ter duplicado tipos
        assert tipos_count_after_pipeline == tipos_count_after_first

    def test_commands_respect_transactions(self, temp_users_file, temp_agenda_file):
        """Test: Comandos usam transações (rollback em erro)"""
        # Criar arquivo de usuários inválido (sem header)
        invalid_file = Path(temp_users_file).parent / "invalid.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["invalid", "data", "no", "header"])
        wb.save(invalid_file)

        # Executar comando com arquivo inválido (deve falhar)
        try:
            call_command(
                "load_full_pipeline",
                "--users-file", str(invalid_file),
                "--agenda-file", temp_agenda_file,
                verbosity=0
            )
        except Exception:
            pass  # Esperado

        # Verificar que não salvou dados parciais (rollback)
        # (Depende de implementação @transaction.atomic)

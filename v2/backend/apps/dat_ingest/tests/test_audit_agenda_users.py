"""
Testes para comando audit_agenda_users.

Valida:
- Gera JSON/CSV com chaves esperadas
- Exclui "Convidados" no matching por nome
- Inclui emails "Convidados" no matching por email
"""

import json
import csv
import pytest
from pathlib import Path
from openpyxl import Workbook
from django.core.management import call_command

pytestmark = pytest.mark.django_db


@pytest.fixture
def users_excel(tmp_path):
    """Cria planilha de usuários cadastrados."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ativos"

    # Header
    ws.append(["Nome", "Nome Completo", "Email", "CPF"])

    # Dados
    ws.append(["João", "João Silva", "joao@example.com", "12345678901"])
    ws.append(["Maria", "Maria Santos", "maria@example.com", "98765432109"])
    ws.append(["Pedro", "Pedro Costa", "pedro@example.com", "11111111111"])
    ws.append(["Convidado", "Ana Lima", "ana@external.com", "22222222222"])  # Email de convidado

    path = tmp_path / "usuarios.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def agenda_excel(tmp_path):
    """Cria planilha de agenda com múltiplas abas."""
    wb = Workbook()

    # Aba ACerta
    ws_acerta = wb.active
    ws_acerta.title = "ACerta"
    ws_acerta.append([
        "Data", "Horário", "Município", "Tipo", "Projeto",
        "Coordenador", "Formador 1", "Formador 2", "Formador 3", "Formador 4", "Formador 5",
        "Convidados"
    ])
    ws_acerta.append([
        "2025-01-10", "08:00-12:00", "Fortaleza", "Formação", "ACerta",
        "João Silva", "Maria Santos", "", "", "", "",
        "ana@external.com"  # Email de convidado
    ])

    # Aba Super
    ws_super = wb.create_sheet("Super")
    ws_super.append([
        "Data", "Horário", "Município", "Tipo", "Projeto",
        "Coordenador", "Formador 1", "Formador 2", "Formador 3", "Formador 4", "Formador 5",
        "Convidados"
    ])
    ws_super.append([
        "2025-01-15", "14:00-17:00", "Caucaia", "Reunião", "Super",
        "Pedro Costa", "Desconhecido Silva", "", "", "", "",  # "Desconhecido" não está cadastrado
        ""
    ])

    # Aba Brincando (sem eventos)
    ws_brincando = wb.create_sheet("Brincando")
    ws_brincando.append([
        "Data", "Horário", "Município", "Tipo", "Projeto",
        "Coordenador", "Formador 1", "Formador 2", "Formador 3", "Formador 4", "Formador 5",
        "Convidados"
    ])

    path = tmp_path / "agenda.xlsx"
    wb.save(path)
    return path


def test_gera_json_csv_com_chaves_esperadas(users_excel, agenda_excel, tmp_path):
    """Comando gera JSON/CSV com chaves esperadas."""
    out_dir = tmp_path / "out_etl"
    out_dir.mkdir()

    call_command(
        "audit_agenda_users",
        users=str(users_excel),
        agenda=str(agenda_excel),
        sheet_users="Ativos",
        out=str(out_dir),
    )

    # Verificar JSON básico
    json_path = out_dir / "audit_users_crosscheck_report.json"
    assert json_path.exists()

    with open(json_path, "r", encoding="utf-8") as f:
        report = json.load(f)

    # Chaves obrigatórias
    assert "abas_processadas" in report
    assert "duplicados_por_aba" in report
    assert "encontrados_por_aba" in report
    assert "usuarios_nao_encontrados" in report
    assert "multi_setor" in report
    assert "nomes_na_agenda_nao_cadastrados" in report

    # Verificar JSON enhanced
    json_enhanced_path = out_dir / "audit_users_crosscheck_enhanced.json"
    assert json_enhanced_path.exists()

    with open(json_enhanced_path, "r", encoding="utf-8") as f:
        enhanced = json.load(f)

    assert "totals" in enhanced
    assert "abas" in enhanced["totals"]
    assert "multi_setor_count" in enhanced["totals"]
    assert "nao_cadastrados_unique" in enhanced["totals"]

    # Verificar CSV
    csv_path = out_dir / "audit_users_encontrados_por_aba.csv"
    assert csv_path.exists()

    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        header = next(reader)
        assert header == ["Nome", "Abas"]


def test_exclui_convidados_no_matching_por_nome(users_excel, agenda_excel, tmp_path):
    """Matching por nome NÃO deve incluir coluna "Convidados"."""
    out_dir = tmp_path / "out_etl"
    out_dir.mkdir()

    call_command(
        "audit_agenda_users",
        users=str(users_excel),
        agenda=str(agenda_excel),
        sheet_users="Ativos",
        out=str(out_dir),
    )

    with open(out_dir / "audit_users_crosscheck_report.json", "r", encoding="utf-8") as f:
        report = json.load(f)

    # Verificar que "Desconhecido Silva" aparece como não cadastrado
    # (porque matching por nome não encontra)
    nao_encontrados_super = report["usuarios_nao_encontrados"].get("Super", [])
    assert any("desconhecido" in nome.lower() for nome in nao_encontrados_super)


def test_inclui_emails_convidados_no_matching(users_excel, agenda_excel, tmp_path):
    """Matching por email DEVE incluir coluna "Convidados"."""
    out_dir = tmp_path / "out_etl"
    out_dir.mkdir()

    call_command(
        "audit_agenda_users",
        users=str(users_excel),
        agenda=str(agenda_excel),
        sheet_users="Ativos",
        out=str(out_dir),
    )

    with open(out_dir / "audit_users_crosscheck_report.json", "r", encoding="utf-8") as f:
        report = json.load(f)

    # Verificar que ana@external.com (convidado) foi encontrado via email
    # (não deve aparecer em não encontrados)
    nao_encontrados_acerta = report["usuarios_nao_encontrados"].get("ACerta", [])
    assert not any("ana@external.com" in nome.lower() for nome in nao_encontrados_acerta)


def test_detecta_multi_setor(tmp_path):
    """Detecta usuários que aparecem em múltiplas abas."""
    # Criar planilha de usuários
    wb_users = Workbook()
    ws_users = wb_users.active
    ws_users.title = "Ativos"
    ws_users.append(["Nome", "Nome Completo", "Email", "CPF"])
    ws_users.append(["João", "João Silva", "joao@example.com", "12345678901"])
    users_path = tmp_path / "usuarios.xlsx"
    wb_users.save(users_path)

    # Criar planilha de agenda com João em múltiplas abas
    wb_agenda = Workbook()

    # Aba ACerta
    ws_acerta = wb_agenda.active
    ws_acerta.title = "ACerta"
    ws_acerta.append([
        "Data", "Horário", "Município", "Tipo", "Projeto",
        "Coordenador", "Formador 1", "Formador 2", "Formador 3", "Formador 4", "Formador 5",
        "Convidados"
    ])
    ws_acerta.append([
        "2025-01-10", "08:00-12:00", "Fortaleza", "Formação", "ACerta",
        "João Silva", "", "", "", "", "", ""
    ])

    # Aba Super (João aparece aqui também)
    ws_super = wb_agenda.create_sheet("Super")
    ws_super.append([
        "Data", "Horário", "Município", "Tipo", "Projeto",
        "Coordenador", "Formador 1", "Formador 2", "Formador 3", "Formador 4", "Formador 5",
        "Convidados"
    ])
    ws_super.append([
        "2025-01-15", "14:00-17:00", "Caucaia", "Reunião", "Super",
        "João Silva", "", "", "", "", "", ""
    ])

    agenda_path = tmp_path / "agenda.xlsx"
    wb_agenda.save(agenda_path)

    out_dir = tmp_path / "out_etl"
    out_dir.mkdir()

    call_command(
        "audit_agenda_users",
        users=str(users_path),
        agenda=str(agenda_path),
        sheet_users="Ativos",
        out=str(out_dir),
    )

    with open(out_dir / "audit_users_crosscheck_report.json", "r", encoding="utf-8") as f:
        report = json.load(f)

    # João deve aparecer em multi_setor (ACerta + Super)
    assert len(report["multi_setor"]) >= 1
    joao_norm = "joao silva"
    assert joao_norm in report["multi_setor"]
    assert set(report["multi_setor"][joao_norm]) == {"ACerta", "Super"}


def test_top_20_nao_cadastrados(tmp_path):
    """Relatório inclui top 20 não cadastrados por frequência."""
    # Criar planilha de usuários (vazia)
    wb_users = Workbook()
    ws_users = wb_users.active
    ws_users.title = "Ativos"
    ws_users.append(["Nome", "Nome Completo", "Email", "CPF"])
    users_path = tmp_path / "usuarios.xlsx"
    wb_users.save(users_path)

    # Criar planilha de agenda com vários não cadastrados
    wb_agenda = Workbook()
    ws_acerta = wb_agenda.active
    ws_acerta.title = "ACerta"
    ws_acerta.append([
        "Data", "Horário", "Município", "Tipo", "Projeto",
        "Coordenador", "Formador 1", "Formador 2", "Formador 3", "Formador 4", "Formador 5",
        "Convidados"
    ])

    # Adicionar 10 eventos com "Fulano" (frequência alta)
    for i in range(10):
        ws_acerta.append([
            f"2025-01-{10+i}", "08:00-12:00", "Fortaleza", "Formação", "ACerta",
            "Fulano Silva", "", "", "", "", "", ""
        ])

    # Adicionar 2 eventos com "Ciclano" (frequência baixa)
    for i in range(2):
        ws_acerta.append([
            f"2025-02-{10+i}", "08:00-12:00", "Caucaia", "Reunião", "ACerta",
            "Ciclano Santos", "", "", "", "", "", ""
        ])

    agenda_path = tmp_path / "agenda.xlsx"
    wb_agenda.save(agenda_path)

    out_dir = tmp_path / "out_etl"
    out_dir.mkdir()

    call_command(
        "audit_agenda_users",
        users=str(users_path),
        agenda=str(agenda_path),
        sheet_users="Ativos",
        out=str(out_dir),
    )

    with open(out_dir / "audit_users_crosscheck_report.json", "r", encoding="utf-8") as f:
        report = json.load(f)

    # Verificar top_20
    top_20 = report["nomes_na_agenda_nao_cadastrados"]["top_20"]
    assert len(top_20) >= 2

    # Fulano deve ter frequência maior que Ciclano
    fulano_entry = next((e for e in top_20 if "fulano" in e["nome"].lower()), None)
    ciclano_entry = next((e for e in top_20 if "ciclano" in e["nome"].lower()), None)

    assert fulano_entry is not None
    assert ciclano_entry is not None
    assert fulano_entry["freq"] > ciclano_entry["freq"]
